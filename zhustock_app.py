# -*- coding: utf-8 -*-
"""
ZHU STOCK APP｜客戶展示版 + 自動回退前一日結果 + 我的持股 + 分析中品牌圖
"""

import os
import re
import sys
import time
import json
import random
import pickle
import threading
import traceback
import socket
import platform
import uuid
import requests
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from urllib.parse import quote_plus
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.patches import Rectangle


# =========================
# 使用者設定
# =========================
OUTPUT_DIR = r"C:\Users\user\Desktop\zhustock"
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "TWSE_ALL.xlsx")
ICON_FILE = "zhu_stock_icon.ico"
SNAPSHOT_DIR = os.path.join(OUTPUT_DIR, "snapshots")
TREND_MEMORY_FILE = os.path.join(OUTPUT_DIR, "trend_memory.pkl")
HOLDINGS_FILE = os.path.join(OUTPUT_DIR, "holdings.json")
AUTH_SESSION_FILE = os.path.join(OUTPUT_DIR, "auth_session.json")
DEVICE_INFO_FILE = os.path.join(OUTPUT_DIR, "device_info.json")
AUTH_SERVER_BASE = "https://zhu-stock-app.onrender.com"

PAYMENT_BANKS = [
    "元大銀行 / 806 / 20342720080940",
    "兆豐銀行 / 017 / 03910980975",
]
PAYMENT_PLAN_LABELS = {
    "monthly": "月訂閱（2888 / 月）",
    "halfyear": "半年訂閱（14888 / 6個月）",
    "yearly": "年訂閱（28888 / 年）",
}

PAYMENT_PLAN_PRICES = {
    "monthly": 2888,
    "halfyear": 14888,
    "yearly": 28888,
}


# 品牌圖檔：請把你的紅底「原」字圖放在程式同層或 OUTPUT_DIR
BRAND_IMAGE_CANDIDATES = [
    "yuan_logo.png",
    "brand_yuan.png",
    "原.png",
]

# 免責聲明（可自行修改）
DISCLAIMER_TEXT = """ZHU STOCK APP 免責聲明

使用本程式前，請您務必詳細閱讀以下內容。當您安裝、登入、註冊、啟用、操作或持續使用本程式，即表示您已閱讀、理解並同意接受本免責聲明之全部內容。

一、資訊與功能性質說明
本程式所提供之選股結果、技術分析畫面、條件篩選資訊、趨勢判讀、訊號提示、會員功能、授權功能、資料彙整、圖表顯示、分類結果、提醒通知及其他相關內容，均僅供使用者作為研究、學習、觀察與投資參考之用，不構成任何形式之投資建議、理財建議、招攬、邀約、推薦買賣或保證獲利之承諾。

二、本程式選股僅供參考
本程式所顯示之標的、篩選條件、技術訊號、突破型態、趨勢判讀、強弱排序、均線條件、成交量條件、籌碼條件及其他相關分析結果，僅為依照系統設定條件所產生之客觀輸出結果，僅供參考，並無任何推薦使用者買進、賣出、放空、加碼、減碼、停損、停利或持有之意思表示。

三、無任何保證獲利或保證正確
本程式不保證任何分析結果、選股條件、歷史回測、技術指標、資料來源、計算結果、趨勢判斷、訊號顯示或策略邏輯之正確性、完整性、即時性、穩定性、適用性、可獲利性、無延遲性、無錯誤性。市場具有高度不確定性，任何歷史表現、模型邏輯、技術分析或條件篩選結果，均不代表未來績效，亦不保證未來報酬。

四、無任何推薦買賣之意思
本程式及其開發者、提供者、維護者、管理者，不對任何特定股票、ETF、期貨、選擇權、指數、基金、加密資產或其他金融商品提供買賣建議。使用者不得將本程式所顯示內容，解釋為明示推薦、暗示推薦、保證獲利、代客操作、招攬投資或專屬投顧建議。

五、操作在個人，盈虧自負
所有投資決策、下單行為、資金配置、風險承擔、進出場時機、部位控制與損益結果，均由使用者自行獨立判斷並承擔全部責任。操作在個人，盈虧自負。使用者因使用本程式所產生或衍生之一切獲利、損失、錯失機會、交易糾紛、資產減損、心理壓力、信用風險或其他損害，均與本程式及其開發者、提供者、維護者、管理者無涉。

六、資料來源與延遲風險
本程式所引用、抓取、整理、轉換或顯示之資料，可能來自公開網站、第三方來源、使用者輸入、API 服務、資料匯入檔案或其他外部來源。由於資料來源可能存在延遲、中斷、缺漏、格式異動、網站更新、抓取失敗、欄位錯置、網路異常、伺服器錯誤，故本程式不保證任何資料之即時性與正確性。使用者應自行再次查證官方資訊與市場即時資料後，再進行任何判斷。

七、技術分析與策略限制
本程式所使用之技術分析、均線、KD、MACD、成交量、趨勢線、突破條件、K 棒型態、篩選條件或其他策略邏輯，僅為特定條件設計下之分析工具。不同市場環境、流動性條件、消息面變化、主力行為、政策風險、國際情勢與系統參數設定，皆可能導致分析結果與實際市場走勢不一致。因此，本程式不保證任何策略在任何時點、任何市場、任何標的皆具有一致效果。

八、非投顧、非證券招攬、非代操服務
本程式僅屬資訊整理、條件篩選與技術分析輔助工具，不屬投資顧問服務、證券招攬服務、保證收益服務、代客操作服務或資產管理服務。使用者應自行評估是否符合自身之財務能力、風險承受度、投資經驗、操作紀律與資金配置原則。

九、系統風險與操作風險
本程式在使用過程中，可能因為設備、作業系統、網路、瀏覽器、第三方套件、伺服器、憑證、更新版本、授權狀態、帳號異常、API 錯誤或其他不可控制因素，導致無法登入、畫面顯示異常、分析中斷、選股結果延遲、付款功能延遲、授權判斷異常、匯出失敗、功能暫時停用或其他未預期錯誤。對此，本程式不保證系統可全天候不中斷運作，亦不負擔因使用中斷、資料延誤或操作失敗所造成之任何損害。

十、會員與授權機制說明
本程式之登入、授權、試用、付費、活動贈送、人工開通、手動審核等功能，僅屬使用權限管理機制，不代表任何投資成果承諾。會員資格之開通、延長、贈送或終止，僅影響功能使用權限，不構成對分析結果、投資報酬或獲利能力之保證。

十一、付款與活動贈送聲明
使用者所進行之付款回報、活動贈送、人工開通、免費試用、會員延長等功能，僅涉及系統權限管理與服務期間設定。任何權限開通、延期或贈送，均不代表推薦特定標的、保證使用結果、保證投資獲利或保證後續市場判斷正確。

十二、使用者應自行審慎判斷
使用者在依據本程式進行任何金融商品研究或交易前，應自行再次確認市場資訊、審慎評估風險、設定停損停利、控制部位大小、注意資金安全，並確認是否符合自身投資目的與能力。對於任何依賴本程式資訊而直接或間接作成之投資決策，均由使用者自行負責。

十三、責任限制
在法律允許之最大範圍內，本程式及其開發者、提供者、維護者、管理者，對於使用者因使用或無法使用本程式所產生之一切直接、間接、附帶、衍生、懲罰性或特殊損害，均不負任何賠償責任，包括但不限於投資損失、機會損失、營業損失、資料遺失、交易延誤、系統中斷、帳號異常、授權錯誤、付款延遲以及任何財務或非財務損害。

十四、使用者同意事項
當您開始使用本程式，即表示您已明確理解並同意：
1. 本程式選股僅供參考。
2. 本程式無任何推薦買賣之意。
3. 本程式無保證獲利或保證正確。
4. 所有操作由使用者自行決定。
5. 操作在個人，盈虧自負。
6. 使用者應自行承擔所有投資與交易風險。

十五、條款調整與最終解釋
本免責聲明得因系統更新、功能變更、法規調整或營運需求而隨時修訂。修訂後內容一經公告或於系統顯示，即視為生效。本程式及本免責聲明之最終解釋權，歸本程式經營與管理方所有。"""

PAYMENT_NOTICE_TEXT = """ZHU STOCK APP 訂閱 / 付款前聲明

為避免任何誤解，請您在查看匯款帳號前，務必先完整閱讀以下內容。當您按下「我同意，顯示匯款帳號」時，即表示您已閱讀、理解並同意以下事項：

一、付款性質說明
您本次匯款，係基於您本人於試用、體驗、觀察或實際使用本系統後，認為本工具在資訊整理、技術分析輔助、畫面操作、選股觀察或個人研究上具有便利性與參考價值，因而出於您個人自由意願，主動決定匯款支持並申請延長或開通會員使用權限。

二、非強迫、非推銷、非保證
本程式並未以任何方式強迫付款，亦無保證獲利、保證勝率、保證選股成功、保證投資績效、保證回本、保證翻倍或任何類似承諾。您了解本次付款僅屬於對程式功能使用權限之申請或支持，並非購買任何保證獲利之金融商品或投資顧問成果。

三、付款後之會員權限性質
付款完成後，系統原則上仍需由創作者或管理員進行人工核對後，始會開通、延長或調整會員方案與使用期限。您理解匯款本身不等於系統立即自動生效，實際開通時間仍以管理端核對結果與處理流程為準。

四、使用者主動通知義務
您理解並同意：完成匯款後，應由您本人主動以系統指定方式、聯絡窗口或管理端可辨識之資訊，通知創作者或管理員進行人工核對。若您未主動通知，或提供資料不足、資訊錯誤、付款帳號末五碼無法辨識、金額不符、方案未註明、匯款人資訊不清楚，可能導致審核延後、無法即時比對或暫時無法開通。

五、應提供之基本資訊
為利管理端核對，您通常應主動提供：
1. 您的帳號或註冊 Email
2. 匯款銀行
3. 匯款金額
4. 匯款帳號末五碼
5. 匯款時間
6. 欲申請之方案
7. 其他可供識別之補充資訊

六、付款與投資結果無關
您清楚知悉，本次付款僅涉及 ZHU STOCK APP 的會員使用權限，不代表任何投資建議、推薦買賣、保證正確、保證獲利、保證不賠、保證勝率或未來績效承諾。您不得將付款行為解釋為創作者對您個人提供投顧、代操、保證收益或一對一投資承諾。

七、不可轉嫁為交易責任
您因參考本程式資訊而進行之任何研究、觀察、買賣、加減碼、停損、停利、持有或其他資金操作，均為您本人獨立判斷之結果，盈虧由您自行承擔，不得因已付款或已成為會員，而將任何投資結果、交易損失、錯失機會或心理壓力歸責於創作者、管理員或本系統。

八、匯款錯誤與例外情況
若因您填寫錯誤、匯錯帳號、匯錯金額、重複匯款、未依規則通知、提供錯誤末五碼、使用非本人可辨識帳戶、或其他非管理端可控制因素而導致核對困難，管理端得視情況要求補充資料後再處理。

九、顯示匯款帳號之條件
本頁之匯款帳號資訊，僅提供給已閱讀並同意本聲明之使用者查看。若您不同意本聲明內容，請勿查看匯款帳號、請勿匯款，亦請勿要求開通任何付費權限。

十、最終確認
當您按下「我同意，顯示匯款帳號」時，即表示您確認：
1. 您是基於個人自由意願決定匯款。
2. 您是因實際使用後覺得方便、實用或具有參考價值，才決定支持並申請會員權限。
3. 您了解付款後仍需主動通知創作者或管理員進行人工核對。
4. 您了解付款僅涉及會員權限，不代表任何投資保證。
5. 您了解所有投資行為仍由您自行判斷，盈虧自負。

若您不同意以上內容，請按「不同意，不顯示帳號」。"""

REQUEST_TIMEOUT = 30
REQUEST_RETRIES = 4
SLEEP_SEC = 0.20
MIN_WEEKLY_VOLUME = 10000

LOOKBACK_WEEKS = 60
MIN_PEAK_GAP = 4
MAX_PEAK_GAP = 40
PEAK_WINDOW = 2

BREAKOUT_BUFFER_PCT = 0.003
BOX_BREAKOUT_BUFFER_PCT = 0.003
HOLD_BUFFER_PCT = 0.0
MIN_DESCENT_PCT = 0.01

BREAKDOWN_BUFFER_PCT = 0.003
HOLD_ABOVE_BUFFER_PCT = 0.0
MIN_ASCENT_PCT = 0.01

BOX_LOOKBACK_WEEKS = 10
BOX_MIN_WEEKS = 4

TRAINING_SCORE_THRESHOLD = 55
BEARISH_TRAINING_SCORE_THRESHOLD = 55

USE_STABLE_COMPLETED_DAY = True
MARKET_FINAL_HOUR = 14
MARKET_FINAL_MINUTE = 10

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Connection": "keep-alive",
}

session = requests.Session()
session.headers.update(HEADERS)


# =========================
# 路徑 / 資源
# =========================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def ensure_dir_safe(path):
    if os.path.exists(path):
        if os.path.isdir(path):
            return
        backup_path = f"{path}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        os.rename(path, backup_path)
    os.makedirs(path, exist_ok=True)


def ensure_output_dir():
    ensure_dir_safe(OUTPUT_DIR)
    ensure_dir_safe(SNAPSHOT_DIR)


def get_snapshot_path(settle_date):
    return os.path.join(SNAPSHOT_DIR, f"snapshot_{settle_date}.pkl")


def find_brand_image_path():
    candidates = []
    for name in BRAND_IMAGE_CANDIDATES:
        candidates.append(resource_path(name))
        candidates.append(os.path.join(OUTPUT_DIR, name))
        candidates.append(os.path.abspath(name))

    for p in candidates:
        if os.path.exists(p):
            return p
    return None


# =========================
# 基本工具函式
# =========================
def rebuild_session():
    global session
    try:
        session.close()
    except Exception:
        pass
    session = requests.Session()
    session.headers.update(HEADERS)


def safe_get(url, params=None, timeout=REQUEST_TIMEOUT):
    last_err = None
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            r = session.get(url, params=params, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception as e:
            last_err = e
            time.sleep(0.8 * attempt)
            rebuild_session()
    raise last_err


def clean_text(x):
    if x is None:
        return ""
    s = str(x).replace("\u3000", " ").replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_number(x, is_int=False):
    if x is None:
        return None
    s = str(x).strip().replace(",", "")
    s = s.replace("X", "").replace("除權息", "").replace("+", "").strip()
    if s in {"", "--", "---", "----", "N/A", "NA", "null", "None"}:
        return None
    try:
        if is_int:
            return int(float(s))
        return float(s)
    except Exception:
        return None


def normalize_code(code):
    code = clean_text(code)
    m = re.match(r"^(\d{4,6})", code)
    return m.group(1) if m else code


def is_valid_stock_code(code):
    return bool(re.fullmatch(r"\d{4}", str(code)))


def is_common_stock_name(name):
    if not name:
        return False
    ng_keywords = [
        "特別股", "受益證券", "ETF", "ETN", "權證", "牛證", "熊證", "公司債",
        "債", "基金", "存託憑證", "TDR", "臺灣存託憑證", "認購", "認售",
        "封閉式", "期貨", "反向", "槓桿"
    ]
    for kw in ng_keywords:
        if kw in name:
            return False
    return True


def roc_date_str(dt_obj):
    roc_year = dt_obj.year - 1911
    return f"{roc_year}/{dt_obj.month:02d}/{dt_obj.day:02d}"


def fmt_date_ymd(dt_obj):
    if isinstance(dt_obj, str):
        return dt_obj
    return dt_obj.strftime("%Y-%m-%d")


def get_week_start_monday(any_date):
    return any_date - timedelta(days=any_date.weekday())


def get_writable_output_path(base_path):
    if not os.path.exists(base_path):
        return base_path

    try:
        with open(base_path, "a+b"):
            return base_path
    except PermissionError:
        pass

    folder = os.path.dirname(base_path)
    filename = os.path.basename(base_path)
    name, ext = os.path.splitext(filename)

    for i in range(1, 1000):
        new_path = os.path.join(folder, f"{name}_{i}{ext}")
        if not os.path.exists(new_path):
            return new_path
        try:
            with open(new_path, "a+b"):
                return new_path
        except PermissionError:
            continue

    raise PermissionError("找不到可寫入的輸出檔名，請先關閉 Excel 後再試一次。")


def safe_pct(a, b):
    if b in [0, None] or pd.isna(a) or pd.isna(b):
        return None
    return (a / b - 1.0) * 100.0


def clamp(x, low, high):
    return max(low, min(high, x))


def body_top(open_price, close_price):
    if pd.isna(open_price) or pd.isna(close_price):
        return None
    return max(open_price, close_price)


def body_bottom(open_price, close_price):
    if pd.isna(open_price) or pd.isna(close_price):
        return None
    return min(open_price, close_price)


def is_red_k(open_price, close_price):
    if pd.isna(open_price) or pd.isna(close_price):
        return False
    return close_price > open_price


def is_black_k(open_price, close_price):
    if pd.isna(open_price) or pd.isna(close_price):
        return False
    return close_price < open_price


def is_real_body_breakout(open_price, close_price, line_y, breakout_buffer_pct=0.0):
    if pd.isna(open_price) or pd.isna(close_price) or pd.isna(line_y):
        return False
    if line_y <= 0:
        return False
    if not is_red_k(open_price, close_price):
        return False

    top_ = body_top(open_price, close_price)
    if pd.isna(top_):
        return False

    if close_price <= line_y * (1 + breakout_buffer_pct):
        return False
    if top_ <= line_y:
        return False

    return True


def is_real_body_breakdown(open_price, close_price, line_y, breakdown_buffer_pct=0.0):
    if pd.isna(open_price) or pd.isna(close_price) or pd.isna(line_y):
        return False
    if line_y <= 0:
        return False
    if not is_black_k(open_price, close_price):
        return False

    bottom_ = body_bottom(open_price, close_price)
    if pd.isna(bottom_):
        return False

    if close_price >= line_y * (1 - breakdown_buffer_pct):
        return False
    if bottom_ >= line_y:
        return False

    return True


def get_effective_reference_today():
    now = datetime.now()
    if USE_STABLE_COMPLETED_DAY:
        cutoff = now.replace(hour=MARKET_FINAL_HOUR, minute=MARKET_FINAL_MINUTE, second=0, microsecond=0)
        if now < cutoff:
            return (now - timedelta(days=1)).date()
    return now.date()


# =========================
# 持股清單
# =========================
def load_holdings():
    ensure_output_dir()
    if not os.path.exists(HOLDINGS_FILE):
        return []
    try:
        with open(HOLDINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            out = []
            for x in data:
                code = normalize_code(str(x))
                if is_valid_stock_code(code) and code not in out:
                    out.append(code)
            return out
        return []
    except Exception:
        return []


def save_holdings(codes):
    ensure_output_dir()
    codes = [normalize_code(str(x)) for x in codes if is_valid_stock_code(normalize_code(str(x)))]
    codes = list(dict.fromkeys(codes))
    with open(HOLDINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)


# =========================
# 趨勢線經驗累積學習（目前僅保留後端記憶結構）
# =========================
def get_default_trend_memory():
    return {
        "bullish": {"exact_good": {}, "exact_bad": {}, "pattern_good": {}, "pattern_bad": {}},
        "bearish": {"exact_good": {}, "exact_bad": {}, "pattern_good": {}, "pattern_bad": {}},
    }


def load_trend_memory():
    ensure_output_dir()
    if not os.path.exists(TREND_MEMORY_FILE):
        return get_default_trend_memory()
    try:
        with open(TREND_MEMORY_FILE, "rb") as f:
            obj = pickle.load(f)
        if not isinstance(obj, dict):
            return get_default_trend_memory()
        for side in ["bullish", "bearish"]:
            if side not in obj:
                obj[side] = {"exact_good": {}, "exact_bad": {}, "pattern_good": {}, "pattern_bad": {}}
            for k in ["exact_good", "exact_bad", "pattern_good", "pattern_bad"]:
                if k not in obj[side]:
                    obj[side][k] = {}
        return obj
    except Exception:
        return get_default_trend_memory()


def save_trend_memory(mem):
    ensure_output_dir()
    with open(TREND_MEMORY_FILE, "wb") as f:
        pickle.dump(mem, f)


def round_bucket(x, step=5, default="NA"):
    if x is None or pd.isna(x):
        return default
    try:
        return str(int(round(float(x) / step) * step))
    except Exception:
        return default


def build_trend_memory_keys(direction, code, trend_info):
    if trend_info is None:
        return None, None

    work = trend_info.get("work_df")
    i = trend_info.get("i")
    j = trend_info.get("j")

    if work is None or i is None or j is None:
        return None, None

    first_date = str(work.iloc[i]["週結算日期"])[:10]
    second_date = str(work.iloc[j]["週結算日期"])[:10]
    gap = j - i

    if direction == "bullish":
        exact_key = f"{code}|{first_date}|{second_date}|BULL"
        pattern_key = "|".join([
            "BULL",
            f"GAP={gap}",
            f"DESC={round_bucket(trend_info.get('descent_pct', 0) * 100 if trend_info.get('descent_pct') is not None else None, 2)}",
            f"LD={round_bucket(trend_info.get('line_distance_pct'), 2)}",
            f"BD={round_bucket(trend_info.get('box_distance_pct'), 2)}",
            f"STRICT={1 if trend_info.get('strict_ok', False) else 0}",
            f"HOLD={1 if trend_info.get('training_hold_ok', False) else 0}",
        ])
    else:
        exact_key = f"{code}|{first_date}|{second_date}|BEAR"
        pattern_key = "|".join([
            "BEAR",
            f"GAP={gap}",
            f"ASC={round_bucket(trend_info.get('ascent_pct', 0) * 100 if trend_info.get('ascent_pct') is not None else None, 2)}",
            f"LD={round_bucket(trend_info.get('line_distance_pct'), 2)}",
            f"BREAK={1 if trend_info.get('line_break_now', False) else 0}",
            f"HOLD={1 if trend_info.get('training_hold_ok', False) else 0}",
        ])

    return exact_key, pattern_key


def get_memory_bonus(direction, code, trend_info):
    mem = load_trend_memory()
    side = mem["bullish"] if direction == "bullish" else mem["bearish"]
    exact_key, pattern_key = build_trend_memory_keys(direction, code, trend_info)

    if exact_key is None or pattern_key is None:
        return 0, ""

    eg = int(side["exact_good"].get(exact_key, 0))
    eb = int(side["exact_bad"].get(exact_key, 0))
    pg = int(side["pattern_good"].get(pattern_key, 0))
    pb = int(side["pattern_bad"].get(pattern_key, 0))

    bonus = eg * 35 - eb * 45 + pg * 8 - pb * 10

    parts = []
    if eg:
        parts.append(f"精準正回饋+{eg}")
    if eb:
        parts.append(f"精準負回饋-{eb}")
    if pg:
        parts.append(f"型態正回饋+{pg}")
    if pb:
        parts.append(f"型態負回饋-{pb}")

    return bonus, "、".join(parts)


# =========================
# Snapshot 相容性檢查
# =========================
def is_snapshot_compatible(snapshot_obj):
    if not isinstance(snapshot_obj, dict):
        return False

    required_keys = [
        "TWSE_ALL",
        "WEEK_20MA",
        "STRICT_BREAKOUT",
        "BEARISH_KEY_BREAKDOWN",
        "TRAINING_POOL",
        "BEARISH_TRAINING_POOL",
        "CLIENT_BULLISH",
        "CLIENT_BULLISH_KEYK",
        "CLIENT_BEARISH",
        "CLIENT_BEARISH_KEYK",
        "WEEKLY_MA_RAW",
        "MASTER_STOCK_LIST",
    ]
    for k in required_keys:
        if k not in snapshot_obj:
            return False

    bull = snapshot_obj.get("CLIENT_BULLISH", pd.DataFrame())
    bear = snapshot_obj.get("CLIENT_BEARISH", pd.DataFrame())
    raw = snapshot_obj.get("WEEKLY_MA_RAW", pd.DataFrame())
    master = snapshot_obj.get("MASTER_STOCK_LIST", pd.DataFrame())

    if not isinstance(bull, pd.DataFrame) or not isinstance(bear, pd.DataFrame) or not isinstance(raw, pd.DataFrame) or not isinstance(master, pd.DataFrame):
        return False
    if raw.empty:
        return False

    bull_need = ["星等", "乖離率(%)", "短線停利Alarm", "長線停利Alarm"]
    bear_need = ["星等", "乖離率(%)", "短線回補Alarm", "長線回補Alarm"]

    for c in bull_need:
        if c not in bull.columns:
            return False
    for c in bear_need:
        if c not in bear.columns:
            return False

    return True


def save_snapshot(settle_date, result_dict):
    path = get_snapshot_path(settle_date)
    with open(path, "wb") as f:
        pickle.dump(result_dict, f)


def load_snapshot(settle_date):
    path = get_snapshot_path(settle_date)
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return pickle.load(f)


def find_latest_compatible_snapshot(days_back=10):
    ensure_output_dir()
    today = get_effective_reference_today()
    for i in range(days_back + 1):
        d = today - timedelta(days=i)
        path = get_snapshot_path(str(d))
        if os.path.exists(path):
            try:
                snap = load_snapshot(str(d))
                if is_snapshot_compatible(snap):
                    snap["settle_date"] = str(d)
                    return snap
            except Exception:
                pass
    return None


# =========================
# 抓上市 / 上櫃一般股票清單 + 產業別
# =========================
def fetch_isin_list(str_mode):
    url = "https://isin.twse.com.tw/isin/C_public.jsp"
    r = safe_get(url, params={"strMode": str_mode})
    r.encoding = "big5-hkscs" if "big5" in r.headers.get("Content-Type", "").lower() else r.apparent_encoding

    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table", class_="h4")
    if table is None:
        raise RuntimeError(f"ISIN 頁面解析失敗，strMode={str_mode}")

    rows = table.find_all("tr")
    data = []
    current_section = ""

    for tr in rows:
        cols = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all(["td", "th"])]
        if not cols:
            continue

        if len(cols) == 1:
            current_section = cols[0]
            continue

        first_col = cols[0]
        m = re.match(r"^(\d{4,6})\s+(.+)$", first_col)
        if not m:
            continue

        code = normalize_code(m.group(1))
        name = clean_text(m.group(2))

        if not is_valid_stock_code(code):
            continue
        if "股票" not in current_section:
            continue
        if not is_common_stock_name(name):
            continue

        industry = cols[4] if len(cols) > 4 else ""
        if not industry:
            industry = "未分類"

        data.append({
            "股票代號": code,
            "股票名稱": name,
            "產業別": industry,
        })

    return pd.DataFrame(data).drop_duplicates(subset=["股票代號"]).reset_index(drop=True)


def get_master_stock_list():
    listed = fetch_isin_list(2).copy()
    listed["市場別"] = "上市"

    otc = fetch_isin_list(4).copy()
    otc["市場別"] = "上櫃"

    master = pd.concat([listed, otc], ignore_index=True)
    master = master.drop_duplicates(subset=["股票代號"]).sort_values(["市場別", "股票代號"]).reset_index(drop=True)
    master["產業別"] = master["產業別"].fillna("").astype(str).str.strip()
    master.loc[master["產業別"] == "", "產業別"] = "未分類"
    return master


# =========================
# TWSE / TPEx 日資料
# =========================
def extract_twse_table(tables):
    if not isinstance(tables, list):
        return None

    for tbl in tables:
        fields = [clean_text(f) for f in tbl.get("fields", [])]
        joined = "|".join(fields)
        need = ["證券代號", "證券名稱", "成交股數", "開盤價", "最高價", "最低價", "收盤價"]
        if all(x in joined for x in need):
            return tbl
    return None


def fetch_twse_daily_all(date_obj):
    url = "https://www.twse.com.tw/exchangeReport/MI_INDEX"
    params = {
        "response": "json",
        "date": date_obj.strftime("%Y%m%d"),
        "type": "ALLBUT0999",
    }

    r = safe_get(url, params=params)
    j = r.json()

    stat = clean_text(j.get("stat", ""))
    if "沒有符合條件的資料" in stat or "很抱歉" in stat or "查詢日期大於今日" in stat:
        return None

    tbl = extract_twse_table(j.get("tables", []))
    if tbl is None:
        return None

    fields = [clean_text(x) for x in tbl["fields"]]
    data = tbl["data"]
    df = pd.DataFrame(data, columns=fields)

    keep_cols = ["證券代號", "證券名稱", "成交股數", "開盤價", "最高價", "最低價", "收盤價"]
    if not all(c in df.columns for c in keep_cols):
        return None

    out = df[keep_cols].copy()
    out["股票代號"] = out["證券代號"].map(normalize_code)
    out["股票名稱"] = out["證券名稱"].map(clean_text)
    out["成交股數"] = out["成交股數"].map(lambda x: parse_number(x, is_int=True))
    out["開盤價"] = out["開盤價"].map(parse_number)
    out["最高價"] = out["最高價"].map(parse_number)
    out["最低價"] = out["最低價"].map(parse_number)
    out["收盤價"] = out["收盤價"].map(parse_number)
    out["日期"] = fmt_date_ymd(date_obj)

    out = out[out["股票代號"].map(is_valid_stock_code)].copy()
    out = out[[
        "日期", "股票代號", "股票名稱",
        "開盤價", "最高價", "最低價", "收盤價", "成交股數"
    ]].reset_index(drop=True)
    return out


def fetch_tpex_daily_all(date_obj):
    url = "https://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php"
    params = {
        "l": "zh-tw",
        "d": roc_date_str(date_obj),
        "o": "json",
        "s": "0,asc,0",
    }

    r = safe_get(url, params=params)
    j = r.json()

    aaData = j.get("aaData", [])
    if not aaData:
        return None

    parsed = []
    for row in aaData:
        if len(row) < 8:
            continue

        code = normalize_code(row[0])
        name = clean_text(row[1])
        if not is_valid_stock_code(code):
            continue

        close_price = parse_number(row[2]) if len(row) > 2 else None
        open_price = parse_number(row[4]) if len(row) > 4 else None
        high_price = parse_number(row[5]) if len(row) > 5 else None
        low_price = parse_number(row[6]) if len(row) > 6 else None
        volume_shares = parse_number(row[7], is_int=True) if len(row) > 7 else None

        parsed.append({
            "日期": fmt_date_ymd(date_obj),
            "股票代號": code,
            "股票名稱": name,
            "開盤價": open_price,
            "最高價": high_price,
            "最低價": low_price,
            "收盤價": close_price,
            "成交股數": volume_shares,
        })

    if not parsed:
        return None

    return pd.DataFrame(parsed)


# =========================
# 最近可用交易日
# =========================
def get_latest_available_trading_date(max_lookback_days=20):
    """
    取得最近可用交易日（穩定修正版）
    邏輯：
    1. 先用原本 TWSE / TPEX 抓法回推搜尋
    2. 若兩邊都因網站異常、連假、暫時阻擋而失敗，改用保守 fallback
    3. fallback 規則：回推到最近的平日（週一～週五），避免整體分析直接中止
    """
    base_today = get_effective_reference_today()
    last_exception_msgs = []

    for i in range(max_lookback_days + 1):
        d = base_today - timedelta(days=i)
        has_data = False

        try:
            df1 = fetch_twse_daily_all(d)
            if df1 is not None and not df1.empty:
                has_data = True
        except Exception as e:
            last_exception_msgs.append(f"TWSE {fmt_date_ymd(d)}: {e}")

        try:
            df2 = fetch_tpex_daily_all(d)
            if df2 is not None and not df2.empty:
                has_data = True
        except Exception as e:
            last_exception_msgs.append(f"TPEX {fmt_date_ymd(d)}: {e}")

        if has_data:
            return d

        time.sleep(0.1)

    # fallback：最近平日
    for i in range(max_lookback_days + 1):
        d = base_today - timedelta(days=i)
        if d.weekday() < 5:
            return d

    detail = "\n".join(last_exception_msgs[-10:]) if last_exception_msgs else "無額外錯誤訊息"
    raise RuntimeError(f"找不到最近可用交易日，請稍後再試。\n{detail}")


# =========================
# 最近週資料
# =========================
def fetch_market_week_data(market, stock_codes, target_settle_date, logger=None):
    monday = get_week_start_monday(target_settle_date)
    dates = []
    cur = monday
    while cur <= target_settle_date:
        dates.append(cur)
        cur += timedelta(days=1)

    daily_frames = []
    hard_failures = []

    for d in dates:
        try:
            if market == "上市":
                df = fetch_twse_daily_all(d)
            else:
                df = fetch_tpex_daily_all(d)

            if df is not None and not df.empty:
                df = df[df["股票代號"].isin(stock_codes)].copy()
                if not df.empty:
                    daily_frames.append(df)
        except Exception as e:
            hard_failures.append((fmt_date_ymd(d), str(e)))
            if logger:
                logger(f"[警告] {market} {fmt_date_ymd(d)} 抓取失敗：{e}")

        time.sleep(SLEEP_SEC)

    if hard_failures:
        raise RuntimeError(f"{market} 本週資料抓取失敗，為避免結果飄動，本次中止。失敗日期數={len(hard_failures)}")

    if not daily_frames:
        return {}

    all_daily = pd.concat(daily_frames, ignore_index=True)
    all_daily["日期_dt"] = pd.to_datetime(all_daily["日期"])

    result = {}
    for code, grp in all_daily.groupby("股票代號"):
        grp = grp.sort_values("日期_dt")
        last_row = grp.iloc[-1]

        weekly_volume_shares = grp["成交股數"].fillna(0).sum()
        weekly_volume_lots = int(round(weekly_volume_shares / 1000)) if pd.notna(weekly_volume_shares) else None

        week_settle_date = last_row["日期"]
        close_price = last_row["收盤價"]

        missing_reason = ""
        if week_settle_date != fmt_date_ymd(target_settle_date):
            missing_reason = "該市場當日無資料，改採本週最後可用交易日"

        result[code] = {
            "週結算日期": week_settle_date,
            "每週五收盤後股價": close_price,
            "每週總成交量(張)": weekly_volume_lots,
            "缺漏原因": missing_reason,
        }

    return result


# =========================
# 歷史日資料 -> 週K
# =========================
def fetch_market_daily_history(market, stock_codes, start_date, end_date, logger=None):
    all_frames = []
    cur = start_date
    total_days = (end_date - start_date).days + 1
    day_count = 0
    hard_failures = []

    while cur <= end_date:
        day_count += 1
        try:
            if market == "上市":
                df = fetch_twse_daily_all(cur)
            else:
                df = fetch_tpex_daily_all(cur)

            if df is not None and not df.empty:
                df = df[df["股票代號"].isin(stock_codes)].copy()
                if not df.empty:
                    df["市場別"] = market
                    all_frames.append(df)
                    if logger:
                        logger(f"[{market}] {fmt_date_ymd(cur)} 抓取成功，筆數={len(df)} ({day_count}/{total_days})")
            else:
                if logger:
                    logger(f"[{market}] {fmt_date_ymd(cur)} 無資料 ({day_count}/{total_days})")
        except Exception as e:
            hard_failures.append((fmt_date_ymd(cur), str(e)))
            if logger:
                logger(f"[警告] {market} {fmt_date_ymd(cur)} 抓取失敗：{e}")

        time.sleep(SLEEP_SEC)
        cur += timedelta(days=1)

    if hard_failures:
        sample = "；".join([f"{d}" for d, _ in hard_failures[:5]])
        raise RuntimeError(f"{market} 歷史資料抓取失敗 {len(hard_failures)} 天，為避免分析不一致，本次中止。失敗日期示例：{sample}")

    if not all_frames:
        return pd.DataFrame(columns=[
            "日期", "股票代號", "股票名稱",
            "開盤價", "最高價", "最低價", "收盤價", "成交股數", "市場別"
        ])

    out = pd.concat(all_frames, ignore_index=True)
    out["日期"] = pd.to_datetime(out["日期"])
    return out


def build_weekly_k_from_daily(daily_df):
    if daily_df.empty:
        return pd.DataFrame(columns=[
            "股票代號", "股票名稱", "市場別", "週別", "週結算日期",
            "週開盤價", "週最高價", "週最低價", "週收盤價", "週成交量(張)"
        ])

    df = daily_df.copy()
    df["日期"] = pd.to_datetime(df["日期"])
    df = df.sort_values(["股票代號", "日期"]).reset_index(drop=True)
    df["週別"] = df["日期"].dt.to_period("W-FRI")

    rows = []
    for (code, week_period), grp in df.groupby(["股票代號", "週別"]):
        grp = grp.sort_values("日期")
        first_row = grp.iloc[0]
        last_row = grp.iloc[-1]

        rows.append({
            "股票代號": code,
            "股票名稱": last_row["股票名稱"],
            "市場別": last_row["市場別"],
            "週別": str(week_period),
            "週結算日期": last_row["日期"].strftime("%Y-%m-%d"),
            "週開盤價": first_row["開盤價"],
            "週最高價": grp["最高價"].max(),
            "週最低價": grp["最低價"].min(),
            "週收盤價": last_row["收盤價"],
            "週成交量(張)": int(round(grp["成交股數"].fillna(0).sum() / 1000)),
        })

    return pd.DataFrame(rows).sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)


def calculate_weekly_indicators(weekly_df):
    if weekly_df.empty:
        return pd.DataFrame()

    df = weekly_df.copy()
    df["週結算日期"] = pd.to_datetime(df["週結算日期"])
    df = df.sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)

    g = df.groupby("股票代號")

    df["週3MA"] = g["週收盤價"].transform(lambda s: s.rolling(3, min_periods=3).mean())
    df["週5MA"] = g["週收盤價"].transform(lambda s: s.rolling(5, min_periods=5).mean())
    df["週10MA"] = g["週收盤價"].transform(lambda s: s.rolling(10, min_periods=10).mean())
    df["週20MA"] = g["週收盤價"].transform(lambda s: s.rolling(20, min_periods=20).mean())

    df["量5MA"] = g["週成交量(張)"].transform(lambda s: s.rolling(5, min_periods=5).mean())
    df["量10MA"] = g["週成交量(張)"].transform(lambda s: s.rolling(10, min_periods=10).mean())
    df["量20MA"] = g["週成交量(張)"].transform(lambda s: s.rolling(20, min_periods=20).mean())

    df["20MA斜率"] = g["週20MA"].transform(lambda s: s.diff())
    df["5MA斜率"] = g["週5MA"].transform(lambda s: s.diff())

    df["近13週最高"] = g["週最高價"].transform(lambda s: s.rolling(13, min_periods=13).max())
    df["近26週最高"] = g["週最高價"].transform(lambda s: s.rolling(26, min_periods=26).max())
    df["近52週最高"] = g["週最高價"].transform(lambda s: s.rolling(52, min_periods=20).max())

    df["近13週最低"] = g["週最低價"].transform(lambda s: s.rolling(13, min_periods=13).min())
    df["近26週最低"] = g["週最低價"].transform(lambda s: s.rolling(26, min_periods=26).min())
    df["近52週最低"] = g["週最低價"].transform(lambda s: s.rolling(52, min_periods=20).min())

    df["是否站上週20MA"] = df["週收盤價"] >= df["週20MA"]
    df["是否跌破週20MA"] = df["週收盤價"] < df["週20MA"]

    df["均線多頭排列"] = (
        (df["週3MA"] >= df["週5MA"]) &
        (df["週5MA"] >= df["週10MA"]) &
        (df["週10MA"] >= df["週20MA"])
    )

    df["均線空頭排列"] = (
        (df["週3MA"] <= df["週5MA"]) &
        (df["週5MA"] <= df["週10MA"]) &
        (df["週10MA"] <= df["週20MA"])
    )

    return df


def get_latest_week_20ma_candidates(weekly_ma_df, master_df, min_weekly_volume=10000):
    empty_cols = [
        "項次", "股票代號", "股票名稱", "市場別", "產業別", "週結算日期",
        "最新週收盤價", "週20MA", "最新一週成交量(張)", "是否站上週20MA"
    ]

    if weekly_ma_df.empty:
        return pd.DataFrame(columns=empty_cols)

    df = weekly_ma_df.copy()
    df = df.dropna(subset=["週20MA"]).copy()
    if df.empty:
        return pd.DataFrame(columns=empty_cols)

    df["週結算日期"] = pd.to_datetime(df["週結算日期"])

    latest = (
        df.sort_values(["股票代號", "週結算日期"])
        .groupby("股票代號", as_index=False)
        .tail(1)
        .copy()
    )

    latest = latest[
        (latest["是否站上週20MA"] == True) &
        (latest["週成交量(張)"] >= min_weekly_volume)
    ].copy()

    industry_map = master_df[["股票代號", "產業別"]].drop_duplicates(subset=["股票代號"]).copy()
    latest = latest.merge(industry_map, on="股票代號", how="left")
    latest["產業別"] = latest["產業別"].fillna("未分類")
    latest = latest.sort_values(["市場別", "股票代號"]).reset_index(drop=True)

    return pd.DataFrame({
        "項次": range(1, len(latest) + 1),
        "股票代號": latest["股票代號"].astype(str),
        "股票名稱": latest["股票名稱"].astype(str),
        "市場別": latest["市場別"].astype(str),
        "產業別": latest["產業別"].astype(str),
        "週結算日期": latest["週結算日期"].dt.strftime("%Y-%m-%d"),
        "最新週收盤價": latest["週收盤價"],
        "週20MA": latest["週20MA"].round(4),
        "最新一週成交量(張)": latest["週成交量(張)"],
        "是否站上週20MA": latest["是否站上週20MA"].map(lambda x: "是" if x else "否"),
    })


# =========================
# 趨勢線工具
# =========================
def line_value(x1, y1, x2, y2, x):
    if x2 == x1:
        return y1
    return y1 + (y2 - y1) * ((x - x1) / (x2 - x1))


def find_pivot_highs_above_20ma(sub_df, window=2):
    pivots = []
    n = len(sub_df)
    for i in range(window, n - window):
        cur_high = sub_df.iloc[i]["週最高價"]
        cur_ma20 = sub_df.iloc[i]["週20MA"]

        if pd.isna(cur_high) or pd.isna(cur_ma20):
            continue
        if cur_high <= cur_ma20:
            continue

        left = sub_df.iloc[i - window:i]["週最高價"].tolist()
        right = sub_df.iloc[i + 1:i + 1 + window]["週最高價"].tolist()

        if any(pd.isna(x) for x in left) or any(pd.isna(x) for x in right):
            continue

        if all(cur_high >= x for x in left) and all(cur_high >= x for x in right):
            pivots.append(i)

    return pivots


def find_pivot_lows_below_20ma(sub_df, window=2):
    pivots = []
    n = len(sub_df)
    for i in range(window, n - window):
        cur_low = sub_df.iloc[i]["週最低價"]
        cur_ma20 = sub_df.iloc[i]["週20MA"]

        if pd.isna(cur_low) or pd.isna(cur_ma20):
            continue
        if cur_low >= cur_ma20:
            continue

        left = sub_df.iloc[i - window:i]["週最低價"].tolist()
        right = sub_df.iloc[i + 1:i + 1 + window]["週最低價"].tolist()

        if any(pd.isna(x) for x in left) or any(pd.isna(x) for x in right):
            continue

        if all(cur_low <= x for x in left) and all(cur_low <= x for x in right):
            pivots.append(i)

    return pivots


def get_recent_box_info(work_df, latest_idx, lookback_weeks=10):
    if latest_idx < 2:
        return None

    start_idx = max(0, latest_idx - lookback_weeks)
    end_idx = latest_idx - 1
    box_df = work_df.iloc[start_idx:end_idx + 1].copy()

    if len(box_df) < BOX_MIN_WEEKS:
        return None

    box_high = box_df["週最高價"].max()
    box_low = box_df["週最低價"].min()

    if pd.isna(box_high) or pd.isna(box_low):
        return None
    if box_high <= 0 or box_low <= 0:
        return None

    return {
        "box_start_idx": start_idx,
        "box_end_idx": end_idx,
        "box_high": float(box_high),
        "box_low": float(box_low),
    }


def no_close_back_below_trendline_after_breakout(work, i, j, breakout_idx, buffer_pct=0.0):
    y1 = work.iloc[i]["週最高價"]
    y2 = work.iloc[j]["週最高價"]
    if pd.isna(y1) or pd.isna(y2):
        return False

    for k in range(breakout_idx + 1, len(work)):
        line_y = line_value(i, y1, j, y2, k)
        close_k = work.iloc[k]["週收盤價"]
        if pd.isna(line_y) or pd.isna(close_k):
            return False
        if close_k < line_y * (1 - buffer_pct):
            return False
    return True


def no_close_back_above_trendline_after_breakdown(work, i, j, breakdown_idx, buffer_pct=0.0):
    y1 = work.iloc[i]["週最低價"]
    y2 = work.iloc[j]["週最低價"]
    if pd.isna(y1) or pd.isna(y2):
        return False

    for k in range(breakdown_idx + 1, len(work)):
        line_y = line_value(i, y1, j, y2, k)
        close_k = work.iloc[k]["週收盤價"]
        if pd.isna(line_y) or pd.isna(close_k):
            return False
        if close_k > line_y * (1 + buffer_pct):
            return False
    return True


def find_first_valid_breakout_and_hold(work, i, j, breakout_buffer_pct=0.003, hold_buffer_pct=0.0):
    y1 = work.iloc[i]["週最高價"]
    y2 = work.iloc[j]["週最高價"]
    if pd.isna(y1) or pd.isna(y2):
        return None

    for k in range(j + 1, len(work)):
        line_y = line_value(i, y1, j, y2, k)
        open_k = work.iloc[k]["週開盤價"]
        close_k = work.iloc[k]["週收盤價"]

        if pd.isna(line_y) or pd.isna(open_k) or pd.isna(close_k):
            continue

        if is_real_body_breakout(open_k, close_k, line_y, breakout_buffer_pct=breakout_buffer_pct):
            ok_hold = no_close_back_below_trendline_after_breakout(work, i, j, k, buffer_pct=hold_buffer_pct)
            if ok_hold:
                return {
                    "breakout_idx": k,
                    "breakout_date": work.iloc[k]["週結算日期"],
                    "breakout_open": open_k,
                    "breakout_close": close_k,
                    "breakout_line": line_y,
                }
    return None


def find_first_valid_breakdown_and_hold(work, i, j, breakdown_buffer_pct=0.003, hold_buffer_pct=0.0):
    y1 = work.iloc[i]["週最低價"]
    y2 = work.iloc[j]["週最低價"]
    if pd.isna(y1) or pd.isna(y2):
        return None

    for k in range(j + 1, len(work)):
        line_y = line_value(i, y1, j, y2, k)
        open_k = work.iloc[k]["週開盤價"]
        close_k = work.iloc[k]["週收盤價"]

        if pd.isna(line_y) or pd.isna(open_k) or pd.isna(close_k):
            continue

        if is_real_body_breakdown(open_k, close_k, line_y, breakdown_buffer_pct=breakdown_buffer_pct):
            ok_hold = no_close_back_above_trendline_after_breakdown(work, i, j, k, buffer_pct=hold_buffer_pct)
            if ok_hold:
                return {
                    "breakdown_idx": k,
                    "breakdown_date": work.iloc[k]["週結算日期"],
                    "breakdown_open": open_k,
                    "breakdown_close": close_k,
                    "breakdown_line": line_y,
                }
    return None


def analyze_best_descending_trendline(sub_df):
    if len(sub_df) < 24:
        return None

    code = str(sub_df.iloc[-1]["股票代號"])
    work = sub_df.tail(LOOKBACK_WEEKS).copy().reset_index(drop=True)
    pivots = find_pivot_highs_above_20ma(work, window=PEAK_WINDOW)
    if len(pivots) < 2:
        return None

    candidates = []
    latest_idx = len(work) - 1
    prev_idx = latest_idx - 1
    if prev_idx < 0:
        return None

    latest_open = work.iloc[latest_idx]["週開盤價"]
    latest_close = work.iloc[latest_idx]["週收盤價"]
    prev_close = work.iloc[prev_idx]["週收盤價"]

    for a in range(len(pivots) - 1):
        for b in range(a + 1, len(pivots)):
            i = pivots[a]
            j = pivots[b]
            gap = j - i

            if gap < MIN_PEAK_GAP or gap > MAX_PEAK_GAP:
                continue

            y1 = work.iloc[i]["週最高價"]
            y2 = work.iloc[j]["週最高價"]
            ma1 = work.iloc[i]["週20MA"]
            ma2 = work.iloc[j]["週20MA"]

            if any(pd.isna(x) for x in [y1, y2, ma1, ma2, latest_open, latest_close, prev_close]):
                continue
            if y1 <= ma1 or y2 <= ma2:
                continue
            if y2 >= y1:
                continue

            descent_pct = (y1 - y2) / y1 if y1 not in [0, None] else 0
            if descent_pct < MIN_DESCENT_PCT:
                continue

            latest_line = line_value(i, y1, j, y2, latest_idx)
            prev_line = line_value(i, y1, j, y2, prev_idx)
            if pd.isna(latest_line) or pd.isna(prev_line) or latest_line <= 0 or prev_line <= 0:
                continue

            line_distance_pct = safe_pct(latest_close, latest_line)
            prev_line_distance_pct = safe_pct(prev_close, prev_line)

            line_break_now = is_real_body_breakout(latest_open, latest_close, latest_line, breakout_buffer_pct=BREAKOUT_BUFFER_PCT)
            line_break_prev = (prev_close > prev_line * (1 + BREAKOUT_BUFFER_PCT))

            box_info = get_recent_box_info(work, latest_idx, BOX_LOOKBACK_WEEKS)
            if box_info is None:
                continue

            box_high = box_info["box_high"]
            box_low = box_info["box_low"]
            box_distance_pct = safe_pct(latest_close, box_high)
            prev_box_distance_pct = safe_pct(prev_close, box_high)

            box_break_now = (is_red_k(latest_open, latest_close) and latest_close > box_high * (1 + BOX_BREAKOUT_BUFFER_PCT))
            box_break_prev = (prev_close > box_high * (1 + BOX_BREAKOUT_BUFFER_PCT))

            strict_ok = (
                line_break_now and
                (not line_break_prev) and
                box_break_now and
                (not box_break_prev)
            )

            hold_info = find_first_valid_breakout_and_hold(
                work=work,
                i=i,
                j=j,
                breakout_buffer_pct=BREAKOUT_BUFFER_PCT,
                hold_buffer_pct=HOLD_BUFFER_PCT
            )

            training_hold_ok = hold_info is not None

            base_score = (
                j * 1000
                + clamp((line_distance_pct if line_distance_pct is not None else -999), -20, 20) * 60
                + clamp((box_distance_pct if box_distance_pct is not None else -999), -20, 20) * 70
                + descent_pct * 1000
                + (250 if strict_ok else 0)
                + (180 if training_hold_ok else 0)
            )

            candidate = {
                "i": i,
                "j": j,
                "y1": y1,
                "y2": y2,
                "latest_line": latest_line,
                "prev_line": prev_line,
                "line_distance_pct": line_distance_pct,
                "prev_line_distance_pct": prev_line_distance_pct,
                "line_break_now": line_break_now,
                "line_break_prev": line_break_prev,
                "box_high": box_high,
                "box_low": box_low,
                "box_start_idx": box_info["box_start_idx"],
                "box_end_idx": box_info["box_end_idx"],
                "box_distance_pct": box_distance_pct,
                "prev_box_distance_pct": prev_box_distance_pct,
                "box_break_now": box_break_now,
                "box_break_prev": box_break_prev,
                "descent_pct": descent_pct,
                "strict_ok": strict_ok,
                "training_hold_ok": training_hold_ok,
                "hold_info": hold_info,
                "work_df": work,
            }

            memory_bonus, memory_text = get_memory_bonus("bullish", code, candidate)
            candidate["memory_bonus"] = memory_bonus
            candidate["memory_text"] = memory_text
            candidate["score"] = base_score + memory_bonus
            candidates.append(candidate)

    if not candidates:
        return None

    candidates = sorted(
        candidates,
        key=lambda x: (
            x["score"],
            x["j"],
            x["descent_pct"],
            x["line_distance_pct"] if x["line_distance_pct"] is not None else -999
        ),
        reverse=True
    )
    return candidates[0]


def analyze_best_ascending_trendline(sub_df):
    if len(sub_df) < 24:
        return None

    code = str(sub_df.iloc[-1]["股票代號"])
    work = sub_df.tail(LOOKBACK_WEEKS).copy().reset_index(drop=True)
    pivots = find_pivot_lows_below_20ma(work, window=PEAK_WINDOW)
    if len(pivots) < 2:
        return None

    candidates = []
    latest_idx = len(work) - 1
    prev_idx = latest_idx - 1
    if prev_idx < 0:
        return None

    latest_open = work.iloc[latest_idx]["週開盤價"]
    latest_close = work.iloc[latest_idx]["週收盤價"]
    prev_close = work.iloc[prev_idx]["週收盤價"]

    for a in range(len(pivots) - 1):
        for b in range(a + 1, len(pivots)):
            i = pivots[a]
            j = pivots[b]
            gap = j - i

            if gap < MIN_PEAK_GAP or gap > MAX_PEAK_GAP:
                continue

            y1 = work.iloc[i]["週最低價"]
            y2 = work.iloc[j]["週最低價"]
            ma1 = work.iloc[i]["週20MA"]
            ma2 = work.iloc[j]["週20MA"]

            if any(pd.isna(x) for x in [y1, y2, ma1, ma2, latest_open, latest_close, prev_close]):
                continue
            if y1 >= ma1 or y2 >= ma2:
                continue
            if y2 <= y1:
                continue

            ascent_pct = (y2 - y1) / y1 if y1 not in [0, None] else 0
            if ascent_pct < MIN_ASCENT_PCT:
                continue

            latest_line = line_value(i, y1, j, y2, latest_idx)
            prev_line = line_value(i, y1, j, y2, prev_idx)
            if pd.isna(latest_line) or pd.isna(prev_line) or latest_line <= 0 or prev_line <= 0:
                continue

            line_distance_pct = safe_pct(latest_close, latest_line)
            prev_line_distance_pct = safe_pct(prev_close, prev_line)

            line_break_now = is_real_body_breakdown(latest_open, latest_close, latest_line, breakdown_buffer_pct=BREAKDOWN_BUFFER_PCT)
            line_break_prev = (prev_close < prev_line * (1 - BREAKDOWN_BUFFER_PCT))

            hold_info = find_first_valid_breakdown_and_hold(
                work=work,
                i=i,
                j=j,
                breakdown_buffer_pct=BREAKDOWN_BUFFER_PCT,
                hold_buffer_pct=HOLD_ABOVE_BUFFER_PCT
            )

            training_hold_ok = hold_info is not None

            base_score = (
                j * 1000
                + clamp(((-line_distance_pct) if line_distance_pct is not None else -999), -20, 20) * 70
                + ascent_pct * 1000
                + (220 if line_break_now and (not line_break_prev) else 0)
                + (180 if training_hold_ok else 0)
            )

            candidate = {
                "i": i,
                "j": j,
                "y1": y1,
                "y2": y2,
                "latest_line": latest_line,
                "prev_line": prev_line,
                "line_distance_pct": line_distance_pct,
                "prev_line_distance_pct": prev_line_distance_pct,
                "line_break_now": line_break_now,
                "line_break_prev": line_break_prev,
                "ascent_pct": ascent_pct,
                "training_hold_ok": training_hold_ok,
                "hold_info": hold_info,
                "work_df": work,
            }

            memory_bonus, memory_text = get_memory_bonus("bearish", code, candidate)
            candidate["memory_bonus"] = memory_bonus
            candidate["memory_text"] = memory_text
            candidate["score"] = base_score + memory_bonus
            candidates.append(candidate)

    if not candidates:
        return None

    candidates = sorted(
        candidates,
        key=lambda x: (
            x["score"],
            x["j"],
            x["ascent_pct"],
            -(x["line_distance_pct"] if x["line_distance_pct"] is not None else 999)
        ),
        reverse=True
    )
    return candidates[0]


# =========================
# 分數計算
# =========================
def calc_training_score(grp, trend_info):
    latest = grp.iloc[-1]
    score = 0
    tags = []

    close_ = latest["週收盤價"]
    ma3 = latest["週3MA"]
    ma5 = latest["週5MA"]
    ma10 = latest["週10MA"]
    ma20 = latest["週20MA"]
    vol = latest["週成交量(張)"]
    vol5 = latest["量5MA"]
    vol20 = latest["量20MA"]
    high13 = latest["近13週最高"]
    high26 = latest["近26週最高"]
    high52 = latest["近52週最高"]
    slope20 = latest["20MA斜率"]
    open_ = latest["週開盤價"]

    if pd.notna(ma20) and close_ >= ma20:
        score += 18
        tags.append("站上20MA")
    elif pd.notna(ma20) and close_ >= ma20 * 0.97:
        score += 8
        tags.append("接近20MA")

    if pd.notna(ma3) and pd.notna(ma5) and pd.notna(ma10) and pd.notna(ma20):
        if ma3 >= ma5 >= ma10 >= ma20:
            score += 18
            tags.append("均線多頭")
        elif ma3 >= ma5 and ma5 >= ma20:
            score += 10
            tags.append("均線轉強")

    if pd.notna(slope20):
        if slope20 > 0:
            score += 8
            tags.append("20MA上彎")
        elif slope20 > -0.05:
            score += 3
            tags.append("20MA走平")

    if pd.notna(vol) and pd.notna(vol20) and vol20 > 0:
        vr20 = vol / vol20
        if vr20 >= 2.0:
            score += 16
            tags.append("量能爆發")
        elif vr20 >= 1.4:
            score += 10
            tags.append("量能放大")
        elif vr20 >= 1.0:
            score += 5
            tags.append("量能不弱")

    if pd.notna(vol) and pd.notna(vol5) and vol5 > 0:
        vr5 = vol / vol5
        if vr5 >= 1.5:
            score += 6
            tags.append("短線量增")

    if pd.notna(high13) and high13 > 0:
        d13 = safe_pct(close_, high13)
        if d13 is not None:
            if d13 >= -2:
                score += 10
                tags.append("近13週高")
            elif d13 >= -5:
                score += 5
                tags.append("接近13週高")

    if pd.notna(high26) and high26 > 0:
        d26 = safe_pct(close_, high26)
        if d26 is not None:
            if d26 >= -3:
                score += 8
                tags.append("近26週高")
            elif d26 >= -8:
                score += 4
                tags.append("接近26週高")

    if pd.notna(high52) and high52 > 0:
        d52 = safe_pct(close_, high52)
        if d52 is not None:
            if d52 >= -5:
                score += 8
                tags.append("接近52週高")
            elif d52 >= -12:
                score += 4
                tags.append("中長期強")

    if trend_info is not None:
        line_dist = trend_info.get("line_distance_pct")
        box_dist = trend_info.get("box_distance_pct")
        strict_ok = trend_info.get("strict_ok", False)
        training_hold_ok = trend_info.get("training_hold_ok", False)

        if strict_ok:
            score += 20
            tags.append("最新正式突破")

        if training_hold_ok:
            score += 24
            tags.append("突破後守穩")
        else:
            if line_dist is not None:
                if line_dist >= -1.5:
                    score += 8
                    tags.append("接近趨勢線")
                elif line_dist >= -4:
                    score += 4
                    tags.append("逼近趨勢線")

            if box_dist is not None:
                if box_dist >= -1.5:
                    score += 8
                    tags.append("接近盤整突破")
                elif box_dist >= -4:
                    score += 4
                    tags.append("逼近盤整突破")

        mem_bonus = trend_info.get("memory_bonus", 0)
        if mem_bonus > 0:
            score += mem_bonus
            tags.append("經驗加分")
        elif mem_bonus < 0:
            score += mem_bonus
            tags.append("經驗扣分")

    if pd.notna(open_) and pd.notna(close_) and open_ > 0:
        body_pct_val = safe_pct(close_, open_)
        if body_pct_val is not None:
            if body_pct_val >= 8:
                score += 10
                tags.append("長紅強攻")
            elif body_pct_val >= 4:
                score += 6
                tags.append("中紅K")

    if pd.notna(vol):
        if vol >= 50000:
            score += 8
            tags.append("週量大")
        elif vol >= 20000:
            score += 5
            tags.append("週量中大")
        elif vol >= 10000:
            score += 3
            tags.append("週量達標")

    uniq_tags = []
    for t in tags:
        if t not in uniq_tags:
            uniq_tags.append(t)

    return round(score, 2), "、".join(uniq_tags[:12])


def calc_bearish_training_score(grp, trend_info):
    latest = grp.iloc[-1]
    score = 0
    tags = []

    close_ = latest["週收盤價"]
    ma3 = latest["週3MA"]
    ma5 = latest["週5MA"]
    ma10 = latest["週10MA"]
    ma20 = latest["週20MA"]
    vol = latest["週成交量(張)"]
    vol5 = latest["量5MA"]
    vol20 = latest["量20MA"]
    low13 = latest["近13週最低"]
    low26 = latest["近26週最低"]
    low52 = latest["近52週最低"]
    slope20 = latest["20MA斜率"]
    open_ = latest["週開盤價"]

    if pd.notna(ma20) and close_ < ma20:
        score += 18
        tags.append("跌破20MA")
    elif pd.notna(ma20) and close_ <= ma20 * 1.03:
        score += 8
        tags.append("接近20MA下")

    if pd.notna(ma3) and pd.notna(ma5) and pd.notna(ma10) and pd.notna(ma20):
        if ma3 <= ma5 <= ma10 <= ma20:
            score += 18
            tags.append("均線空頭")
        elif ma3 <= ma5 and ma5 <= ma20:
            score += 10
            tags.append("均線轉弱")

    if pd.notna(slope20):
        if slope20 < 0:
            score += 8
            tags.append("20MA下彎")
        elif slope20 < 0.05:
            score += 3
            tags.append("20MA走平偏弱")

    if pd.notna(vol) and pd.notna(vol20) and vol20 > 0:
        vr20 = vol / vol20
        if vr20 >= 2.0:
            score += 16
            tags.append("量能放大")
        elif vr20 >= 1.4:
            score += 10
            tags.append("量能增強")
        elif vr20 >= 1.0:
            score += 5
            tags.append("量能不弱")

    if pd.notna(vol) and pd.notna(vol5) and vol5 > 0:
        vr5 = vol / vol5
        if vr5 >= 1.5:
            score += 6
            tags.append("短線量增")

    if pd.notna(low13) and low13 > 0:
        d13 = safe_pct(close_, low13)
        if d13 is not None:
            if d13 <= 2:
                score += 10
                tags.append("近13週低")
            elif d13 <= 5:
                score += 5
                tags.append("接近13週低")

    if pd.notna(low26) and low26 > 0:
        d26 = safe_pct(close_, low26)
        if d26 is not None:
            if d26 <= 3:
                score += 8
                tags.append("近26週低")
            elif d26 <= 8:
                score += 4
                tags.append("接近26週低")

    if pd.notna(low52) and low52 > 0:
        d52 = safe_pct(close_, low52)
        if d52 is not None:
            if d52 <= 5:
                score += 8
                tags.append("接近52週低")
            elif d52 <= 12:
                score += 4
                tags.append("中長期弱勢")

    if trend_info is not None:
        line_dist = trend_info.get("line_distance_pct")
        training_hold_ok = trend_info.get("training_hold_ok", False)
        break_now = trend_info.get("line_break_now", False)

        if break_now:
            score += 20
            tags.append("最新正式跌破")

        if training_hold_ok:
            score += 24
            tags.append("跌破後守弱")

        if line_dist is not None:
            if line_dist <= 1.5:
                score += 8
                tags.append("接近上升趨勢線")
            elif line_dist <= 4:
                score += 4
                tags.append("逼近上升趨勢線")

        mem_bonus = trend_info.get("memory_bonus", 0)
        if mem_bonus > 0:
            score += mem_bonus
            tags.append("經驗加分")
        elif mem_bonus < 0:
            score += mem_bonus
            tags.append("經驗扣分")

    if pd.notna(open_) and pd.notna(close_) and open_ > 0:
        body_pct_val = safe_pct(close_, open_)
        if body_pct_val is not None:
            if body_pct_val <= -8:
                score += 10
                tags.append("長黑轉弱")
            elif body_pct_val <= -4:
                score += 6
                tags.append("中黑K")

    if pd.notna(vol):
        if vol >= 50000:
            score += 8
            tags.append("週量大")
        elif vol >= 20000:
            score += 5
            tags.append("週量中大")
        elif vol >= 10000:
            score += 3
            tags.append("週量達標")

    uniq_tags = []
    for t in tags:
        if t not in uniq_tags:
            uniq_tags.append(t)

    return round(score, 2), "、".join(uniq_tags[:12])


# =========================
# 星等 / Alarm
# =========================
def star_text(n):
    n = int(clamp(n, 1, 5))
    return "★" * n + "☆" * (5 - n)


def score_to_star(score):
    if score >= 90:
        return 5
    elif score >= 75:
        return 4
    elif score >= 60:
        return 3
    elif score >= 45:
        return 2
    return 1


def calc_bias_pct(close_, ma20):
    if pd.isna(close_) or pd.isna(ma20) or ma20 in [0, None]:
        return None
    return (close_ - ma20) / ma20 * 100.0


def calc_bullish_star_and_alarm_by_score(score, grp):
    latest = grp.iloc[-1]
    prev = grp.iloc[-2] if len(grp) >= 2 else latest

    close_ = latest["週收盤價"]
    ma3 = latest["週3MA"]
    ma5 = latest["週5MA"]
    ma20 = latest["週20MA"]
    prev_ma3 = prev["週3MA"]
    prev_ma5 = prev["週5MA"]

    bias = calc_bias_pct(close_, ma20)
    star = score_to_star(score)

    short_alarm = "否"
    long_alarm = "否"

    if pd.notna(prev_ma3) and pd.notna(prev_ma5) and pd.notna(ma3) and pd.notna(ma5):
        if prev_ma3 >= prev_ma5 and ma3 < ma5:
            short_alarm = "是"

    if pd.notna(ma20) and pd.notna(close_) and close_ < ma20:
        long_alarm = "是"

    return {
        "星等數值": star,
        "星等": star_text(star),
        "乖離率(%)": round(bias, 2) if bias is not None else None,
        "短線停利Alarm": short_alarm,
        "長線停利Alarm": long_alarm,
    }


def calc_bearish_star_and_alarm_by_score(score, grp):
    latest = grp.iloc[-1]
    prev = grp.iloc[-2] if len(grp) >= 2 else latest

    close_ = latest["週收盤價"]
    ma3 = latest["週3MA"]
    ma5 = latest["週5MA"]
    ma20 = latest["週20MA"]
    prev_ma3 = prev["週3MA"]
    prev_ma5 = prev["週5MA"]

    bias = calc_bias_pct(close_, ma20)
    star = score_to_star(score)

    short_alarm = "否"
    long_alarm = "否"

    if pd.notna(prev_ma3) and pd.notna(prev_ma5) and pd.notna(ma3) and pd.notna(ma5):
        if prev_ma3 <= prev_ma5 and ma3 > ma5:
            short_alarm = "是"

    if pd.notna(ma20) and pd.notna(close_) and close_ > ma20:
        long_alarm = "是"

    return {
        "星等數值": star,
        "星等": star_text(star),
        "乖離率(%)": round(bias, 2) if bias is not None else None,
        "短線回補Alarm": short_alarm,
        "長線回補Alarm": long_alarm,
    }


# =========================
# 建立分頁
# =========================
def build_strict_breakout_sheet(weekly_ma_df, master_df):
    empty_cols = [
        "項次", "股票代號", "股票名稱", "市場別", "產業別",
        "週結算日期", "最新週收盤價", "週20MA", "最新一週成交量(張)",
        "第一高點日期", "第一高點價格", "第二高點日期", "第二高點價格",
        "盤整區高點", "盤整區低點", "最新壓力線價位",
        "最後一根是否突破趨勢線", "最後一根是否突破盤整",
        "最後一根是否紅K", "是否實K突破趨勢線",
        "趨勢線距離(%)", "盤整區距離(%)", "穿越K棒數", "記憶回饋"
    ]
    if weekly_ma_df.empty:
        return pd.DataFrame(columns=empty_cols)

    df = weekly_ma_df.copy()
    df["週結算日期"] = pd.to_datetime(df["週結算日期"])
    df = df.sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)

    industry_map = master_df[["股票代號", "產業別"]].drop_duplicates(subset=["股票代號"]).copy()
    rows = []

    for code, grp in df.groupby("股票代號"):
        grp = grp.sort_values("週結算日期").reset_index(drop=True).copy()
        latest = grp.iloc[-1]

        if pd.isna(latest["週20MA"]):
            continue
        if latest["週收盤價"] < latest["週20MA"]:
            continue
        if latest["週成交量(張)"] < MIN_WEEKLY_VOLUME:
            continue

        trend = analyze_best_descending_trendline(grp)
        if trend is None or not trend["strict_ok"]:
            continue

        work = trend["work_df"]
        i = trend["i"]
        j = trend["j"]

        industry = industry_map.loc[industry_map["股票代號"] == code, "產業別"]
        industry_val = industry.iloc[0] if len(industry) > 0 else "未分類"

        latest_open = latest["週開盤價"]
        latest_close = latest["週收盤價"]
        latest_line = trend["latest_line"]

        rows.append({
            "股票代號": str(code),
            "股票名稱": str(latest["股票名稱"]),
            "市場別": str(latest["市場別"]),
            "產業別": str(industry_val),
            "週結算日期": latest["週結算日期"].strftime("%Y-%m-%d"),
            "最新週收盤價": round(float(latest["週收盤價"]), 4),
            "週20MA": round(float(latest["週20MA"]), 4),
            "最新一週成交量(張)": int(latest["週成交量(張)"]),
            "第一高點日期": work.iloc[i]["週結算日期"].strftime("%Y-%m-%d"),
            "第一高點價格": round(float(work.iloc[i]["週最高價"]), 4),
            "第二高點日期": work.iloc[j]["週結算日期"].strftime("%Y-%m-%d"),
            "第二高點價格": round(float(work.iloc[j]["週最高價"]), 4),
            "盤整區高點": round(float(trend["box_high"]), 4),
            "盤整區低點": round(float(trend["box_low"]), 4),
            "最新壓力線價位": round(float(latest_line), 4),
            "最後一根是否突破趨勢線": "是" if trend["line_break_now"] else "否",
            "最後一根是否突破盤整": "是" if trend["box_break_now"] else "否",
            "最後一根是否紅K": "是" if is_red_k(latest_open, latest_close) else "否",
            "是否實K突破趨勢線": "是" if is_real_body_breakout(latest_open, latest_close, latest_line, BREAKOUT_BUFFER_PCT) else "否",
            "趨勢線距離(%)": round(float(trend["line_distance_pct"]), 2) if trend["line_distance_pct"] is not None else None,
            "盤整區距離(%)": round(float(trend["box_distance_pct"]), 2) if trend["box_distance_pct"] is not None else None,
            "穿越K棒數": 1,
            "記憶回饋": trend.get("memory_text", ""),
        })

    if not rows:
        return pd.DataFrame(columns=empty_cols)

    out_df = pd.DataFrame(rows)
    out_df = out_df.sort_values(
        by=["盤整區距離(%)", "趨勢線距離(%)", "最新一週成交量(張)", "股票代號"],
        ascending=[False, False, False, True]
    ).reset_index(drop=True)
    out_df.insert(0, "項次", range(1, len(out_df) + 1))
    return out_df


def build_training_pool_sheet(weekly_ma_df, master_df):
    empty_cols = [
        "項次", "股票代號", "股票名稱", "市場別", "產業別",
        "週結算日期", "最新週收盤價", "週3MA", "週5MA", "週10MA", "週20MA",
        "最新一週成交量(張)", "量20MA", "近13週最高", "近26週最高", "近52週最高",
        "最佳第一高點日期", "最佳第一高點價格", "最佳第二高點日期", "最佳第二高點價格",
        "最新壓力線價位", "盤整區高點", "盤整區低點", "趨勢線距離(%)", "盤整區距離(%)",
        "是否最新正式突破", "是否突破後守穩趨勢線", "有效突破日期", "有效突破開盤價", "有效突破收盤價",
        "有效突破時壓力線價位", "有效突破是否紅K", "有效突破是否實K站上",
        "StrongScore", "技術標籤", "記憶回饋", "記憶分數",
        "星等", "星等數值", "乖離率(%)", "短線停利Alarm", "長線停利Alarm"
    ]

    if weekly_ma_df.empty:
        return pd.DataFrame(columns=empty_cols)

    df = weekly_ma_df.copy()
    df["週結算日期"] = pd.to_datetime(df["週結算日期"])
    df = df.sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)

    industry_map = master_df[["股票代號", "產業別"]].drop_duplicates(subset=["股票代號"]).copy()
    rows = []

    for code, grp in df.groupby("股票代號"):
        grp = grp.sort_values("週結算日期").reset_index(drop=True).copy()
        latest = grp.iloc[-1]

        if pd.isna(latest["週20MA"]):
            continue
        if pd.isna(latest["週成交量(張)"]):
            continue
        if latest["週成交量(張)"] < MIN_WEEKLY_VOLUME:
            continue

        trend = analyze_best_descending_trendline(grp)
        score, tags = calc_training_score(grp, trend)

        base_ok = (
            (latest["週收盤價"] >= latest["週20MA"] * 0.95 if pd.notna(latest["週20MA"]) else False)
            or
            (latest["均線多頭排列"] if pd.notna(latest["均線多頭排列"]) else False)
        )

        if not base_ok:
            continue
        if trend is None:
            continue
        if not trend["training_hold_ok"]:
            continue
        if score < TRAINING_SCORE_THRESHOLD:
            continue

        industry = industry_map.loc[industry_map["股票代號"] == code, "產業別"]
        industry_val = industry.iloc[0] if len(industry) > 0 else "未分類"

        work = trend["work_df"]
        i = trend["i"]
        j = trend["j"]
        hold_info = trend["hold_info"]

        hold_open = hold_info["breakout_open"] if hold_info else None
        hold_close = hold_info["breakout_close"] if hold_info else None
        hold_line = hold_info["breakout_line"] if hold_info else None

        extra = calc_bullish_star_and_alarm_by_score(score, grp)

        rows.append({
            "股票代號": str(code),
            "股票名稱": str(latest["股票名稱"]),
            "市場別": str(latest["市場別"]),
            "產業別": str(industry_val),
            "週結算日期": latest["週結算日期"].strftime("%Y-%m-%d"),
            "最新週收盤價": round(float(latest["週收盤價"]), 4),
            "週3MA": round(float(latest["週3MA"]), 4) if pd.notna(latest["週3MA"]) else None,
            "週5MA": round(float(latest["週5MA"]), 4) if pd.notna(latest["週5MA"]) else None,
            "週10MA": round(float(latest["週10MA"]), 4) if pd.notna(latest["週10MA"]) else None,
            "週20MA": round(float(latest["週20MA"]), 4) if pd.notna(latest["週20MA"]) else None,
            "最新一週成交量(張)": int(latest["週成交量(張)"]),
            "量20MA": round(float(latest["量20MA"]), 2) if pd.notna(latest["量20MA"]) else None,
            "近13週最高": round(float(latest["近13週最高"]), 4) if pd.notna(latest["近13週最高"]) else None,
            "近26週最高": round(float(latest["近26週最高"]), 4) if pd.notna(latest["近26週最高"]) else None,
            "近52週最高": round(float(latest["近52週最高"]), 4) if pd.notna(latest["近52週最高"]) else None,
            "最佳第一高點日期": work.iloc[i]["週結算日期"].strftime("%Y-%m-%d"),
            "最佳第一高點價格": round(float(work.iloc[i]["週最高價"]), 4),
            "最佳第二高點日期": work.iloc[j]["週結算日期"].strftime("%Y-%m-%d"),
            "最佳第二高點價格": round(float(work.iloc[j]["週最高價"]), 4),
            "最新壓力線價位": round(float(trend["latest_line"]), 4) if pd.notna(trend["latest_line"]) else None,
            "盤整區高點": round(float(trend["box_high"]), 4) if pd.notna(trend["box_high"]) else None,
            "盤整區低點": round(float(trend["box_low"]), 4) if pd.notna(trend["box_low"]) else None,
            "趨勢線距離(%)": round(float(trend["line_distance_pct"]), 2) if trend["line_distance_pct"] is not None else None,
            "盤整區距離(%)": round(float(trend["box_distance_pct"]), 2) if trend["box_distance_pct"] is not None else None,
            "是否最新正式突破": "是" if trend["strict_ok"] else "否",
            "是否突破後守穩趨勢線": "是",
            "有效突破日期": hold_info["breakout_date"].strftime("%Y-%m-%d") if hold_info is not None else None,
            "有效突破開盤價": round(float(hold_open), 4) if hold_open is not None else None,
            "有效突破收盤價": round(float(hold_close), 4) if hold_close is not None else None,
            "有效突破時壓力線價位": round(float(hold_line), 4) if hold_line is not None else None,
            "有效突破是否紅K": "是" if is_red_k(hold_open, hold_close) else "否",
            "有效突破是否實K站上": "是" if is_real_body_breakout(hold_open, hold_close, hold_line, BREAKOUT_BUFFER_PCT) else "否",
            "StrongScore": score,
            "技術標籤": tags,
            "記憶回饋": trend.get("memory_text", ""),
            "記憶分數": trend.get("memory_bonus", 0),
            "星等": extra["星等"],
            "星等數值": extra["星等數值"],
            "乖離率(%)": extra["乖離率(%)"],
            "短線停利Alarm": extra["短線停利Alarm"],
            "長線停利Alarm": extra["長線停利Alarm"],
        })

    if not rows:
        return pd.DataFrame(columns=empty_cols)

    out_df = pd.DataFrame(rows)
    out_df = out_df.sort_values(
        by=["StrongScore", "星等數值", "是否最新正式突破", "最新一週成交量(張)", "股票代號"],
        ascending=[False, False, False, False, True]
    ).reset_index(drop=True)
    out_df.insert(0, "項次", range(1, len(out_df) + 1))
    return out_df


def build_bearish_key_breakdown_sheet(weekly_ma_df, master_df):
    empty_cols = [
        "項次", "股票代號", "股票名稱", "市場別", "產業別",
        "週結算日期", "最新週收盤價", "週20MA", "最新一週成交量(張)",
        "第一低點日期", "第一低點價格", "第二低點日期", "第二低點價格",
        "最新支撐線價位", "最後一根是否跌破趨勢線",
        "最後一根是否黑K", "是否實K跌破趨勢線",
        "趨勢線距離(%)", "穿越K棒數", "記憶回饋"
    ]
    if weekly_ma_df.empty:
        return pd.DataFrame(columns=empty_cols)

    df = weekly_ma_df.copy()
    df["週結算日期"] = pd.to_datetime(df["週結算日期"])
    df = df.sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)

    industry_map = master_df[["股票代號", "產業別"]].drop_duplicates(subset=["股票代號"]).copy()
    rows = []

    for code, grp in df.groupby("股票代號"):
        grp = grp.sort_values("週結算日期").reset_index(drop=True).copy()
        latest = grp.iloc[-1]

        if pd.isna(latest["週20MA"]):
            continue
        if latest["週收盤價"] >= latest["週20MA"]:
            continue
        if latest["週成交量(張)"] < MIN_WEEKLY_VOLUME:
            continue

        trend = analyze_best_ascending_trendline(grp)
        if trend is None:
            continue
        if not (trend["line_break_now"] and (not trend["line_break_prev"])):
            continue

        work = trend["work_df"]
        i = trend["i"]
        j = trend["j"]

        industry = industry_map.loc[industry_map["股票代號"] == code, "產業別"]
        industry_val = industry.iloc[0] if len(industry) > 0 else "未分類"

        latest_open = latest["週開盤價"]
        latest_close = latest["週收盤價"]
        latest_line = trend["latest_line"]

        rows.append({
            "股票代號": str(code),
            "股票名稱": str(latest["股票名稱"]),
            "市場別": str(latest["市場別"]),
            "產業別": str(industry_val),
            "週結算日期": latest["週結算日期"].strftime("%Y-%m-%d"),
            "最新週收盤價": round(float(latest["週收盤價"]), 4),
            "週20MA": round(float(latest["週20MA"]), 4),
            "最新一週成交量(張)": int(latest["週成交量(張)"]),
            "第一低點日期": work.iloc[i]["週結算日期"].strftime("%Y-%m-%d"),
            "第一低點價格": round(float(work.iloc[i]["週最低價"]), 4),
            "第二低點日期": work.iloc[j]["週結算日期"].strftime("%Y-%m-%d"),
            "第二低點價格": round(float(work.iloc[j]["週最低價"]), 4),
            "最新支撐線價位": round(float(latest_line), 4),
            "最後一根是否跌破趨勢線": "是" if trend["line_break_now"] else "否",
            "最後一根是否黑K": "是" if is_black_k(latest_open, latest_close) else "否",
            "是否實K跌破趨勢線": "是" if is_real_body_breakdown(latest_open, latest_close, latest_line, BREAKDOWN_BUFFER_PCT) else "否",
            "趨勢線距離(%)": round(float(trend["line_distance_pct"]), 2) if trend["line_distance_pct"] is not None else None,
            "穿越K棒數": 1,
            "記憶回饋": trend.get("memory_text", ""),
        })

    if not rows:
        return pd.DataFrame(columns=empty_cols)

    out_df = pd.DataFrame(rows)
    out_df = out_df.sort_values(
        by=["趨勢線距離(%)", "最新一週成交量(張)", "股票代號"],
        ascending=[False, False, True]
    ).reset_index(drop=True)
    out_df.insert(0, "項次", range(1, len(out_df) + 1))
    return out_df


def build_bearish_training_pool_sheet(weekly_ma_df, master_df):
    empty_cols = [
        "項次", "股票代號", "股票名稱", "市場別", "產業別",
        "週結算日期", "最新週收盤價", "週3MA", "週5MA", "週10MA", "週20MA",
        "最新一週成交量(張)", "量20MA", "近13週最低", "近26週最低", "近52週最低",
        "最佳第一低點日期", "最佳第一低點價格", "最佳第二低點日期", "最佳第二低點價格",
        "最新支撐線價位", "趨勢線距離(%)", "是否最新正式跌破", "是否跌破後持續在趨勢線下",
        "有效跌破日期", "有效跌破開盤價", "有效跌破收盤價",
        "有效跌破時支撐線價位", "有效跌破是否黑K", "有效跌破是否實K站下",
        "BearishScore", "空方技術標籤", "記憶回饋", "記憶分數",
        "星等", "星等數值", "乖離率(%)", "短線回補Alarm", "長線回補Alarm"
    ]

    if weekly_ma_df.empty:
        return pd.DataFrame(columns=empty_cols)

    df = weekly_ma_df.copy()
    df["週結算日期"] = pd.to_datetime(df["週結算日期"])
    df = df.sort_values(["股票代號", "週結算日期"]).reset_index(drop=True)

    industry_map = master_df[["股票代號", "產業別"]].drop_duplicates(subset=["股票代號"]).copy()
    rows = []

    for code, grp in df.groupby("股票代號"):
        grp = grp.sort_values("週結算日期").reset_index(drop=True).copy()
        latest = grp.iloc[-1]

        if pd.isna(latest["週20MA"]):
            continue
        if pd.isna(latest["週成交量(張)"]):
            continue
        if latest["週成交量(張)"] < MIN_WEEKLY_VOLUME:
            continue
        if latest["週收盤價"] >= latest["週20MA"]:
            continue

        trend = analyze_best_ascending_trendline(grp)
        score, tags = calc_bearish_training_score(grp, trend)

        base_ok = (
            (latest["週收盤價"] <= latest["週20MA"] * 1.05 if pd.notna(latest["週20MA"]) else False)
            or
            (latest["均線空頭排列"] if pd.notna(latest["均線空頭排列"]) else False)
        )

        if not base_ok:
            continue
        if trend is None:
            continue
        if not trend["training_hold_ok"]:
            continue
        if score < BEARISH_TRAINING_SCORE_THRESHOLD:
            continue

        industry = industry_map.loc[industry_map["股票代號"] == code, "產業別"]
        industry_val = industry.iloc[0] if len(industry) > 0 else "未分類"

        work = trend["work_df"]
        i = trend["i"]
        j = trend["j"]
        hold_info = trend["hold_info"]

        hold_open = hold_info["breakdown_open"] if hold_info else None
        hold_close = hold_info["breakdown_close"] if hold_info else None
        hold_line = hold_info["breakdown_line"] if hold_info else None

        extra = calc_bearish_star_and_alarm_by_score(score, grp)

        rows.append({
            "股票代號": str(code),
            "股票名稱": str(latest["股票名稱"]),
            "市場別": str(latest["市場別"]),
            "產業別": str(industry_val),
            "週結算日期": latest["週結算日期"].strftime("%Y-%m-%d"),
            "最新週收盤價": round(float(latest["週收盤價"]), 4),
            "週3MA": round(float(latest["週3MA"]), 4) if pd.notna(latest["週3MA"]) else None,
            "週5MA": round(float(latest["週5MA"]), 4) if pd.notna(latest["週5MA"]) else None,
            "週10MA": round(float(latest["週10MA"]), 4) if pd.notna(latest["週10MA"]) else None,
            "週20MA": round(float(latest["週20MA"]), 4) if pd.notna(latest["週20MA"]) else None,
            "最新一週成交量(張)": int(latest["週成交量(張)"]),
            "量20MA": round(float(latest["量20MA"]), 2) if pd.notna(latest["量20MA"]) else None,
            "近13週最低": round(float(latest["近13週最低"]), 4) if pd.notna(latest["近13週最低"]) else None,
            "近26週最低": round(float(latest["近26週最低"]), 4) if pd.notna(latest["近26週最低"]) else None,
            "近52週最低": round(float(latest["近52週最低"]), 4) if pd.notna(latest["近52週最低"]) else None,
            "最佳第一低點日期": work.iloc[i]["週結算日期"].strftime("%Y-%m-%d"),
            "最佳第一低點價格": round(float(work.iloc[i]["週最低價"]), 4),
            "最佳第二低點日期": work.iloc[j]["週結算日期"].strftime("%Y-%m-%d"),
            "最佳第二低點價格": round(float(work.iloc[j]["週最低價"]), 4),
            "最新支撐線價位": round(float(trend["latest_line"]), 4) if pd.notna(trend["latest_line"]) else None,
            "趨勢線距離(%)": round(float(trend["line_distance_pct"]), 2) if trend["line_distance_pct"] is not None else None,
            "是否最新正式跌破": "是" if trend["line_break_now"] and (not trend["line_break_prev"]) else "否",
            "是否跌破後持續在趨勢線下": "是",
            "有效跌破日期": hold_info["breakdown_date"].strftime("%Y-%m-%d") if hold_info is not None else None,
            "有效跌破開盤價": round(float(hold_open), 4) if hold_open is not None else None,
            "有效跌破收盤價": round(float(hold_close), 4) if hold_close is not None else None,
            "有效跌破時支撐線價位": round(float(hold_line), 4) if hold_line is not None else None,
            "有效跌破是否黑K": "是" if is_black_k(hold_open, hold_close) else "否",
            "有效跌破是否實K站下": "是" if is_real_body_breakdown(hold_open, hold_close, hold_line, BREAKDOWN_BUFFER_PCT) else "否",
            "BearishScore": score,
            "空方技術標籤": tags,
            "記憶回饋": trend.get("memory_text", ""),
            "記憶分數": trend.get("memory_bonus", 0),
            "星等": extra["星等"],
            "星等數值": extra["星等數值"],
            "乖離率(%)": extra["乖離率(%)"],
            "短線回補Alarm": extra["短線回補Alarm"],
            "長線回補Alarm": extra["長線回補Alarm"],
        })

    if not rows:
        return pd.DataFrame(columns=empty_cols)

    out_df = pd.DataFrame(rows)
    out_df = out_df.sort_values(
        by=["BearishScore", "星等數值", "是否最新正式跌破", "最新一週成交量(張)", "股票代號"],
        ascending=[False, False, False, False, True]
    ).reset_index(drop=True)
    out_df.insert(0, "項次", range(1, len(out_df) + 1))
    return out_df


# =========================
# 客戶端顯示資料
# =========================
def build_client_bullish_view(training_df):
    cols = [
        "股票代號", "股票名稱", "產業別", "週結算日期",
        "星等", "乖離率(%)", "短線停利Alarm", "長線停利Alarm"
    ]
    if training_df is None or training_df.empty:
        return pd.DataFrame(columns=["項次"] + cols)

    out = training_df[cols].copy()
    out = out.drop_duplicates(subset=["股票代號"]).reset_index(drop=True)
    out = out.sort_values(["股票代號"]).reset_index(drop=True)
    out.insert(0, "項次", range(1, len(out) + 1))
    return out


def build_client_bearish_view(bearish_df):
    cols = [
        "股票代號", "股票名稱", "產業別", "週結算日期",
        "星等", "乖離率(%)", "短線回補Alarm", "長線回補Alarm"
    ]
    if bearish_df is None or bearish_df.empty:
        return pd.DataFrame(columns=["項次"] + cols)

    out = bearish_df[cols].copy()
    out = out.drop_duplicates(subset=["股票代號"]).reset_index(drop=True)
    out = out.sort_values(["股票代號"]).reset_index(drop=True)
    out.insert(0, "項次", range(1, len(out) + 1))
    return out


def build_client_bullish_keyk_view(strict_df):
    cols = [
        "股票代號", "股票名稱", "產業別", "週結算日期",
        "最新週收盤價", "週20MA", "最新一週成交量(張)",
        "趨勢線距離(%)", "盤整區距離(%)"
    ]
    if strict_df is None or strict_df.empty:
        return pd.DataFrame(columns=["項次"] + cols)

    use_cols = [c for c in cols if c in strict_df.columns]
    out = strict_df[use_cols].copy()
    out = out.drop_duplicates(subset=["股票代號"]).reset_index(drop=True)
    sort_cols = [c for c in ["盤整區距離(%)", "趨勢線距離(%)", "最新一週成交量(張)", "股票代號"] if c in out.columns]
    ascending = []
    for c in sort_cols:
        ascending.append(False if c != "股票代號" else True)
    if sort_cols:
        out = out.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)
    out.insert(0, "項次", range(1, len(out) + 1))
    return out


def build_client_bearish_keyk_view(bearish_key_df):
    cols = [
        "股票代號", "股票名稱", "產業別", "週結算日期",
        "最新週收盤價", "週20MA", "最新一週成交量(張)",
        "趨勢線距離(%)"
    ]
    if bearish_key_df is None or bearish_key_df.empty:
        return pd.DataFrame(columns=["項次"] + cols)

    use_cols = [c for c in cols if c in bearish_key_df.columns]
    out = bearish_key_df[use_cols].copy()
    out = out.drop_duplicates(subset=["股票代號"]).reset_index(drop=True)
    sort_cols = [c for c in ["趨勢線距離(%)", "最新一週成交量(張)", "股票代號"] if c in out.columns]
    ascending = []
    for c in sort_cols:
        ascending.append(False if c != "股票代號" else True)
    if sort_cols:
        out = out.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)
    out.insert(0, "項次", range(1, len(out) + 1))
    return out


def build_holdings_view(holdings_codes, result_dict):
    master = result_dict.get("MASTER_STOCK_LIST", pd.DataFrame())
    bull = result_dict.get("TRAINING_POOL", pd.DataFrame())
    bear = result_dict.get("BEARISH_TRAINING_POOL", pd.DataFrame())

    rows = []
    for code in holdings_codes:
        code = normalize_code(code)
        if not is_valid_stock_code(code):
            continue

        name = ""
        industry = ""
        if isinstance(master, pd.DataFrame) and not master.empty:
            m = master[master["股票代號"].astype(str) == code]
            if not m.empty:
                name = str(m.iloc[0].get("股票名稱", ""))
                industry = str(m.iloc[0].get("產業別", ""))

        bull_one = bull[bull["股票代號"].astype(str) == code] if isinstance(bull, pd.DataFrame) and not bull.empty else pd.DataFrame()
        bear_one = bear[bear["股票代號"].astype(str) == code] if isinstance(bear, pd.DataFrame) and not bear.empty else pd.DataFrame()

        if not bull_one.empty:
            r = bull_one.iloc[0]
            rows.append({
                "股票代號": code,
                "股票名稱": r.get("股票名稱", name),
                "產業別": r.get("產業別", industry),
                "來源": "系統看多池",
                "狀態": "系統選出",
                "星等": r.get("星等", ""),
                "乖離率(%)": r.get("乖離率(%)", ""),
                "短線提醒": r.get("短線停利Alarm", ""),
                "長線提醒": r.get("長線停利Alarm", ""),
                "意見": "可提供意見",
            })
        elif not bear_one.empty:
            r = bear_one.iloc[0]
            rows.append({
                "股票代號": code,
                "股票名稱": r.get("股票名稱", name),
                "產業別": r.get("產業別", industry),
                "來源": "系統看空池",
                "狀態": "系統選出",
                "星等": r.get("星等", ""),
                "乖離率(%)": r.get("乖離率(%)", ""),
                "短線提醒": r.get("短線回補Alarm", ""),
                "長線提醒": r.get("長線回補Alarm", ""),
                "意見": "可提供意見",
            })
        else:
            rows.append({
                "股票代號": code,
                "股票名稱": name if name else "",
                "產業別": industry if industry else "",
                "來源": "持股池外",
                "狀態": "客戶自選",
                "星等": "",
                "乖離率(%)": "",
                "短線提醒": "",
                "長線提醒": "",
                "意見": "無法提供意見",
            })

    if not rows:
        return pd.DataFrame(columns=[
            "項次", "股票代號", "股票名稱", "產業別", "來源", "狀態",
            "星等", "乖離率(%)", "短線提醒", "長線提醒", "意見"
        ])

    out = pd.DataFrame(rows)
    out = out.drop_duplicates(subset=["股票代號"]).reset_index(drop=True)
    out.insert(0, "項次", range(1, len(out) + 1))
    return out


# =========================
# Excel 格式
# =========================
def beautify_sheet(ws, widths=None):
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if widths:
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


# =========================
# 主分析函式
# =========================
def build_all_excel(logger=None, force_rebuild_snapshot=False):
    def log(msg):
        if logger:
            logger(msg)
        else:
            print(msg)

    log("=== 開始建立 TWSE_ALL.xlsx（穩定快照版）===")
    ensure_output_dir()

    target_settle_date = get_latest_available_trading_date(max_lookback_days=20)
    target_settle_date_str = str(target_settle_date)
    snapshot_path = get_snapshot_path(target_settle_date_str)

    if (not force_rebuild_snapshot) and os.path.exists(snapshot_path):
        log(f"偵測到當日快照，檢查相容性：{snapshot_path}")
        result = load_snapshot(target_settle_date_str)

        if is_snapshot_compatible(result):
            log("snapshot 相容，直接載入。")
            result["excel_path"] = OUTPUT_XLSX
            return result
        else:
            log("舊 snapshot 與目前版本不相容，將自動重新建構新快照。")

    log("1. 抓取上市/上櫃一般股票清單 + 產業別...")
    master = get_master_stock_list()
    log(f"股票總數：{len(master)}")

    log(f"2. 本次固定使用結算日：{target_settle_date}")

    listed_codes = master.loc[master["市場別"] == "上市", "股票代號"].tolist()
    otc_codes = master.loc[master["市場別"] == "上櫃", "股票代號"].tolist()

    log("3. 抓取上市本週最新資料...")
    listed_week_map = fetch_market_week_data("上市", listed_codes, target_settle_date, logger=log)

    log("4. 抓取上櫃本週最新資料...")
    otc_week_map = fetch_market_week_data("上櫃", otc_codes, target_settle_date, logger=log)

    log("5. 整理 TWSE_ALL 分頁...")
    rows = []
    for idx, row in master.reset_index(drop=True).iterrows():
        code = row["股票代號"]
        market = row["市場別"]

        if market == "上市":
            wk = listed_week_map.get(code, None)
            source = "TWSE ISIN + TWSE 每日收盤行情"
        else:
            wk = otc_week_map.get(code, None)
            source = "TWSE ISIN + TPEx 日收盤行情"

        if wk is None:
            week_settle_date = ""
            friday_close = None
            weekly_volume = None
            missing_reason = "本週查無交易資料或網站暫無回傳"
        else:
            week_settle_date = wk.get("週結算日期", "")
            friday_close = wk.get("每週五收盤後股價", None)
            weekly_volume = wk.get("每週總成交量(張)", None)
            missing_reason = wk.get("缺漏原因", "")

        rows.append({
            "項次": idx + 1,
            "股票代號": row["股票代號"],
            "股票名稱": row["股票名稱"],
            "市場別": row["市場別"],
            "產業別": row["產業別"],
            "週結算日期": week_settle_date,
            "每週五收盤後股價": friday_close,
            "每週總成交量(張)": weekly_volume,
            "資料來源": source,
            "缺漏原因": missing_reason,
        })

    twse_all_df = pd.DataFrame(rows)

    log("6. 抓取週K計算所需歷史日資料...")
    history_end = target_settle_date
    history_start = history_end - timedelta(days=460)
    log(f"歷史資料區間：{history_start} ~ {history_end}")

    listed_daily = fetch_market_daily_history("上市", listed_codes, history_start, history_end, logger=log)
    otc_daily = fetch_market_daily_history("上櫃", otc_codes, history_start, history_end, logger=log)

    if listed_daily.empty and otc_daily.empty:
        daily_all = pd.DataFrame()
    elif listed_daily.empty:
        daily_all = otc_daily.copy()
    elif otc_daily.empty:
        daily_all = listed_daily.copy()
    else:
        daily_all = pd.concat([listed_daily, otc_daily], ignore_index=True)

    log("7. 轉換週K...")
    weekly_df = build_weekly_k_from_daily(daily_all)

    log("8. 計算週指標...")
    weekly_ma_df = calculate_weekly_indicators(weekly_df)

    log("9. 建立 WEEK_20MA...")
    week_20ma_df = get_latest_week_20ma_candidates(weekly_ma_df, master, MIN_WEEKLY_VOLUME)

    log("10. 建立 STRICT_BREAKOUT...")
    strict_df = build_strict_breakout_sheet(weekly_ma_df, master)

    log("11. 建立 TRAINING_POOL...")
    training_df = build_training_pool_sheet(weekly_ma_df, master)

    log("12. 建立 BEARISH_KEY_BREAKDOWN...")
    bearish_key_df = build_bearish_key_breakdown_sheet(weekly_ma_df, master)

    log("13. 建立 BEARISH_TRAINING_POOL...")
    bearish_training_df = build_bearish_training_pool_sheet(weekly_ma_df, master)

    log("14. 建立客戶展示頁...")
    client_bullish_df = build_client_bullish_view(training_df)
    client_bullish_keyk_df = build_client_bullish_keyk_view(strict_df)
    client_bearish_df = build_client_bearish_view(bearish_training_df)
    client_bearish_keyk_df = build_client_bearish_keyk_view(bearish_key_df)

    save_path = get_writable_output_path(OUTPUT_XLSX)

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        twse_all_df.to_excel(writer, index=False, sheet_name="TWSE_ALL")
        week_20ma_df.to_excel(writer, index=False, sheet_name="WEEK_20MA")
        strict_df.to_excel(writer, index=False, sheet_name="STRICT_BREAKOUT")
        bearish_key_df.to_excel(writer, index=False, sheet_name="BEARISH_KEY_BREAKDOWN")
        training_df.to_excel(writer, index=False, sheet_name="TRAINING_POOL")
        bearish_training_df.to_excel(writer, index=False, sheet_name="BEARISH_TRAINING_POOL")

        beautify_sheet(writer.book["TWSE_ALL"])
        beautify_sheet(writer.book["WEEK_20MA"])
        beautify_sheet(writer.book["STRICT_BREAKOUT"])
        beautify_sheet(writer.book["BEARISH_KEY_BREAKDOWN"])
        beautify_sheet(writer.book["TRAINING_POOL"])
        beautify_sheet(writer.book["BEARISH_TRAINING_POOL"])

    result = {
        "excel_path": save_path,
        "settle_date": target_settle_date_str,
        "TWSE_ALL": twse_all_df,
        "WEEK_20MA": week_20ma_df,
        "STRICT_BREAKOUT": strict_df,
        "BEARISH_KEY_BREAKDOWN": bearish_key_df,
        "TRAINING_POOL": training_df,
        "BEARISH_TRAINING_POOL": bearish_training_df,
        "CLIENT_BULLISH": client_bullish_df,
        "CLIENT_BULLISH_KEYK": client_bullish_keyk_df,
        "CLIENT_BEARISH": client_bearish_df,
        "CLIENT_BEARISH_KEYK": client_bearish_keyk_df,
        "WEEKLY_MA_RAW": weekly_ma_df,
        "MASTER_STOCK_LIST": master,
    }

    save_snapshot(target_settle_date_str, result)

    log("=== 完成 ===")
    log(f"本次固定結算日：{target_settle_date}")
    log(f"快照已儲存：{snapshot_path}")
    log(f"TWSE_ALL 筆數：{len(twse_all_df)}")
    log(f"WEEK_20MA 筆數：{len(week_20ma_df)}")
    log(f"STRICT_BREAKOUT 筆數：{len(strict_df)}")
    log(f"BEARISH_KEY_BREAKDOWN 筆數：{len(bearish_key_df)}")
    log(f"TRAINING_POOL 筆數：{len(training_df)}")
    log(f"BEARISH_TRAINING_POOL 筆數：{len(bearish_training_df)}")
    log(f"客戶看多頁筆數：{len(client_bullish_df)}")
    log(f"客戶多方關鍵K頁筆數：{len(client_bullish_keyk_df)}")
    log(f"客戶看空頁筆數：{len(client_bearish_df)}")
    log(f"客戶空方關鍵K頁筆數：{len(client_bearish_keyk_df)}")

    return result


def load_json_file(path, default=None):
    if default is None:
        default = {}
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return default


def save_json_file(path, data):
    ensure_output_dir()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_local_device_info():
    data = load_json_file(DEVICE_INFO_FILE, default={})
    device_id = clean_text(data.get("device_id"))
    device_name = clean_text(data.get("device_name"))

    if not device_id:
        raw = f"{platform.node()}-{uuid.getnode()}"
        device_id = re.sub(r"[^A-Za-z0-9_-]", "", raw)[:64] or f"ZhuStock-{uuid.uuid4().hex[:12]}"

    if not device_name:
        device_name = platform.node() or socket.gethostname() or "Windows-PC"

    payload = {"device_id": device_id, "device_name": device_name}
    save_json_file(DEVICE_INFO_FILE, payload)
    return payload


def load_auth_session():
    return load_json_file(AUTH_SESSION_FILE, default={})


def save_auth_session(data):
    save_json_file(AUTH_SESSION_FILE, data)


def clear_auth_session():
    try:
        if os.path.exists(AUTH_SESSION_FILE):
            os.remove(AUTH_SESSION_FILE)
    except Exception:
        pass


def auth_post(path, payload, timeout=20, headers=None):
    url = f"{AUTH_SERVER_BASE}{path}"
    r = requests.post(url, json=payload, headers=headers or {}, timeout=timeout)
    try:
        r.raise_for_status()
    except requests.HTTPError:
        try:
            err = r.json()
            detail = err.get("detail") or err.get("message") or str(err)
        except Exception:
            detail = r.text
        raise RuntimeError(detail)
    return r.json()


def auth_get(path, headers=None, timeout=20):
    url = f"{AUTH_SERVER_BASE}{path}"
    r = requests.get(url, headers=headers or {}, timeout=timeout)
    try:
        r.raise_for_status()
    except requests.HTTPError:
        try:
            err = r.json()
            detail = err.get("detail") or err.get("message") or str(err)
        except Exception:
            detail = r.text
        raise RuntimeError(detail)
    return r.json()


def auth_request(method, path, payload=None, timeout=20, headers=None, params=None):
    url = f"{AUTH_SERVER_BASE}{path}"
    r = requests.request(method.upper(), url, json=payload, headers=headers or {}, timeout=timeout, params=params)
    try:
        r.raise_for_status()
    except requests.HTTPError:
        try:
            err = r.json()
            detail = err.get("detail") or err.get("message") or str(err)
        except Exception:
            detail = r.text or str(r.status_code)
        raise RuntimeError(detail)
    try:
        return r.json()
    except Exception:
        return {"success": True, "message": r.text.strip()}


class AuthDialog(tk.Toplevel):
    def __init__(self, master, device_info):
        super().__init__(master)
        self.master = master
        self.device_info = device_info
        self.result = None
        self.pending_register_payload = None
        self.pending_register_email = ""
        self.title("ZHU STOCK 授權登入")
        self.geometry("560x700")
        self.resizable(False, False)
        self.transient(master)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        try:
            self.iconbitmap(resource_path(ICON_FILE))
        except Exception:
            pass

        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="ZHU STOCK APP", font=("Microsoft JhengHei", 16, "bold")).pack(anchor="center", pady=(0, 4))
        ttk.Label(outer, text="請先登入或註冊後再使用系統", font=("Microsoft JhengHei", 10)).pack(anchor="center", pady=(0, 10))

        self.msg_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=self.msg_var, foreground="#B00020", wraplength=510).pack(anchor="w", pady=(0, 8))

        notebook = ttk.Notebook(outer)
        notebook.pack(fill="both", expand=True)

        login_tab = ttk.Frame(notebook, padding=14)
        reg_tab = ttk.Frame(notebook, padding=14)
        reset_tab = ttk.Frame(notebook, padding=14)
        notebook.add(login_tab, text="登入")
        notebook.add(reg_tab, text="註冊")
        notebook.add(reset_tab, text="忘記密碼")

        # login fields
        saved = load_auth_session()
        self.login_account = tk.StringVar(value=saved.get('login_account', '') or saved.get('email', '') or saved.get('account', ''))
        self.login_password = tk.StringVar()
        self.login_remember_me = tk.BooleanVar(value=bool(saved.get('remember_me', False) and self.login_account.get()))

        for label, var, show in [
            ("帳號或 Email", self.login_account, None),
            ("密碼", self.login_password, "*"),
        ]:
            ttk.Label(login_tab, text=label).pack(anchor="w", pady=(6, 0))
            ttk.Entry(login_tab, textvariable=var, width=46, show=show).pack(anchor="w", fill="x", pady=(2, 0))

        ttk.Checkbutton(login_tab, text="記住我", variable=self.login_remember_me).pack(anchor="w", pady=(8, 0))
        ttk.Button(login_tab, text="登入", command=self.do_login).pack(anchor="center", pady=16)
        ttk.Label(login_tab, text="提示：登入時系統會在背景自動驗證此電腦的授權，不需要手動輸入裝置代號。", foreground="#555555", wraplength=480).pack(anchor="w", pady=(4, 0))

        # register fields
        self.reg_username = tk.StringVar()
        self.reg_password = tk.StringVar()
        self.reg_confirm_password = tk.StringVar()
        self.reg_phone = tk.StringVar()
        self.reg_email = tk.StringVar()
        self.reg_code = tk.StringVar()

        reg_fields = [
            ("帳號", self.reg_username, None),
            ("密碼", self.reg_password, "*"),
            ("確認密碼", self.reg_confirm_password, "*"),
            ("手機號碼", self.reg_phone, None),
        ]
        for label, var, show in reg_fields:
            ttk.Label(reg_tab, text=label).pack(anchor="w", pady=(5, 0))
            ttk.Entry(reg_tab, textvariable=var, width=46, show=show).pack(anchor="w", fill="x", pady=(2, 0))

        ttk.Label(reg_tab, text="Email").pack(anchor="w", pady=(5, 0))
        email_row = ttk.Frame(reg_tab)
        email_row.pack(anchor="w", fill="x", pady=(2, 0))
        ttk.Entry(email_row, textvariable=self.reg_email, width=40).pack(side="left", fill="x", expand=True)
        ttk.Button(email_row, text="取得驗證碼", command=self.send_register_code).pack(side="left", padx=(8, 0))

        ttk.Label(reg_tab, text="驗證碼").pack(anchor="w", pady=(5, 0))
        ttk.Entry(reg_tab, textvariable=self.reg_code, width=46).pack(anchor="w", fill="x", pady=(2, 0))

        ttk.Label(reg_tab, text="註冊前請完整閱讀免責聲明", foreground="#8B0000", font=("Microsoft JhengHei", 10, "bold")).pack(anchor="w", pady=(10, 4))
        self.reg_agree_var = tk.BooleanVar(value=False)
        agree_btn_frame = ttk.Frame(reg_tab)
        agree_btn_frame.pack(anchor="w", pady=(0, 6))
        ttk.Button(agree_btn_frame, text="我同意，繼續註冊", command=self.accept_disclaimer).pack(side="left", padx=(0, 8))
        ttk.Button(agree_btn_frame, text="不同意，離開", command=self.decline_disclaimer).pack(side="left")
        self.reg_agree_status_var = tk.StringVar(value="目前狀態：尚未同意免責聲明")
        ttk.Label(reg_tab, textvariable=self.reg_agree_status_var, foreground="#8B0000").pack(anchor="w", pady=(0, 6))
        disclaimer_frame = ttk.Frame(reg_tab)
        disclaimer_frame.pack(fill="both", expand=False, pady=(0, 6))
        disclaimer_scroll = ttk.Scrollbar(disclaimer_frame, orient="vertical")
        disclaimer_scroll.pack(side="right", fill="y")
        self.reg_disclaimer_text = tk.Text(disclaimer_frame, height=10, wrap="word", yscrollcommand=disclaimer_scroll.set)
        self.reg_disclaimer_text.pack(side="left", fill="both", expand=True)
        disclaimer_scroll.config(command=self.reg_disclaimer_text.yview)
        self.reg_disclaimer_text.insert("1.0", DISCLAIMER_TEXT)
        self.reg_disclaimer_text.config(state="disabled")

        btn_frame = ttk.Frame(reg_tab)
        btn_frame.pack(anchor="center", pady=14)
        ttk.Button(btn_frame, text="完成註冊", command=self.verify_register_code).pack(side="left", padx=6)
        ttk.Label(reg_tab, text="註冊成功後，系統會自動開通 30 天免費試用。", foreground="#555555", wraplength=480).pack(anchor="w", pady=(4, 0))

        # reset password fields
        self.reset_account = tk.StringVar()
        self.reset_code = tk.StringVar()
        self.reset_new_password = tk.StringVar()
        self.reset_confirm_password = tk.StringVar()

        for label, var, show in [
            ("帳號或 Email", self.reset_account, None),
            ("驗證碼", self.reset_code, None),
            ("新密碼", self.reset_new_password, "*"),
            ("確認新密碼", self.reset_confirm_password, "*"),
        ]:
            ttk.Label(reset_tab, text=label).pack(anchor="w", pady=(6, 0))
            ttk.Entry(reset_tab, textvariable=var, width=46, show=show).pack(anchor="w", fill="x", pady=(2, 0))

        reset_btns = ttk.Frame(reset_tab)
        reset_btns.pack(anchor="center", pady=14)
        ttk.Button(reset_btns, text="寄送重設驗證碼", command=self.send_reset_code).pack(side="left", padx=6)
        ttk.Button(reset_btns, text="重設密碼", command=self.reset_password).pack(side="left", padx=6)
        ttk.Label(reset_tab, text="忘記密碼時，系統會將純數字驗證碼寄到註冊 Email。", foreground="#555555", wraplength=480).pack(anchor="w", pady=(4, 0))

        self.grab_set()

    def set_message(self, text, ok=False):
        self.msg_var.set(text)

    def _session_from_login(self, account, login_data):
        user = login_data.get("user", {}) or {}
        email = clean_text(user.get("email", ""))
        username = clean_text(user.get("username", ""))
        return {
            "login_account": account,
            "account": username or email or account,
            "email": email,
            "username": username,
            "phone": clean_text(user.get("phone") or user.get("mobile") or ""),
            "token": login_data.get("access_token", login_data.get("token", "")),
            "device_id": self.device_info.get("device_id", ""),
            "device_name": self.device_info.get("device_name", ""),
            "subscription_status": user.get("subscription_status", login_data.get("subscription_status", "trial")),
            "plan_type": user.get("plan_type", login_data.get("plan_type", "trial")),
            "trial_end_at": str(user.get("trial_end_at", login_data.get("trial_end_at", ""))),
            "subscription_end_at": str(user.get("subscription_end_at", login_data.get("subscription_end_at", ""))),
            "role": user.get("role", login_data.get("role", "")),
            "is_creator": user.get("is_creator", login_data.get("is_creator", False)),
            "is_admin": user.get("is_admin", login_data.get("is_admin", False)),
        }

    def do_login(self):
        payload = {
            "account": clean_text(self.login_account.get()),
            "password": self.login_password.get(),
        }
        if not payload['account'] or not payload['password']:
            self.set_message("請輸入帳號與密碼。")
            return
        try:
            data = auth_post("/auth/login", payload)
            if not data.get("success"):
                self.set_message(data.get("message", "登入失敗"))
                return
            session_payload = self._session_from_login(payload['account'], data)
            session_payload["remember_me"] = bool(self.login_remember_me.get())
            if self.login_remember_me.get():
                save_auth_session(session_payload)
            else:
                clear_auth_session()
            self.result = session_payload
            self.destroy()
        except Exception as e:
            self.set_message(f"登入失敗：{e}")

    def send_register_code(self):
        payload = {
            "username": clean_text(self.reg_username.get()),
            "password": self.reg_password.get(),
            "confirm_password": self.reg_confirm_password.get(),
            "phone": clean_text(self.reg_phone.get()),
            "email": clean_text(self.reg_email.get()).lower(),
            "device_id": self.device_info.get("device_id", ""),
            "device_name": self.device_info.get("device_name", ""),
        }
        if not payload["username"] or not payload["password"] or not payload["confirm_password"] or not payload["phone"] or not payload["email"]:
            self.set_message("請完整填寫帳號、密碼、確認密碼、手機號碼、Email。")
            return
        if not self.reg_agree_var.get():
            self.set_message("請先勾選『我已完整閱讀並同意上述免責聲明』後，再開始註冊程序。")
            return
        try:
            data = auth_post("/auth/send-register-code", payload)
            self.pending_register_payload = payload
            self.pending_register_email = payload["email"]
            msg = data.get("message", "驗證碼已寄出")
            if data.get("dev_code"):
                msg += f"（測試碼：{data.get('dev_code')}）"
            self.set_message(msg)
        except requests.HTTPError as e:
            try:
                detail = e.response.json().get("detail", str(e))
            except Exception:
                detail = str(e)
            self.set_message(f"寄送失敗：{detail}")
        except Exception as e:
            self.set_message(f"寄送失敗：{e}")

    def verify_register_code(self):
        email = clean_text(self.reg_email.get()).lower()
        code = clean_text(self.reg_code.get())
        if not email or not code:
            self.set_message("請輸入 Email 與驗證碼。")
            return
        if not self.reg_agree_var.get():
            self.set_message("請先勾選『我已完整閱讀並同意上述免責聲明』後，再完成註冊。")
            return
        try:
            data = auth_post("/auth/verify-register-code", {"email": email, "code": code})
            if not data.get("success"):
                self.set_message(data.get("message", "驗證失敗"))
                return
            login_data = auth_post("/auth/login", {
                "account": clean_text(self.reg_username.get()),
                "password": self.reg_password.get(),
            })
            if not login_data.get("success"):
                self.set_message(f"註冊成功，但自動登入失敗：{login_data.get('message', '')}")
                return
            session_payload = self._session_from_login(clean_text(self.reg_username.get()), login_data)
            session_payload["remember_me"] = False
            clear_auth_session()
            self.result = session_payload
            self.destroy()
        except requests.HTTPError as e:
            try:
                detail = e.response.json().get("detail", str(e))
            except Exception:
                detail = str(e)
            self.set_message(f"註冊失敗：{detail}")
        except Exception as e:
            self.set_message(f"註冊失敗：{e}")

    def accept_disclaimer(self):
        self.reg_agree_var.set(True)
        self.reg_agree_status_var.set("目前狀態：已同意免責聲明，可繼續註冊")
        self.set_message("您已同意免責聲明，現在可以發送驗證碼並完成註冊。")

    def decline_disclaimer(self):
        self.reg_agree_var.set(False)
        self.reg_agree_status_var.set("目前狀態：未同意免責聲明")
        if messagebox.askyesno("免責聲明", "您尚未同意免責聲明，是否離開登入視窗？"):
            self.result = None
            self.destroy()
        else:
            self.set_message("若要註冊，請先閱讀免責聲明並按『我同意，繼續註冊』。")

    def send_reset_code(self):
        account = clean_text(self.reset_account.get())
        if not account:
            self.set_message("請輸入帳號或 Email。")
            return
        try:
            data = auth_post("/auth/forgot-password", {"account": account})
            msg = data.get("message", "若帳號存在，重設驗證碼已寄送到 Email")
            if data.get("dev_code"):
                msg += f"（測試碼：{data.get('dev_code')}）"
            self.set_message(msg)
        except Exception as e:
            self.set_message(f"寄送失敗：{e}")

    def reset_password(self):
        payload = {
            "account": clean_text(self.reset_account.get()),
            "code": clean_text(self.reset_code.get()),
            "new_password": self.reset_new_password.get(),
            "confirm_password": self.reset_confirm_password.get(),
        }
        if not payload["account"] or not payload["code"] or not payload["new_password"] or not payload["confirm_password"]:
            self.set_message("請完整填寫帳號、驗證碼、新密碼、確認新密碼。")
            return
        try:
            data = auth_post("/auth/reset-password", payload)
            self.set_message(data.get("message", "密碼已重設，請重新登入"))
        except requests.HTTPError as e:
            try:
                detail = e.response.json().get("detail", str(e))
            except Exception:
                detail = str(e)
            self.set_message(f"重設失敗：{detail}")
        except Exception as e:
            self.set_message(f"重設失敗：{e}")

    def on_close(self):
        if messagebox.askyesno("離開", "尚未登入，是否直接關閉程式？", parent=self):
            self.result = None
            self.destroy()

# =========================
# GUI
# =========================
class StockApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ZHU STOCK APP｜正式授權版")
        self.geometry("1600x980")
        self.minsize(1280, 820)

        try:
            self.iconbitmap(resource_path(ICON_FILE))
        except Exception:
            pass

        ensure_output_dir()

        self.running = False
        self.auto_refresh_job = None
        self.latest_result = {}
        self.latest_excel_path = OUTPUT_XLSX
        self.chart_canvas = None
        self.user_holdings = load_holdings()
        self.loading_popup = None
        self.loading_img = None
        self.device_info = get_local_device_info()
        self.auth_session = load_auth_session()
        self.license_info = {}
        self.analysis_allowed = False
        self.user_email_var = tk.StringVar(value="帳號：尚未登入")
        self.license_var = tk.StringVar(value="授權：尚未驗證")
        self.payment_plan_var = tk.StringVar(value="monthly")
        self.payment_bank_var = tk.StringVar(value=PAYMENT_BANKS[0])
        self.payment_last5_var = tk.StringVar()
        self.payment_amount_var = tk.StringVar()
        self.payment_payer_name_var = tk.StringVar()
        self.payment_phone_var = tk.StringVar()
        self.payment_status_var = tk.StringVar(value="請先閱讀付款前聲明。")
        self.payment_agree_var = tk.BooleanVar(value=False)
        self.payment_consent_status_var = tk.StringVar(value="目前狀態：尚未同意付款前聲明，匯款帳號尚未顯示。")
        self.admin_payment_status_var = tk.StringVar(value="請按「重新整理」載入付款回報清單。")
        self.admin_payment_reports = []
        self.admin_payment_filtered_reports = []
        self.admin_seen_pending_report_ids = set()
        self.admin_payment_loaded_once = False
        self.admin_payment_filter_status_var = tk.StringVar(value="pending")
        self.admin_payment_filter_keyword_var = tk.StringVar()
        self.admin_payment_pending_only_var = tk.BooleanVar(value=True)
        self.admin_payment_summary_var = tk.StringVar(value="待審核：0｜已核准：0｜已駁回：0｜全部：0")
        self.selected_payment_report_id = None
        self.admin_users_status_var = tk.StringVar(value="請按「重新整理會員」載入註冊會員清單。")
        self.admin_users = []
        self.selected_admin_user_account = None
        self.selected_admin_user_record = {}
        self.admin_members_plan_var = tk.StringVar(value="monthly")
        self.admin_members_days_var = tk.StringVar(value="30")
        self.admin_members_end_var = tk.StringVar()
        self.admin_members_reason_var = tk.StringVar(value="活動贈送")
        self.admin_members_note_var = tk.StringVar()
        self.auth_profile = {}
        self.is_creator = False

        self.sheet_order = ["CLIENT_BULLISH", "CLIENT_BULLISH_KEYK", "CLIENT_BEARISH", "CLIENT_BEARISH_KEYK", "MY_HOLDINGS", "PAYMENT", "ADMIN_USERS", "ADMIN_PAYMENT_REVIEW"]

        self.tree_sort_state = {
            "CLIENT_BULLISH": {"column": None, "ascending": True},
            "CLIENT_BULLISH_KEYK": {"column": None, "ascending": True},
            "CLIENT_BEARISH": {"column": None, "ascending": True},
            "CLIENT_BEARISH_KEYK": {"column": None, "ascending": True},
            "MY_HOLDINGS": {"column": None, "ascending": True},
        }
        self.current_view_data = {
            "CLIENT_BULLISH": pd.DataFrame(),
            "CLIENT_BULLISH_KEYK": pd.DataFrame(),
            "CLIENT_BEARISH": pd.DataFrame(),
            "CLIENT_BEARISH_KEYK": pd.DataFrame(),
            "MY_HOLDINGS": pd.DataFrame(),
        }

        self.create_widgets()
        self.after(200, self.initialize_auth_flow)

    def create_widgets(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        self.run_btn = ttk.Button(top, text="立即執行分析", command=self.run_analysis_thread)
        self.run_btn.pack(side="left", padx=4)

        self.rebuild_btn = ttk.Button(top, text="強制重建快照", command=self.run_force_rebuild_thread)
        self.rebuild_btn.pack(side="left", padx=4)
        self.logout_btn = ttk.Button(top, text="登出 / 切換帳號", command=self.logout)
        self.logout_btn.pack(side="left", padx=4)
        self.payment_btn = ttk.Button(top, text="訂閱 / 付款", command=self.open_payment_tab)
        self.payment_btn.pack(side="left", padx=4)

        ttk.Label(top, text="自動刷新(分鐘)：").pack(side="left", padx=(20, 4))
        self.auto_var = tk.StringVar(value="30")
        self.auto_entry = ttk.Entry(top, textvariable=self.auto_var, width=8)
        self.auto_entry.pack(side="left", padx=4)

        self.auto_on_btn = ttk.Button(top, text="啟動自動刷新", command=self.start_auto_refresh)
        self.auto_on_btn.pack(side="left", padx=4)

        self.auto_off_btn = ttk.Button(top, text="停止自動刷新", command=self.stop_auto_refresh)
        self.auto_off_btn.pack(side="left", padx=4)

        self.status_var = tk.StringVar(value="狀態：待命中")
        ttk.Label(top, textvariable=self.status_var).pack(side="right", padx=8)
        ttk.Label(top, textvariable=self.license_var).pack(side="right", padx=8)
        ttk.Label(top, textvariable=self.user_email_var).pack(side="right", padx=8)

        info = ttk.Frame(self)
        info.pack(fill="x", padx=10, pady=(0, 6))

        self.settle_var = tk.StringVar(value="本次結算日：尚未執行")
        ttk.Label(info, textvariable=self.settle_var, font=("Microsoft JhengHei", 10, "bold")).pack(side="left", padx=4)

        self.summary_var = tk.StringVar(value="看多：0 檔｜多方關鍵K：0 檔｜看空：0 檔｜空方關鍵K：0 檔｜持股：0 檔")
        ttk.Label(info, textvariable=self.summary_var, font=("Microsoft JhengHei", 10, "bold")).pack(side="right", padx=4)

        self.main_pane = ttk.Panedwindow(self, orient="vertical")
        self.main_pane.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.upper_pane = ttk.Frame(self.main_pane)
        self.lower_pane = ttk.Panedwindow(self.main_pane, orient="horizontal")

        self.main_pane.add(self.upper_pane, weight=4)
        self.main_pane.add(self.lower_pane, weight=3)

        self.notebook = ttk.Notebook(self.upper_pane)
        self.notebook.pack(fill="both", expand=True)

        self.sheet_frames = {}
        self.sheet_trees = {}
        self.sheet_titles = {
            "CLIENT_BULLISH": "看多",
            "CLIENT_BULLISH_KEYK": "多方關鍵K",
            "CLIENT_BEARISH": "看空",
            "CLIENT_BEARISH_KEYK": "空方關鍵K",
            "MY_HOLDINGS": "我的持股",
            "PAYMENT": "訂閱 / 付款",
            "ADMIN_USERS": "會員管理",
            "ADMIN_PAYMENT_REVIEW": "付款審核",
        }

        for sheet, title in [
            ("CLIENT_BULLISH", "看多"),
            ("CLIENT_BULLISH_KEYK", "多方關鍵K"),
            ("CLIENT_BEARISH", "看空"),
            ("CLIENT_BEARISH_KEYK", "空方關鍵K"),
            ("MY_HOLDINGS", "我的持股"),
            ("PAYMENT", "訂閱 / 付款"),
            ("ADMIN_USERS", "會員管理"),
            ("ADMIN_PAYMENT_REVIEW", "付款審核"),
        ]:
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=title)
            self.sheet_frames[sheet] = frame

            if sheet == "PAYMENT":
                self.build_payment_tab(frame)
                continue

            if sheet == "ADMIN_USERS":
                self.build_admin_users_tab(frame)
                continue

            if sheet == "ADMIN_PAYMENT_REVIEW":
                self.build_admin_payment_review_tab(frame)
                continue

            if sheet == "MY_HOLDINGS":
                hold_bar = ttk.Frame(frame)
                hold_bar.pack(fill="x", padx=6, pady=6)

                ttk.Label(hold_bar, text="股票代號：").pack(side="left")
                self.holding_code_var = tk.StringVar()
                self.holding_entry = ttk.Entry(hold_bar, textvariable=self.holding_code_var, width=12)
                self.holding_entry.pack(side="left", padx=4)

                ttk.Button(hold_bar, text="加入持股", command=self.add_holding_code).pack(side="left", padx=4)
                ttk.Button(hold_bar, text="刪除選取持股", command=self.remove_selected_holding).pack(side="left", padx=4)
                ttk.Button(hold_bar, text="重新整理持股", command=self.refresh_holdings_view).pack(side="left", padx=4)

            tree_frame = ttk.Frame(frame)
            tree_frame.pack(fill="both", expand=True)

            tree = ttk.Treeview(tree_frame, show="headings")
            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")

            tree_frame.rowconfigure(0, weight=1)
            tree_frame.columnconfigure(0, weight=1)

            tree.bind("<<TreeviewSelect>>", self.on_tree_select)

            self.sheet_trees[sheet] = tree

        chart_frame = ttk.LabelFrame(self.lower_pane, text="週K線圖")
        self.lower_pane.add(chart_frame, weight=3)

        self.chart_title_var = tk.StringVar(value="請先執行分析，然後點選一檔標的")
        ttk.Label(chart_frame, textvariable=self.chart_title_var, font=("Microsoft JhengHei", 10, "bold")).pack(anchor="w", padx=8, pady=6)

        self.chart_container = ttk.Frame(chart_frame)
        self.chart_container.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.right_pane = ttk.Panedwindow(self.lower_pane, orient="vertical")
        self.lower_pane.add(self.right_pane, weight=2)

        detail_frame = ttk.LabelFrame(self.right_pane, text="標的資訊")
        self.right_pane.add(detail_frame, weight=2)

        self.detail_notebook = ttk.Notebook(detail_frame)
        self.detail_notebook.pack(fill="both", expand=True, padx=8, pady=8)

        tech_tab = ttk.Frame(self.detail_notebook)
        company_tab = ttk.Frame(self.detail_notebook)

        self.detail_notebook.add(tech_tab, text="技術資訊")
        self.detail_notebook.add(company_tab, text="公司基本資訊")

        self.detail_text = tk.Text(tech_tab, height=16, wrap="word")
        self.detail_text.pack(fill="both", expand=True)

        self.company_text = tk.Text(company_tab, height=16, wrap="word")
        self.company_text.pack(fill="both", expand=True)

        self.detail_text.config(state="disabled")
        self.company_text.config(state="disabled")

        log_frame = ttk.LabelFrame(self.right_pane, text="執行紀錄")
        self.right_pane.add(log_frame, weight=2)

        self.log_text = tk.Text(log_frame, height=12, wrap="word", state="disabled")
        self.log_text.pack(fill="both", expand=True, padx=8, pady=8)

        self.notebook.bind("<<NotebookTabChanged>>", self.on_notebook_tab_changed)
        self.update_context_panels_by_tab()

    def _get_current_sheet_key(self):
        try:
            current_tab = self.notebook.select()
        except Exception:
            return ""
        for key, frame in self.sheet_frames.items():
            try:
                if str(frame) == str(current_tab):
                    return key
            except Exception:
                pass
        return ""

    def _set_context_panels_visible(self, visible=True):
        if not hasattr(self, "main_pane") or not hasattr(self, "lower_pane"):
            return
        try:
            panes = list(self.main_pane.panes())
        except Exception:
            panes = []

        lower_str = str(self.lower_pane)
        exists = lower_str in panes

        if visible and not exists:
            try:
                self.main_pane.add(self.lower_pane, weight=3)
            except Exception:
                pass
        elif (not visible) and exists:
            try:
                self.main_pane.forget(self.lower_pane)
            except Exception:
                pass

    def update_context_panels_by_tab(self, event=None):
        current_sheet = self._get_current_sheet_key()
        hide_for = {"PAYMENT", "ADMIN_USERS", "ADMIN_PAYMENT_REVIEW"}
        self._set_context_panels_visible(current_sheet not in hide_for)

    def on_notebook_tab_changed(self, event=None):
        self.update_context_panels_by_tab()

    def build_payment_tab(self, frame):
        container = ttk.Frame(frame)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, highlightthickness=0)
        vscroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        outer = ttk.Frame(canvas, padding=18)
        canvas_window = canvas.create_window((0, 0), window=outer, anchor="nw")

        def _sync_scrollregion(event=None):
            try:
                canvas.configure(scrollregion=canvas.bbox("all"))
            except Exception:
                pass

        def _resize_inner(event):
            try:
                canvas.itemconfigure(canvas_window, width=event.width)
            except Exception:
                pass

        def _on_mousewheel(event):
            try:
                if event.delta:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                elif getattr(event, "num", None) == 4:
                    canvas.yview_scroll(-1, "units")
                elif getattr(event, "num", None) == 5:
                    canvas.yview_scroll(1, "units")
            except Exception:
                pass

        outer.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<Configure>", _resize_inner)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", _on_mousewheel)
        canvas.bind_all("<Button-5>", _on_mousewheel)

        ttk.Label(outer, text="訂閱 / 付款資訊", font=("Microsoft JhengHei", 14, "bold")).pack(anchor="w", pady=(0, 10))
        ttk.Label(
            outer,
            text="請先閱讀以下付款前聲明。只有在您按下『我同意，顯示匯款資訊』後，系統才會顯示方案、匯款帳號與匯款回報欄位。",
            justify="left",
            wraplength=1100,
            foreground="#555555",
        ).pack(anchor="w", pady=(0, 10))

        notice_box = ttk.LabelFrame(outer, text="付款前聲明")
        notice_box.pack(fill="x", expand=False, pady=(0, 12))

        notice_scroll_wrap = ttk.Frame(notice_box)
        notice_scroll_wrap.pack(fill="both", expand=True, padx=10, pady=(10, 8))
        notice_scroll = ttk.Scrollbar(notice_scroll_wrap, orient="vertical")
        notice_scroll.pack(side="right", fill="y")
        self.payment_notice_text = tk.Text(notice_scroll_wrap, height=12, wrap="word", yscrollcommand=notice_scroll.set)
        self.payment_notice_text.pack(side="left", fill="both", expand=True)
        notice_scroll.config(command=self.payment_notice_text.yview)
        self.payment_notice_text.insert("1.0", PAYMENT_NOTICE_TEXT)
        self.payment_notice_text.config(state="disabled")

        consent_row = ttk.Frame(notice_box)
        consent_row.pack(anchor="w", padx=10, pady=(0, 8))
        ttk.Button(consent_row, text="我同意，顯示匯款資訊", command=self.accept_payment_notice).pack(side="left", padx=(0, 8))
        ttk.Button(consent_row, text="不同意，不顯示帳號", command=self.decline_payment_notice).pack(side="left")
        ttk.Label(notice_box, textvariable=self.payment_consent_status_var, foreground="#8B0000", wraplength=1100).pack(anchor="w", padx=10, pady=(0, 10))

        self.payment_submit_box = ttk.LabelFrame(outer, text="匯款完成後請填寫")
        form = ttk.Frame(self.payment_submit_box)
        form.pack(anchor="w", padx=18, pady=16)

        plan_values = list(PAYMENT_PLAN_LABELS.values())
        plan_width = 22
        bank_width = 32
        entry_width = 22

        ttk.Label(form, text="訂閱方案", font=("Microsoft JhengHei", 10, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        self.payment_plan_combo = ttk.Combobox(
            form,
            textvariable=self.payment_plan_var,
            state="readonly",
            width=plan_width,
            values=plan_values,
        )
        self.payment_plan_combo.grid(row=0, column=1, sticky="w", pady=6)
        self.payment_plan_combo.bind("<<ComboboxSelected>>", self._on_payment_plan_combo_change)

        ttk.Label(form, text="匯款銀行", font=("Microsoft JhengHei", 10, "bold")).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        self.payment_bank_combo = ttk.Combobox(
            form,
            textvariable=self.payment_bank_var,
            state="readonly",
            width=bank_width,
            values=PAYMENT_BANKS,
        )
        self.payment_bank_combo.grid(row=1, column=1, sticky="w", pady=6)

        ttk.Label(form, text="匯款人姓名 / 暱稱", font=("Microsoft JhengHei", 10, "bold")).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(form, textvariable=self.payment_payer_name_var, width=entry_width).grid(row=2, column=1, sticky="w", pady=6)

        ttk.Label(form, text="匯款金額", font=("Microsoft JhengHei", 10, "bold")).grid(row=3, column=0, sticky="w", padx=(0, 10), pady=6)
        self.payment_amount_entry = ttk.Entry(form, textvariable=self.payment_amount_var, width=entry_width, state="readonly")
        self.payment_amount_entry.grid(row=3, column=1, sticky="w", pady=6)

        ttk.Label(form, text="匯款後末五碼", font=("Microsoft JhengHei", 10, "bold")).grid(row=4, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(form, textvariable=self.payment_last5_var, width=entry_width).grid(row=4, column=1, sticky="w", pady=6)

        action_row = ttk.Frame(self.payment_submit_box)
        action_row.pack(anchor="w", padx=18, pady=(0, 8))
        ttk.Button(action_row, text="確認送出", command=self.submit_payment_report, width=12).pack(side="left")
        ttk.Button(action_row, text="清除重填", command=self.reset_payment_form, width=12).pack(side="left", padx=(8, 0))

        ttk.Label(
            self.payment_submit_box,
            text="提醒：請先選擇方案與匯款帳號，完成匯款後再填寫匯款人與末五碼後送出。資料送出後會進入管理端付款審核頁面，由管理員人工核對。",
            foreground="#1F5E2E",
            wraplength=1100,
        ).pack(anchor="w", padx=18, pady=(0, 8))

        ttk.Label(outer, textvariable=self.payment_status_var, foreground="#1F5E2E", wraplength=1100).pack(anchor="w", pady=(6, 0))
        self.update_payment_notice_visibility()
        _sync_scrollregion()

    def update_payment_notice_visibility(self):
        show_accounts = bool(getattr(self, "payment_agree_var", tk.BooleanVar(value=False)).get())
        submit_box = getattr(self, "payment_submit_box", None)
        if show_accounts:
            if submit_box is not None and not submit_box.winfo_ismapped():
                submit_box.pack(fill="x", expand=False, pady=(0, 12))
        else:
            if submit_box is not None and submit_box.winfo_ismapped():
                submit_box.pack_forget()

    def accept_payment_notice(self):
        self.payment_agree_var.set(True)
        self.payment_consent_status_var.set("目前狀態：已同意付款前聲明，匯款資訊已顯示。")
        self.payment_status_var.set("您已同意付款前聲明，請先選擇方案、匯款，再填寫末五碼與手機驗證後送出匯款資料。")
        self.update_payment_notice_visibility()
        self.show_payment_info_popup()

    def decline_payment_notice(self):
        self.payment_agree_var.set(False)
        self.payment_consent_status_var.set("目前狀態：未同意付款前聲明，匯款資訊不顯示。")
        self.payment_status_var.set("您尚未同意付款前聲明，因此系統不顯示方案、匯款帳號與送出欄位。")
        self.update_payment_notice_visibility()

    def _normalize_phone_digits(self, value):
        return "".join(ch for ch in clean_text(value) if ch.isdigit())

    def _get_current_bound_phone(self):
        candidates = []
        for payload in [self.auth_profile, self.license_info, self.auth_session]:
            if isinstance(payload, dict):
                user_obj = payload.get("user") if isinstance(payload.get("user"), dict) else {}
                candidates.extend([
                    payload.get("phone"), payload.get("mobile"), payload.get("tel"),
                    user_obj.get("phone"), user_obj.get("mobile"), user_obj.get("tel"),
                ])
        for value in candidates:
            digits = self._normalize_phone_digits(value)
            if digits:
                return digits
        token = clean_text((self.auth_session or {}).get("token"))
        if token:
            try:
                profile = auth_get("/auth/me", headers=self._auth_headers())
                if isinstance(profile, dict):
                    self.auth_profile = profile
                    user_obj = profile.get("user") if isinstance(profile.get("user"), dict) else {}
                    digits = self._normalize_phone_digits(profile.get("phone") or profile.get("mobile") or user_obj.get("phone") or user_obj.get("mobile") or "")
                    if digits:
                        return digits
            except Exception:
                pass
        return ""

    def reset_payment_form(self):
        self.payment_plan_var.set(PAYMENT_PLAN_LABELS.get("monthly", "月訂閱（2888／月）"))
        self.payment_bank_var.set(PAYMENT_BANKS[0] if PAYMENT_BANKS else "")
        self.payment_last5_var.set("")
        self.payment_amount_var.set("")
        self.payment_payer_name_var.set("")
        self.payment_phone_var.set(clean_text((self.auth_session or {}).get("phone") or ""))
        self.payment_status_var.set("匯款回報欄位已清空。")
        self._on_payment_plan_combo_change()

    def show_payment_info_popup(self):
        try:
            win = tk.Toplevel(self)
            win.title("付款資訊提醒")
            win.geometry("760x520")
            win.transient(self)
            win.grab_set()
            wrap = ttk.Frame(win, padding=14)
            wrap.pack(fill="both", expand=True)
            ttk.Label(wrap, text="您已同意付款前聲明", font=("Microsoft JhengHei", 13, "bold")).pack(anchor="w", pady=(0, 8))
            ttk.Label(wrap, text="請依序完成：選方案 → 匯款 → 填寫末五碼 → 驗證手機 → 送出資料", foreground="#1F5E2E").pack(anchor="w", pady=(0, 8))
            box = tk.Text(wrap, wrap="word", height=20)
            box.pack(fill="both", expand=True)
            summary = [
                "【訂閱方案】",
                *[f"- {label}" for label in PAYMENT_PLAN_LABELS.values()],
                "",
                "【匯款帳號】",
                *[f"- {bank}" for bank in PAYMENT_BANKS],
                "",
                "【送出前需填寫】",
                "- 訂閱方案",
                "- 匯款銀行",
                "- 匯款人姓名 / 暱稱",
                "- 匯款金額",
                "- 匯款後末五碼",
                "- 手機號碼並完成驗證",
                "",
                "送出後資料會進入管理員的付款審核頁面，供人工比對。",
            ]
            box.insert("1.0", "\n".join(summary))
            box.config(state="disabled")
            ttk.Button(wrap, text="我知道了", command=win.destroy).pack(anchor="e", pady=(10, 0))
        except Exception:
            pass

    def submit_payment_report(self):
        plan = clean_text(self.payment_plan_var.get())
        reverse_plan_map = {v: k for k, v in PAYMENT_PLAN_LABELS.items()}
        plan_key = reverse_plan_map.get(plan, plan)

        bank = clean_text(self.payment_bank_var.get())
        last5 = clean_text(self.payment_last5_var.get())
        payer_name = clean_text(self.payment_payer_name_var.get())
        amount = clean_text(self.payment_amount_var.get())

        if not plan_key or plan_key not in PAYMENT_PLAN_LABELS:
            messagebox.showwarning("提醒", "請選擇訂閱方案。")
            return
        if not bank:
            messagebox.showwarning("提醒", "請選擇匯款銀行。")
            return
        if not payer_name:
            messagebox.showwarning("提醒", "請填寫匯款人姓名 / 暱稱。")
            return
        if not amount:
            messagebox.showwarning("提醒", "尚未帶入匯款金額，請重新選擇方案。")
            return
        if not last5 or len(last5) < 5:
            messagebox.showwarning("提醒", "請填寫匯款後末五碼。")
            return

        payload = {
            "report_id": str(uuid.uuid4()),
            "account": clean_text((self.auth_session or {}).get("account") or ""),
            "email": clean_text((self.auth_session or {}).get("email") or ""),
            "plan": plan_key,
            "plan_label": PAYMENT_PLAN_LABELS.get(plan_key, plan),
            "bank": bank,
            "payer_name": payer_name,
            "amount": amount,
            "last5": last5,
            "status": "pending",
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }

        reports = load_payment_reports()
        reports.append(payload)
        save_payment_reports(reports)

        self.payment_status_var.set("匯款資料已送出，請等待管理員人工審核。")
        messagebox.showinfo("完成", "匯款資料已送出，請等待管理員人工審核。")
        self.refresh_admin_payment_review_tab_if_possible()
        self.reset_payment_form()

    def build_admin_users_tab(self, frame):
        outer = ttk.Frame(frame, padding=12)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="管理員會員總覽", font=("Microsoft JhengHei", 13, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Label(
            outer,
            text="此頁可查看所有註冊會員的訂閱狀態、待審核情況，並可手動開通方案、活動贈送天數，以及刪除會員以避免殭屍帳號累積。",
            foreground="#555555",
            wraplength=980,
        ).pack(anchor="w", pady=(0, 8))

        ctrl = ttk.LabelFrame(outer, text="手動開通 / 活動贈送")
        ctrl.pack(fill="x", pady=(0, 8))

        row1 = ttk.Frame(ctrl)
        row1.pack(fill="x", padx=8, pady=(8, 4))
        ttk.Button(row1, text="重新整理會員", command=self.load_admin_users).pack(side="left")
        ttk.Button(row1, text="重新整理付款待審核", command=self.load_admin_overview).pack(side="left", padx=6)
        ttk.Button(row1, text="開啟付款審核頁", command=lambda: self.notebook.select(self.sheet_frames.get("ADMIN_PAYMENT_REVIEW"))).pack(side="left", padx=6)
        ttk.Button(row1, text="匯出Excel", command=self.export_admin_overview_excel).pack(side="left", padx=6)
        ttk.Button(row1, text="刪除選取會員", command=self.delete_selected_user).pack(side="left", padx=6)

        row2 = ttk.Frame(ctrl)
        row2.pack(fill="x", padx=8, pady=(4, 8))
        ttk.Label(row2, text="方案：").pack(side="left")
        self.admin_members_plan_combo = ttk.Combobox(
            row2,
            textvariable=self.admin_members_plan_var,
            state="readonly",
            values=["monthly", "quarterly", "yearly", "trial", "free_grant", "none"],
            width=14,
        )
        self.admin_members_plan_combo.pack(side="left", padx=(0, 10))
        ttk.Label(row2, text="天數：").pack(side="left")
        ttk.Entry(row2, textvariable=self.admin_members_days_var, width=8).pack(side="left", padx=(0, 10))
        ttk.Label(row2, text="到期日(YYYY-MM-DD，可留空)：").pack(side="left")
        ttk.Entry(row2, textvariable=self.admin_members_end_var, width=16).pack(side="left", padx=(0, 10))
        ttk.Button(row2, text="手動開通選取會員", command=self.manual_set_plan_selected_user).pack(side="left", padx=4)
        ttk.Button(row2, text="活動贈送選取會員", command=self.grant_free_selected_user).pack(side="left", padx=4)

        row3 = ttk.Frame(ctrl)
        row3.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(row3, text="贈送原因：").pack(side="left")
        self.admin_members_reason_combo = ttk.Combobox(
            row3,
            textvariable=self.admin_members_reason_var,
            state="readonly",
            values=["活動贈送", "客服補償", "新會員體驗", "人工補發", "VIP 回饋", "其他"],
            width=16,
        )
        self.admin_members_reason_combo.pack(side="left", padx=(0, 10))
        ttk.Label(row3, text="備註：").pack(side="left")
        ttk.Entry(row3, textvariable=self.admin_members_note_var, width=56).pack(side="left", padx=(0, 10), fill="x", expand=True)

        columns = ("account", "email", "role", "subscription_status", "plan_type", "subscription_end_at", "trial_end_at", "pending_review", "phone")
        tree_frame = ttk.Frame(outer)
        tree_frame.pack(fill="both", expand=True)

        self.admin_users_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        self.admin_users_tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.admin_users_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.admin_users_tree.xview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.admin_users_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        widths = {
            "account": 160, "email": 220, "role": 100, "subscription_status": 120,
            "plan_type": 100, "subscription_end_at": 140, "trial_end_at": 140,
            "pending_review": 100, "phone": 130,
        }
        headings = {
            "account": "帳號", "email": "Email", "role": "角色", "subscription_status": "訂閱狀態",
            "plan_type": "方案", "subscription_end_at": "訂閱到期", "trial_end_at": "試用到期",
            "pending_review": "待審核", "phone": "手機",
        }
        for col in columns:
            self.admin_users_tree.heading(col, text=headings.get(col, col))
            self.admin_users_tree.column(col, width=widths.get(col, 120), anchor="center", stretch=True)

        self.admin_users_tree.bind("<<TreeviewSelect>>", self.on_admin_user_select)
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        detail_box = ttk.LabelFrame(outer, text="會員明細")
        detail_box.pack(fill="both", expand=False, pady=(8, 0))
        self.admin_user_detail_text = tk.Text(detail_box, height=11, wrap="word", state="disabled")
        self.admin_user_detail_text.pack(fill="both", expand=True, padx=8, pady=8)

        ttk.Label(outer, textvariable=self.admin_users_status_var, foreground="#1F5E2E", wraplength=980).pack(anchor="w", pady=(8, 0))

    def build_admin_payment_review_tab(self, frame):
        outer = ttk.Frame(frame, padding=12)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="管理員付款審核", font=("Microsoft JhengHei", 13, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Label(outer, text="此頁可查看付款回報清單，並執行核准或駁回。", foreground="#555555").pack(anchor="w", pady=(0, 8))

        action_bar = ttk.Frame(outer)
        action_bar.pack(fill="x", pady=(0, 6))
        ttk.Button(action_bar, text="重新整理", command=self.load_admin_payment_reports).pack(side="left")
        ttk.Button(action_bar, text="核准選取", command=self.approve_selected_payment_report).pack(side="left", padx=6)
        ttk.Button(action_bar, text="駁回選取", command=self.reject_selected_payment_report).pack(side="left", padx=6)

        filter_bar = ttk.Frame(outer)
        filter_bar.pack(fill="x", pady=(0, 8))
        ttk.Label(filter_bar, text="狀態篩選：").pack(side="left")
        self.admin_payment_status_combo = ttk.Combobox(
            filter_bar,
            textvariable=self.admin_payment_filter_status_var,
            state="readonly",
            width=14,
            values=["pending", "approved", "rejected", "all"],
        )
        self.admin_payment_status_combo.pack(side="left")
        self.admin_payment_status_combo.bind("<<ComboboxSelected>>", self.on_admin_payment_filter_change)
        ttk.Checkbutton(
            filter_bar,
            text="只看待審核",
            variable=self.admin_payment_pending_only_var,
            command=self.on_admin_payment_filter_change,
        ).pack(side="left", padx=(8, 0))
        ttk.Label(filter_bar, text="關鍵字：").pack(side="left", padx=(12, 0))
        keyword_entry = ttk.Entry(filter_bar, textvariable=self.admin_payment_filter_keyword_var, width=28)
        keyword_entry.pack(side="left", padx=(0, 6))
        keyword_entry.bind("<KeyRelease>", self.on_admin_payment_filter_change)
        ttk.Button(filter_bar, text="套用篩選", command=self.apply_admin_payment_filters).pack(side="left")
        ttk.Button(filter_bar, text="清除篩選", command=self.clear_admin_payment_filters).pack(side="left", padx=6)
        ttk.Label(filter_bar, textvariable=self.admin_payment_summary_var, foreground="#1F5E2E").pack(side="right")

        columns = ("report_id", "account", "payer_name", "plan_type", "amount", "transfer_last5", "bank", "status", "created_at")
        tree_frame = ttk.Frame(outer)
        tree_frame.pack(fill="both", expand=True)

        self.admin_payment_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="browse")
        self.admin_payment_tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.admin_payment_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.admin_payment_tree.xview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.admin_payment_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        widths = {
            "report_id": 90, "account": 160, "payer_name": 120, "plan_type": 100,
            "amount": 90, "transfer_last5": 90, "bank": 220, "status": 110, "created_at": 180
        }
        headings = {
            "report_id": "報表ID", "account": "帳號", "payer_name": "付款人", "plan_type": "方案",
            "amount": "金額", "transfer_last5": "末五碼", "bank": "銀行", "status": "狀態", "created_at": "建立時間"
        }
        for col in columns:
            self.admin_payment_tree.heading(col, text=headings.get(col, col))
            anchor = "e" if col == "amount" else "center"
            self.admin_payment_tree.column(col, width=widths.get(col, 120), anchor=anchor, stretch=True)

        self.admin_payment_tree.bind("<<TreeviewSelect>>", self.on_admin_payment_select)

        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        detail_box = ttk.LabelFrame(outer, text="付款回報明細")
        detail_box.pack(fill="both", expand=False, pady=(8, 0))

        self.admin_payment_detail_text = tk.Text(detail_box, height=10, wrap="word", state="disabled")
        self.admin_payment_detail_text.pack(fill="both", expand=True, padx=8, pady=8)

        ttk.Label(outer, textvariable=self.admin_payment_status_var, foreground="#1F5E2E", wraplength=900).pack(anchor="w", pady=(8, 0))

    def hide_admin_users_tab(self):
        frame = self.sheet_frames.get("ADMIN_USERS")
        if frame is None:
            return
        try:
            self.notebook.hide(frame)
        except Exception:
            pass

    def show_admin_users_tab(self):
        frame = self.sheet_frames.get("ADMIN_USERS")
        if frame is None:
            return
        try:
            self.notebook.index(frame)
            return
        except Exception:
            pass
        try:
            self.notebook.add(frame, text=self.sheet_titles.get("ADMIN_USERS", "會員管理"))
        except Exception:
            pass

    def _normalize_admin_user_item(self, item):
        if not isinstance(item, dict):
            return {}
        account = clean_text(item.get("username") or item.get("account") or item.get("email") or item.get("user_id") or "")
        email = clean_text(item.get("email") or "")
        role = clean_text(item.get("role") or ("creator" if self._coerce_bool(item.get("is_creator")) else ("admin" if self._coerce_bool(item.get("is_admin")) else "user")))
        sub = clean_text(item.get("subscription_status") or item.get("status") or "none")
        plan = clean_text(item.get("plan_type") or item.get("plan") or "")
        return {
            "account": account,
            "email": email,
            "role": role,
            "phone": clean_text(item.get("phone") or item.get("mobile") or ""),
            "subscription_status": sub,
            "plan_type": plan,
            "trial_end_at": clean_text(str(item.get("trial_end_at") or "")),
            "subscription_end_at": clean_text(str(item.get("subscription_end_at") or item.get("end_at") or "")),
            "is_creator": self._coerce_bool(item.get("is_creator")),
            "is_admin": self._coerce_bool(item.get("is_admin")),
            "raw": item,
            "pending_review": 0,
        }

    def _extract_admin_users(self, data):
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for key in ["items", "data", "users", "results", "members"]:
                value = data.get(key)
                if isinstance(value, list):
                    return value
        return []

    def _merge_pending_reviews_into_users(self, rows):
        pending_map = {}
        try:
            data = auth_get("/admin/payment-reports", headers=self._auth_headers())
            reports = [self._normalize_payment_report_item(x) for x in self._extract_payment_reports(data)]
            self.admin_payment_reports = reports
            for r in reports:
                if clean_text(r.get("status", "pending")).lower() not in {"approved", "rejected", "done", "success"}:
                    acc = clean_text(r.get("account", ""))
                    if acc:
                        pending_map[acc] = pending_map.get(acc, 0) + 1
        except Exception:
            pass
        for row in rows:
            acc = clean_text(row.get("account", ""))
            email = clean_text(row.get("email", ""))
            row["pending_review"] = pending_map.get(acc, pending_map.get(email, 0))
        return rows

    def load_admin_overview(self):
        self.load_admin_users()
        try:
            self.load_admin_payment_reports()
        except Exception:
            pass

    def export_admin_overview_excel(self):
        if not self.is_creator:
            messagebox.showwarning("提醒", "只有創作者帳號可以匯出會員資料。")
            return
        try:
            if not self.admin_users:
                self.load_admin_users()
            if not self.admin_payment_reports:
                try:
                    self.load_admin_payment_reports()
                except Exception:
                    pass

            rows = list(self.admin_users or [])
            payment_rows = list(getattr(self, "admin_payment_reports", []) or [])
            if not rows and not payment_rows:
                messagebox.showwarning("提醒", "目前沒有可匯出的會員或付款資料。")
                return

            os.makedirs(OUTPUT_DIR, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_path = os.path.join(OUTPUT_DIR, f"會員總覽匯出_{ts}.xlsx")

            member_records = []
            for row in rows:
                raw = row.get("raw", {}) if isinstance(row.get("raw"), dict) else {}
                member_records.append({
                    "帳號": row.get("account", ""),
                    "Email": row.get("email", ""),
                    "角色": row.get("role", ""),
                    "訂閱狀態": row.get("subscription_status", ""),
                    "方案": PAYMENT_PLAN_LABELS.get(row.get("plan_type", ""), row.get("plan_type", "")),
                    "方案代碼": row.get("plan_type", ""),
                    "訂閱到期": (row.get("subscription_end_at", "") or "")[:19],
                    "試用到期": (row.get("trial_end_at", "") or "")[:19],
                    "待審核付款": row.get("pending_review", 0),
                    "手機": row.get("phone", ""),
                    "建立時間": clean_text(raw.get("created_at", "")),
                    "更新時間": clean_text(raw.get("updated_at", "")),
                    "裝置ID": clean_text(raw.get("device_id", "")),
                    "裝置名稱": clean_text(raw.get("device_name", "")),
                })

            payment_records = []
            for row in payment_rows:
                payment_records.append({
                    "報表ID": row.get("report_id", ""),
                    "帳號": row.get("account", ""),
                    "付款人": row.get("payer_name", ""),
                    "方案": row.get("plan_label", row.get("plan_type", "")),
                    "方案代碼": row.get("plan_type", ""),
                    "金額": row.get("amount", 0),
                    "末五碼": row.get("transfer_last5", ""),
                    "銀行": row.get("bank", ""),
                    "狀態": row.get("status", ""),
                    "建立時間": row.get("created_at", ""),
                    "備註": row.get("note", ""),
                })

            summary_rows = [{
                "匯出時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "匯出者": clean_text((self.auth_session or {}).get("username") or (self.auth_session or {}).get("account") or "admin"),
                "會員總數": len(member_records),
                "待審核付款總數": sum(int(x.get("待審核付款") or 0) for x in member_records),
                "付款回報總數": len(payment_records),
                "待審核回報": sum(1 for x in payment_records if clean_text(x.get("狀態", "")).lower() == "pending"),
                "已核准回報": sum(1 for x in payment_records if clean_text(x.get("狀態", "")).lower() == "approved"),
                "已駁回回報": sum(1 for x in payment_records if clean_text(x.get("狀態", "")).lower() == "rejected"),
            }]

            with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name="匯出摘要", index=False)
                pd.DataFrame(member_records).to_excel(writer, sheet_name="會員總覽", index=False)
                pd.DataFrame(payment_records).to_excel(writer, sheet_name="付款回報", index=False)

                for sheet_name, df in {
                    "匯出摘要": pd.DataFrame(summary_rows),
                    "會員總覽": pd.DataFrame(member_records),
                    "付款回報": pd.DataFrame(payment_records),
                }.items():
                    ws = writer.book[sheet_name]
                    ws.freeze_panes = "A2"
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    for row_cells in ws.iter_rows(min_row=2):
                        for cell in row_cells:
                            cell.alignment = Alignment(vertical="center")
                    for idx, col_name in enumerate(df.columns, start=1):
                        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].fillna("").tolist()]) if not df.empty else len(str(col_name))
                        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 36)

            self.log(f"管理員匯出會員總覽 Excel：{export_path}")
            messagebox.showinfo("成功", "會員總覽已匯出：\n" + str(export_path))
        except Exception as e:
            messagebox.showerror("匯出Excel失敗", str(e))

    def load_admin_users(self):
        if not self.is_creator:
            self.admin_users_status_var.set("載入失敗：非創作者權限")
            return
        try:
            data = auth_get("/admin/users", headers=self._auth_headers())
            rows = [self._normalize_admin_user_item(x) for x in self._extract_admin_users(data)]
            rows = [x for x in rows if x.get("account")]
            rows = self._merge_pending_reviews_into_users(rows)
            self.admin_users = rows
            self.refresh_admin_users_tree()
            pending_total = sum(int(x.get("pending_review") or 0) for x in rows)
            self.admin_users_status_var.set(f"已載入 {len(rows)} 位會員，待審核付款 {pending_total} 筆。")
        except Exception as e:
            self.admin_users_status_var.set(f"載入失敗：{e}")
            self.log(f"載入會員失敗：{e}")

    def refresh_admin_users_tree(self):
        tree = getattr(self, "admin_users_tree", None)
        if tree is None:
            return
        for iid in tree.get_children():
            tree.delete(iid)
        self.selected_admin_user_account = None
        self.selected_admin_user_record = {}
        for row in self.admin_users:
            values = (
                row.get("account", ""), row.get("email", ""), row.get("role", ""),
                row.get("subscription_status", ""), PAYMENT_PLAN_LABELS.get(row.get("plan_type", ""), row.get("plan_type", "")),
                (row.get("subscription_end_at", "") or "")[:10], (row.get("trial_end_at", "") or "")[:10],
                row.get("pending_review", 0), row.get("phone", ""),
            )
            iid = row.get("account") or row.get("email")
            tree.insert("", "end", iid=iid, values=values)
        self.set_readonly_text(getattr(self, "admin_user_detail_text", self.detail_text), "請先選取一位會員。")

    def on_admin_user_select(self, event=None):
        tree = getattr(self, "admin_users_tree", None)
        if tree is None:
            return
        sel = tree.selection()
        if not sel:
            self.selected_admin_user_account = None
            self.selected_admin_user_record = {}
            return
        iid = sel[0]
        self.selected_admin_user_account = iid
        row = next((x for x in self.admin_users if str(x.get("account", "")) == str(iid) or str(x.get("email", "")) == str(iid)), None)
        self.selected_admin_user_record = row or {}
        if not row:
            return
        raw = row.get("raw", {}) if isinstance(row.get("raw"), dict) else {}
        detail = [
            f"帳號：{row.get('account','-')}",
            f"Email：{row.get('email','-')}",
            f"角色：{row.get('role','-')}",
            f"手機：{row.get('phone','-')}",
            f"訂閱狀態：{row.get('subscription_status','-')}",
            f"方案：{PAYMENT_PLAN_LABELS.get(row.get('plan_type',''), row.get('plan_type','-') or '-')}",
            f"訂閱到期：{row.get('subscription_end_at','-')}",
            f"試用到期：{row.get('trial_end_at','-')}",
            f"待審核付款：{row.get('pending_review',0)}",
        ]
        for key in ["created_at", "updated_at", "device_id", "device_name"]:
            val = clean_text(raw.get(key, ""))
            if val:
                detail.append(f"{key}：{val}")
        self.set_readonly_text(getattr(self, "admin_user_detail_text", self.detail_text), "\n".join(detail))

    def _selected_admin_identity(self):
        row = self.selected_admin_user_record or {}
        raw = (row.get("raw", {}) or {}) if isinstance(row.get("raw"), dict) else {}
        return {
            "account": clean_text(row.get("account", "")),
            "username": clean_text(raw.get("username") or row.get("account", "")),
            "email": clean_text(row.get("email", "")),
            "user_id": clean_text(raw.get("user_id") or raw.get("id") or raw.get("member_id") or raw.get("uuid") or ""),
        }

    def manual_set_plan_selected_user(self):
        if not self.is_creator:
            messagebox.showwarning("提醒", "只有創作者帳號可以手動開通會員。")
            return
        identity = self._selected_admin_identity()
        if not identity.get("account") and not identity.get("email"):
            messagebox.showwarning("提醒", "請先選取一位會員。")
            return
        plan_type = clean_text(self.admin_members_plan_var.get()) or "monthly"
        days_text = clean_text(self.admin_members_days_var.get())
        end_at = clean_text(self.admin_members_end_var.get())
        try:
            days = int(days_text) if days_text else 30
            if days <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning("提醒", "請輸入正確的天數。")
            return
        payload = {
            "account": identity.get("account"),
            "username": identity.get("username"),
            "email": identity.get("email"),
            "target_account": identity.get("account"),
            "target_username": identity.get("username"),
            "target_email": identity.get("email"),
            "plan_type": plan_type,
            "subscription_status": "active" if plan_type not in {"none", "trial"} else plan_type,
            "days": days,
            "duration_days": days,
            "grant_days": days,
            "subscription_days": days,
            "end_at": end_at,
            "subscription_end_at": end_at,
            "note": "admin manual activation",
        }
        try:
            data = auth_post("/admin/set-plan", payload, headers=self._auth_headers())
            msg = data.get("message", f"已為 {identity.get('account') or identity.get('email')} 手動開通 {PAYMENT_PLAN_LABELS.get(plan_type, plan_type)}。")
            self.admin_users_status_var.set(msg)
            self.log(msg)
            messagebox.showinfo("成功", msg)
            self.load_admin_overview()
        except Exception as e:
            messagebox.showerror("手動開通失敗", str(e))

    def grant_free_selected_user(self):
        if not self.is_creator:
            messagebox.showwarning("提醒", "只有創作者帳號可以活動贈送會員。")
            return
        identity = self._selected_admin_identity()
        target_label = identity.get("account") or identity.get("email")
        if not target_label:
            messagebox.showwarning("提醒", "請先選取一位會員。")
            return
        days_text = clean_text(self.admin_members_days_var.get())
        reason = clean_text(self.admin_members_reason_var.get()) or "活動贈送"
        note = clean_text(self.admin_members_note_var.get())
        try:
            days = int(days_text) if days_text else 30
            if days <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning("提醒", "請輸入正確的天數。")
            return

        note_text = f"{reason}｜{note}" if note else reason
        confirm = (
            f"確定要贈送給：{target_label}\n"
            f"贈送天數：{days} 天\n"
            f"贈送原因：{reason}\n"
            f"備註：{note or '無'}"
        )
        if not messagebox.askyesno("確認活動贈送", confirm):
            return

        payload = {
            "account": identity.get("account"),
            "username": identity.get("username"),
            "email": identity.get("email"),
            "target_account": identity.get("account"),
            "target_username": identity.get("username"),
            "target_email": identity.get("email"),
            "days": days,
            "free_days": days,
            "grant_days": days,
            "duration_days": days,
            "reason": reason,
            "grant_reason": reason,
            "note": note_text,
        }
        try:
            data = auth_post("/admin/grant-free", payload, headers=self._auth_headers())
            msg = data.get("message", f"已為 {target_label} 贈送 {days} 天使用期限。")
            self.admin_users_status_var.set(msg)
            self.log(f"{msg}（原因：{reason}；備註：{note or '無'}）")
            messagebox.showinfo("成功", msg)
            self.admin_members_note_var.set("")
            self.load_admin_overview()
            self.load_admin_users()
        except Exception as e:
            messagebox.showerror("活動贈送失敗", str(e))

    def delete_selected_user(self):
        if not self.is_creator:
            messagebox.showwarning("提醒", "只有創作者帳號可以刪除會員。")
            return

        identity = self._selected_admin_identity()
        target_label = identity.get("account") or identity.get("email")
        if not target_label:
            messagebox.showwarning("提醒", "請先選取一位會員。")
            return

        if clean_text(target_label).lower() == "admin":
            messagebox.showwarning("提醒", "為避免誤刪，系統不允許直接刪除 admin 帳號。")
            return

        confirm_text = (
            f"確定要刪除會員：{target_label}？\n\n"
            "此動作主要用來清除殭屍帳號、測試帳號或無效帳號。\n"
            "刪除成功後，系統會重新整理會員清單。"
        )
        if not messagebox.askyesno("確認刪除會員", confirm_text):
            return

        payload = {
            "account": identity.get("account"),
            "username": identity.get("username"),
            "email": identity.get("email"),
            "user_id": identity.get("user_id"),
            "target_account": identity.get("account"),
            "target_username": identity.get("username"),
            "target_email": identity.get("email"),
            "target_user_id": identity.get("user_id"),
        }

        headers = self._auth_headers()
        user_id = clean_text(identity.get("user_id", ""))
        account = clean_text(identity.get("account", ""))
        email = clean_text(identity.get("email", ""))

        attempts = [
            ("POST", "/admin/delete-user", payload, None),
            ("POST", "/admin/users/delete", payload, None),
            ("POST", "/admin/remove-user", payload, None),
            ("POST", "/admin/user/delete", payload, None),
            ("DELETE", "/admin/delete-user", payload, None),
            ("DELETE", "/admin/users/delete", payload, None),
            ("DELETE", "/admin/remove-user", payload, None),
            ("DELETE", "/admin/user/delete", payload, None),
        ]
        if user_id:
            attempts.extend([
                ("DELETE", f"/admin/users/{user_id}", None, None),
                ("DELETE", f"/admin/user/{user_id}", None, None),
                ("POST", f"/admin/users/{user_id}/delete", payload, None),
            ])
        if account or email:
            attempts.extend([
                ("DELETE", "/admin/users", None, {"account": account, "email": email, "user_id": user_id}),
                ("DELETE", "/admin/user", None, {"account": account, "email": email, "user_id": user_id}),
            ])

        seen = set()
        last_error = None
        for method, path, body, params in attempts:
            key = (method, path, json.dumps(body, ensure_ascii=False, sort_keys=True) if isinstance(body, dict) else str(body), json.dumps(params, ensure_ascii=False, sort_keys=True) if isinstance(params, dict) else str(params))
            if key in seen:
                continue
            seen.add(key)
            try:
                data = auth_request(method, path, payload=body, headers=headers, params=params)
                msg = ""
                if isinstance(data, dict):
                    msg = clean_text(data.get("message") or data.get("detail") or "")
                if not msg:
                    msg = f"已刪除會員：{target_label}"
                self.admin_users_status_var.set(msg)
                self.log(f"{msg}｜method={method}｜path={path}")
                messagebox.showinfo("成功", msg)
                self.load_admin_overview()
                return
            except Exception as e:
                last_error = e
                err_text = clean_text(str(e)).lower()
                if any(x in err_text for x in ["404", "not found", "no route"]):
                    continue
                if any(x in err_text for x in ["405", "method not allowed"]):
                    continue
                if any(x in err_text for x in ["422", "field required", "validation"]):
                    continue
                break

        messagebox.showerror(
            "刪除會員失敗",
            "後端目前沒有對上可用的刪除會員接口，或接口欄位名稱與前端不同。\n\n"
            f"最後錯誤：{last_error or '未知錯誤'}"
        )
    def _coerce_bool(self, value):
        if isinstance(value, bool):
            return value
        s = clean_text(value).lower()
        return s in {"1", "true", "yes", "y", "on", "creator", "admin", "superadmin"}

    def _extract_creator_flag(self, payload):
        if not isinstance(payload, dict):
            return False
        user = payload.get("user") if isinstance(payload.get("user"), dict) else {}
        role_candidates = [
            payload.get("role"), payload.get("user_role"), payload.get("account_role"),
            user.get("role"), user.get("user_role"), user.get("account_role"),
        ]
        for role in role_candidates:
            role_text = clean_text(role).lower()
            if role_text in {"creator", "admin", "superadmin", "owner"}:
                return True
        bool_candidates = [
            payload.get("is_creator"), payload.get("creator"), payload.get("is_admin"), payload.get("admin"),
            user.get("is_creator"), user.get("creator"), user.get("is_admin"), user.get("admin"),
        ]
        return any(self._coerce_bool(x) for x in bool_candidates)

    def hide_admin_payment_review_tab(self):
        frame = self.sheet_frames.get("ADMIN_PAYMENT_REVIEW")
        if frame is None:
            return
        try:
            self.notebook.hide(frame)
        except Exception:
            pass

    def show_admin_payment_review_tab(self):
        frame = self.sheet_frames.get("ADMIN_PAYMENT_REVIEW")
        if frame is None:
            return
        try:
            self.notebook.index(frame)
            return
        except Exception:
            pass
        try:
            self.notebook.add(frame, text=self.sheet_titles.get("ADMIN_PAYMENT_REVIEW", "付款審核"))
        except Exception:
            pass

    def update_admin_payment_review_visibility(self, force_refresh=False):
        token = clean_text((self.auth_session or {}).get("token"))
        creator = self._extract_creator_flag(self.auth_session)
        if token and (force_refresh or (not creator and not self.auth_profile)):
            try:
                profile = auth_get("/auth/me", headers=self._auth_headers())
                self.auth_profile = profile if isinstance(profile, dict) else {}
                creator = creator or self._extract_creator_flag(self.auth_profile)
            except Exception:
                pass
        self.is_creator = bool(creator)
        if self.is_creator:
            self.show_admin_users_tab()
            self.show_admin_payment_review_tab()
            if force_refresh:
                try:
                    self.load_admin_overview()
                except Exception:
                    pass
        else:
            self.hide_admin_users_tab()
            self.hide_admin_payment_review_tab()

    def _auth_headers(self):
        token = clean_text((self.auth_session or {}).get("token"))
        if not token:
            return {}
        return {"Authorization": f"Bearer {token}"}

    def _normalize_payment_report_item(self, item):
        if not isinstance(item, dict):
            return {}
        plan_type = clean_text(item.get("plan_type") or item.get("plan") or "")
        status = clean_text(item.get("status") or item.get("review_status") or item.get("state") or "pending")
        return {
            "report_id": item.get("report_id") or item.get("id") or item.get("payment_report_id") or "",
            "account": clean_text(item.get("account") or item.get("username") or item.get("email") or ""),
            "payer_name": clean_text(item.get("payer_name") or item.get("name") or ""),
            "plan_type": plan_type,
            "plan_label": PAYMENT_PLAN_LABELS.get(plan_type, plan_type or "-"),
            "amount": item.get("amount") or item.get("transfer_amount") or 0,
            "transfer_last5": clean_text(item.get("transfer_last5") or item.get("last5") or ""),
            "bank": clean_text(item.get("bank") or item.get("bank_name") or item.get("transfer_bank") or ""),
            "status": status or "pending",
            "created_at": clean_text(str(item.get("created_at") or item.get("transfer_time") or item.get("created") or "")),
            "note": clean_text(item.get("note") or ""),
            "raw": item,
        }

    def _extract_payment_reports(self, data):
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for key in ["items", "data", "reports", "payment_reports", "results"]:
                value = data.get(key)
                if isinstance(value, list):
                    return value
        return []

    def load_admin_payment_reports(self):
        token = clean_text((self.auth_session or {}).get("token"))
        if not token:
            messagebox.showwarning("提醒", "請先登入管理員帳號後再載入付款回報。")
            return
        try:
            data = auth_get("/admin/payment-reports", headers=self._auth_headers())
            rows = [self._normalize_payment_report_item(x) for x in self._extract_payment_reports(data)]
            self.admin_payment_reports = rows
            self.apply_admin_payment_filters(update_status=True)
            self.log(f"管理員付款審核：已載入 {len(rows)} 筆付款回報")

            pending_ids = {str(x.get("report_id", "")) for x in rows if self._payment_status_bucket(x.get("status", "")) == "pending" and str(x.get("report_id", ""))}
            old_pending_ids = set(getattr(self, "admin_seen_pending_report_ids", set()) or set())
            new_pending_ids = pending_ids - old_pending_ids
            if self.admin_payment_loaded_once and new_pending_ids:
                alarm_text = f"付款審核提醒：目前有 {len(new_pending_ids)} 筆新的待審核匯款回報。"
                self.admin_payment_status_var.set(alarm_text)
                try:
                    self.bell()
                except Exception:
                    pass
                messagebox.showwarning("付款審核提醒", alarm_text)
            self.admin_seen_pending_report_ids = pending_ids
            self.admin_payment_loaded_once = True
        except Exception as e:
            self.admin_payment_status_var.set(f"載入失敗：{e}")
            messagebox.showerror("載入付款回報失敗", str(e))

    def _payment_status_bucket(self, status):
        s = clean_text(status).lower()
        if s in {"approved", "done", "success", "paid", "completed"}:
            return "approved"
        if s in {"rejected", "reject", "denied", "failed", "void"}:
            return "rejected"
        return "pending"

    def _payment_filter_match(self, row, selected_status, keyword, pending_only):
        bucket = self._payment_status_bucket(row.get("status", ""))
        if pending_only and bucket != "pending":
            return False
        if selected_status and selected_status != "all" and bucket != selected_status:
            return False
        if keyword:
            hay = " ".join([
                clean_text(row.get("account", "")),
                clean_text(row.get("payer_name", "")),
                clean_text(row.get("plan_label", row.get("plan_type", ""))),
                clean_text(row.get("amount", "")),
                clean_text(row.get("transfer_last5", "")),
                clean_text(row.get("bank", "")),
                clean_text(row.get("status", "")),
                clean_text(row.get("created_at", "")),
                clean_text(row.get("note", "")),
            ]).lower()
            if keyword.lower() not in hay:
                return False
        return True

    def _update_admin_payment_summary(self, rows=None):
        rows = self.admin_payment_reports if rows is None else rows
        pending = approved = rejected = 0
        for row in rows:
            bucket = self._payment_status_bucket(row.get("status", ""))
            if bucket == "approved":
                approved += 1
            elif bucket == "rejected":
                rejected += 1
            else:
                pending += 1
        self.admin_payment_summary_var.set(f"待審核：{pending}｜已核准：{approved}｜已駁回：{rejected}｜全部：{len(rows)}")

    def apply_admin_payment_filters(self, update_status=False):
        selected_status = clean_text(self.admin_payment_filter_status_var.get()).lower() or "all"
        keyword = clean_text(self.admin_payment_filter_keyword_var.get())
        pending_only = bool(self.admin_payment_pending_only_var.get())
        if pending_only:
            selected_status = "pending"
            self.admin_payment_filter_status_var.set("pending")
        filtered = [
            row for row in self.admin_payment_reports
            if self._payment_filter_match(row, selected_status, keyword, pending_only)
        ]
        self.admin_payment_filtered_reports = filtered
        self._update_admin_payment_summary(self.admin_payment_reports)
        self.refresh_admin_payment_tree(filtered_rows=filtered)
        if update_status:
            extra = f"，目前顯示 {len(filtered)} 筆" if len(filtered) != len(self.admin_payment_reports) else ""
            self.admin_payment_status_var.set(f"已載入 {len(self.admin_payment_reports)} 筆付款回報{extra}。")
        else:
            self.admin_payment_status_var.set(f"篩選完成：目前顯示 {len(filtered)} / {len(self.admin_payment_reports)} 筆。")

    def clear_admin_payment_filters(self):
        self.admin_payment_filter_status_var.set("pending")
        self.admin_payment_pending_only_var.set(True)
        self.admin_payment_filter_keyword_var.set("")
        self.apply_admin_payment_filters(update_status=False)

    def on_admin_payment_filter_change(self, event=None):
        if not self.admin_payment_pending_only_var.get() and clean_text(self.admin_payment_filter_status_var.get()).lower() not in {"pending", "approved", "rejected", "all"}:
            self.admin_payment_filter_status_var.set("all")
        self.apply_admin_payment_filters(update_status=False)

    def refresh_admin_payment_tree(self, filtered_rows=None):
        tree = getattr(self, "admin_payment_tree", None)
        if tree is None:
            return
        for iid in tree.get_children():
            tree.delete(iid)
        self.selected_payment_report_id = None
        rows = self.admin_payment_filtered_reports if filtered_rows is None else filtered_rows
        for row in rows:
            created_at = row.get("created_at", "")
            values = (
                row.get("report_id", ""),
                row.get("account", ""),
                row.get("payer_name", ""),
                row.get("plan_label", row.get("plan_type", "")),
                row.get("amount", 0),
                row.get("transfer_last5", ""),
                row.get("bank", ""),
                row.get("status", ""),
                created_at[:19].replace("T", " "),
            )
            tree.insert("", "end", iid=str(row.get("report_id", "")), values=values)
        self.set_readonly_text(getattr(self, "admin_payment_detail_text", self.detail_text), "請先選取一筆付款回報。")

    def on_admin_payment_select(self, event=None):
        tree = getattr(self, "admin_payment_tree", None)
        if tree is None:
            return
        selected = tree.selection()
        if not selected:
            self.selected_payment_report_id = None
            return
        iid = selected[0]
        self.selected_payment_report_id = iid
        source_rows = self.admin_payment_filtered_reports or self.admin_payment_reports
        row = next((x for x in source_rows if str(x.get("report_id", "")) == str(iid)), None)
        if not row:
            return
        detail = [
            f"報表ID：{row.get('report_id','')}",
            f"帳號：{row.get('account','')}",
            f"付款人：{row.get('payer_name','')}",
            f"方案：{row.get('plan_label', row.get('plan_type',''))}",
            f"金額：{row.get('amount',0)}",
            f"末五碼：{row.get('transfer_last5','')}",
            f"銀行：{row.get('bank','')}",
            f"狀態：{row.get('status','')}",
            f"建立時間：{row.get('created_at','')}",
            f"備註：{row.get('note','')}",
            "",
            "原始資料：",
            json.dumps(row.get('raw', {}), ensure_ascii=False, indent=2),
        ]
        self.set_readonly_text(self.admin_payment_detail_text, "\n".join(detail))

    def _run_payment_review_action(self, action):
        report_id = self.selected_payment_report_id
        if not report_id:
            messagebox.showwarning("提醒", "請先選取一筆付款回報。")
            return
        path = f"/admin/approve-payment-report/{report_id}" if action == "approve" else f"/admin/reject-payment-report/{report_id}"
        confirm_text = "確定要核准這筆付款回報嗎？" if action == "approve" else "確定要駁回這筆付款回報嗎？"
        if not messagebox.askyesno("確認", confirm_text):
            return
        try:
            data = auth_post(path, {}, timeout=20, headers=self._auth_headers())
            msg = data.get("message") or ("已核准付款回報。" if action == "approve" else "已駁回付款回報。")
            self.admin_payment_status_var.set(msg)
            self.log(f"管理員付款審核：{action} report_id={report_id}")
            messagebox.showinfo("成功", msg)
            self.load_admin_payment_reports()
        except Exception as e:
            messagebox.showerror("付款審核失敗", str(e))

    def approve_selected_payment_report(self):
        self._run_payment_review_action("approve")

    def reject_selected_payment_report(self):
        self._run_payment_review_action("reject")

    def _on_payment_plan_combo_change(self, event=None):
        selected = clean_text(self.payment_plan_var.get())
        reverse_map = {v: k for k, v in PAYMENT_PLAN_LABELS.items()}
        plan_key = reverse_map.get(selected, selected)
        if plan_key not in PAYMENT_PLAN_LABELS:
            plan_key = "monthly"
        self.payment_plan_var.set(PAYMENT_PLAN_LABELS[plan_key])
        self.payment_amount_var.set(str(PAYMENT_PLAN_PRICES.get(plan_key, "")))

    def open_payment_tab(self):
        try:
            frame = self.sheet_frames.get('PAYMENT')
            if frame is not None:
                self.notebook.select(frame)
        except Exception:
            pass

    def clear_payment_form(self):
        self.payment_plan_var.set('monthly')
        try:
            self.payment_plan_combo.current(0)
        except Exception:
            pass
        self.payment_bank_var.set(PAYMENT_BANKS[0])
        try:
            self.payment_bank_combo.current(0)
        except Exception:
            pass
        self.payment_last5_var.set('')
        self.payment_amount_var.set('')
        self.payment_status_var.set('請選擇方案並完成付款回報。')

    def update_payment_tab_state(self):
        account = clean_text((self.auth_session or {}).get('username') or (self.auth_session or {}).get('account') or (self.auth_session or {}).get('email'))
        enabled = bool(account)
        state = 'normal' if enabled else 'disabled'
        readonly = 'readonly' if enabled else 'disabled'
        try:
            self.payment_plan_combo.config(state=readonly)
            self.payment_bank_combo.config(state=readonly)
            self.payment_last5_entry.config(state=state)
            self.payment_amount_entry.config(state=state)
            self.payment_submit_btn.config(state=state)
        except Exception:
            pass
        if not enabled:
            self.payment_status_var.set('請先登入後再送出付款回報。')

    def submit_payment_report_legacy(self):
        account = clean_text((self.auth_session or {}).get('username') or (self.auth_session or {}).get('account') or (self.auth_session or {}).get('email'))
        if not account:
            messagebox.showwarning('提醒', '請先登入後再送出付款回報。')
            return
        plan_type = clean_text(self.payment_plan_var.get())
        bank = clean_text(self.payment_bank_var.get())
        last5 = clean_text(self.payment_last5_var.get())
        amount_text = clean_text(self.payment_amount_var.get())
        if plan_type not in PAYMENT_PLAN_LABELS:
            messagebox.showwarning('提醒', '請選擇訂閱方案。')
            return
        if not bank:
            messagebox.showwarning('提醒', '請選擇匯款銀行。')
            return
        if len(last5) != 5 or not last5.isdigit():
            messagebox.showwarning('提醒', '請輸入 5 碼匯款末五碼。')
            return
        try:
            amount = int(float(amount_text))
            if amount <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning('提醒', '請輸入正確的匯款金額。')
            return

        token = clean_text((self.auth_session or {}).get('token'))
        if not token:
            messagebox.showwarning('提醒', '登入資訊失效，請先重新登入。')
            return

        payload = {
            'account': account,
            'token': token,
            'plan_type': plan_type,
            'bank_name': bank,
            'bank': bank,
            'transfer_bank': bank,
            'last5': last5,
            'transfer_last5': last5,
            'amount': amount,
            'transfer_amount': amount,
        }
        try:
            data = auth_post('/payments/report', payload, headers=self._auth_headers())
            msg = data.get('message', '付款回報已送出，請等待管理員確認。')
            self.payment_status_var.set(msg)
            self.log(f'付款回報已送出：{account} / {PAYMENT_PLAN_LABELS.get(plan_type, plan_type)} / {bank}')
            messagebox.showinfo('成功', msg)
            self.payment_last5_var.set('')
            self.payment_amount_var.set('')
        except Exception as e:
            messagebox.showerror('付款回報失敗', f'送出付款回報失敗：{e}')

    def log(self, msg):
        now = datetime.now().strftime("%H:%M:%S")
        text = f"[{now}] {msg}\n"
        self.log_text.config(state="normal")
        self.log_text.insert("end", text)
        self.log_text.see("end")
        self.log_text.config(state="disabled")
        self.update_idletasks()

    def set_status(self, text):
        self.status_var.set(f"狀態：{text}")
        self.update_idletasks()

    def set_readonly_text(self, widget, text):
        widget.config(state="normal")
        widget.delete("1.0", "end")
        widget.insert("end", text)
        widget.config(state="disabled")

    def show_loading_popup(self):
        if self.loading_popup is not None:
            return

        pop = tk.Toplevel(self)
        pop.title("分析中")
        pop.geometry("520x500")
        pop.resizable(False, False)
        pop.transient(self)
        pop.grab_set()

        try:
            pop.iconbitmap(resource_path(ICON_FILE))
        except Exception:
            pass

        frm = ttk.Frame(pop, padding=18)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="系統分析中，請稍候...", font=("Microsoft JhengHei", 14, "bold")).pack(pady=(0, 12))

        img_path = find_brand_image_path()
        if img_path is not None:
            try:
                self.loading_img = tk.PhotoImage(file=img_path)
                lbl = ttk.Label(frm, image=self.loading_img)
                lbl.pack(pady=8)
            except Exception:
                canvas = tk.Canvas(frm, width=260, height=260, bg="#A11B1B", highlightthickness=0)
                canvas.pack(pady=8)
                canvas.create_text(130, 130, text="原", font=("Microsoft JhengHei", 90, "bold"), fill="black")
        else:
            canvas = tk.Canvas(frm, width=260, height=260, bg="#A11B1B", highlightthickness=0)
            canvas.pack(pady=8)
            canvas.create_text(130, 130, text="原", font=("Microsoft JhengHei", 90, "bold"), fill="black")

        ttk.Label(
            frm,
            text=DISCLAIMER_TEXT,
            wraplength=450,
            justify="left",
            font=("Microsoft JhengHei", 10)
        ).pack(pady=(14, 8))

        self.loading_popup = pop
        self.update_idletasks()

    def hide_loading_popup(self):
        if self.loading_popup is not None:
            try:
                self.loading_popup.grab_release()
            except Exception:
                pass
            try:
                self.loading_popup.destroy()
            except Exception:
                pass
            self.loading_popup = None
            self.loading_img = None

    def initialize_auth_flow(self):
        self.lock_features(reason="尚未登入")
        if self.try_restore_session():
            return
        self.open_auth_dialog()

    def try_restore_session(self):
        sess = self.auth_session or {}
        if not bool(sess.get("remember_me", False)):
            return False
        account = clean_text(sess.get("username") or sess.get("account") or sess.get("email"))
        device_id = clean_text(sess.get("device_id") or self.device_info.get("device_id"))
        token = clean_text(sess.get("token"))
        if not account or not device_id:
            return False
        ok = self.check_license_and_apply(account, device_id, token=token, silent=True)
        return ok

    def open_auth_dialog(self):
        dlg = AuthDialog(self, self.device_info)
        self.wait_window(dlg)
        if dlg.result is None:
            self.destroy()
            return
        self.auth_session = dlg.result
        ok = self.check_license_and_apply(
            self.auth_session.get("username")
            or self.auth_session.get("account", "")
            or self.auth_session.get("email", ""),
            self.auth_session.get("device_id", self.device_info.get("device_id", "")),
            token=self.auth_session.get("token", ""),
            silent=False,
        )
        if not ok:
            self.open_auth_dialog()

    def check_license_and_apply(self, account, device_id, token="", silent=False):
        try:
            data = auth_get(f"/license/status?account={quote_plus(account)}&device_id={quote_plus(device_id)}")
            self.license_info = data
            allowed = bool(data.get("allowed", False))
            sub = clean_text(data.get("subscription_status", "none")) or "none"
            plan_type = clean_text(data.get("plan_type", ""))
            trial_end_at = clean_text(str(data.get("trial_end_at", "") or ""))
            subscription_end_at = clean_text(str(data.get("subscription_end_at", "") or ""))
            remain_text = ""
            if trial_end_at and trial_end_at != 'None':
                try:
                    dt = datetime.fromisoformat(trial_end_at.replace('Z', '+00:00').replace(' ', 'T'))
                    now = datetime.utcnow().replace(tzinfo=dt.tzinfo)
                    days_left = (dt - now).days
                    remain_text = f"，剩餘約 {max(days_left,0)} 天"
                except Exception:
                    pass
            self.user_email_var.set(f"帳號：{account}")
            if sub == 'trial':
                self.license_var.set(f"授權：免費試用中{remain_text}")
            elif sub == 'free_grant':
                end_text = f"，到期：{subscription_end_at[:10]}" if subscription_end_at else ""
                self.license_var.set(f"授權：活動贈送{end_text}")
            elif sub == 'active':
                end_text = f"，到期：{subscription_end_at[:10]}" if subscription_end_at else ""
                name_map = {'monthly': '月訂閱', 'halfyear': '半年訂閱', 'yearly': '年訂閱'}
                self.license_var.set(f"授權：{name_map.get(plan_type, '已訂閱')}{end_text}")
            else:
                self.license_var.set(f"授權：{data.get('message','未開通')}")

            session_payload = {
                "login_account": (self.auth_session or {}).get("login_account", account),
                "username": clean_text(data.get("username", "")) or clean_text((self.auth_session or {}).get("username", "")) or account,
                "account": clean_text(data.get("username", "")) or account,
                "email": clean_text(data.get("email", "")) or clean_text((self.auth_session or {}).get("email", "")),
                "phone": clean_text(data.get("phone") or data.get("mobile") or "") or clean_text((self.auth_session or {}).get("phone", "")) or clean_text((self.auth_profile or {}).get("phone", "")) or clean_text((self.auth_profile or {}).get("mobile", "")),
                "token": token,
                "device_id": device_id,
                "device_name": self.device_info.get("device_name", ""),
                "subscription_status": sub,
                "plan_type": plan_type,
                "trial_end_at": trial_end_at,
                "subscription_end_at": subscription_end_at,
                "role": clean_text(data.get("role", "")) or clean_text((self.auth_session or {}).get("role", "")) or clean_text((self.auth_profile or {}).get("role", "")),
                "is_creator": bool(data.get("is_creator", (self.auth_session or {}).get("is_creator", False) or (self.auth_profile or {}).get("is_creator", False))),
                "is_admin": bool(data.get("is_admin", (self.auth_session or {}).get("is_admin", False) or (self.auth_profile or {}).get("is_admin", False))),
            }
            self.auth_session = session_payload
            save_auth_session(session_payload)
            self.update_payment_tab_state()
            self.update_admin_payment_review_visibility(force_refresh=True)

            if allowed:
                self.unlock_features()
                self.set_status("授權有效")
                if not silent:
                    self.log(f"登入成功：{account}（{sub}）")
                return True

            self.lock_features(reason=data.get("message", "尚未開通授權"))
            if not silent:
                messagebox.showwarning("授權提醒", data.get("message", "尚未開通授權"))
            return False
        except Exception as e:
            self.lock_features(reason="無法連線到授權伺服器")
            if not silent:
                messagebox.showerror("授權失敗", f"無法驗證授權：{e}")
            return False

    def lock_features(self, reason="未授權"):
        self.analysis_allowed = False
        self.run_btn.config(state="disabled")
        self.rebuild_btn.config(state="disabled")
        self.auto_on_btn.config(state="disabled")
        self.auto_off_btn.config(state="disabled")
        self.update_payment_tab_state()
        self.set_status(reason)

    def unlock_features(self):
        self.analysis_allowed = True
        self.run_btn.config(state="normal")
        self.rebuild_btn.config(state="normal")
        self.auto_on_btn.config(state="normal")
        self.auto_off_btn.config(state="normal")
        self.update_payment_tab_state()

    def logout(self):
        clear_auth_session()
        self.auth_session = {}
        self.auth_profile = {}
        self.is_creator = False
        self.hide_admin_users_tab()
        self.hide_admin_payment_review_tab()
        self.license_info = {}
        self.user_email_var.set("帳號：尚未登入")
        self.license_var.set("授權：尚未驗證")
        self.update_payment_tab_state()
        self.lock_features(reason="已登出")
        self.open_auth_dialog()

    def run_analysis_thread(self):
        if not self.analysis_allowed:
            messagebox.showwarning("授權提醒", "目前尚未登入，或試用已到期／未訂閱，無法執行分析。")
            return
        if self.running:
            messagebox.showinfo("提醒", "分析仍在執行中，請稍候。")
            return
        th = threading.Thread(target=self.run_analysis, args=(False,), daemon=True)
        th.start()

    def run_force_rebuild_thread(self):
        if not self.analysis_allowed:
            messagebox.showwarning("授權提醒", "目前尚未登入，或試用已到期／未訂閱，無法強制重建。")
            return
        if self.running:
            messagebox.showinfo("提醒", "分析仍在執行中，請稍候。")
            return
        ok = messagebox.askyesno("確認", "這會重新抓網路資料並覆蓋同日快照，是否繼續？")
        if not ok:
            return
        th = threading.Thread(target=self.run_analysis, args=(True,), daemon=True)
        th.start()

    def run_analysis(self, force_rebuild_snapshot=False):
        self.running = True
        self.run_btn.config(state="disabled")
        self.rebuild_btn.config(state="disabled")
        self.set_status("執行分析中...")
        self.after(0, self.show_loading_popup)

        try:
            if force_rebuild_snapshot:
                self.log("開始強制重建當日快照...")
            else:
                self.log("開始執行分析...")

            result = build_all_excel(logger=self.log, force_rebuild_snapshot=force_rebuild_snapshot)
            self.latest_result = result
            self.latest_excel_path = result.get("excel_path", OUTPUT_XLSX)

            self.refresh_notebook(result)

            settle_date = result.get("settle_date", "未知")
            bull_count = len(result.get("CLIENT_BULLISH", pd.DataFrame()))
            bull_keyk_count = len(result.get("CLIENT_BULLISH_KEYK", pd.DataFrame()))
            bear_count = len(result.get("CLIENT_BEARISH", pd.DataFrame()))
            bear_keyk_count = len(result.get("CLIENT_BEARISH_KEYK", pd.DataFrame()))
            hold_count = len(self.user_holdings)

            self.settle_var.set(f"本次結算日：{settle_date}")
            self.summary_var.set(f"看多：{bull_count} 檔｜多方關鍵K：{bull_keyk_count} 檔｜看空：{bear_count} 檔｜空方關鍵K：{bear_keyk_count} 檔｜持股：{hold_count} 檔")

            self.set_status("分析完成")
            self.log("分析完成。")
            self.show_alarm_summary()

        except Exception as e:
            self.log(f"即時抓取失敗：{e}")
            self.log("改為嘗試載入最近一日可用快照...")

            fallback = find_latest_compatible_snapshot(days_back=10)
            if fallback is not None:
                self.latest_result = fallback
                self.latest_excel_path = fallback.get("excel_path", OUTPUT_XLSX)
                self.refresh_notebook(fallback)

                settle_date = fallback.get("settle_date", "未知")
                bull_count = len(fallback.get("CLIENT_BULLISH", pd.DataFrame()))
                bear_count = len(fallback.get("CLIENT_BEARISH", pd.DataFrame()))
                hold_count = len(self.user_holdings)

                self.settle_var.set(f"本次結算日：{settle_date}（回退結果）")
                self.summary_var.set(f"看多：{bull_count} 檔｜看空：{bear_count} 檔｜持股：{hold_count} 檔")
                self.set_status("已載入前一可用結果")
                self.log(f"已回退載入：{settle_date}")
                messagebox.showwarning("提示", f"最新資料暫時抓不到，已自動改用 {settle_date} 的最近可用結果。")
            else:
                self.set_status("執行失敗")
                self.log(traceback.format_exc())
                messagebox.showerror("執行失敗", f"{e}\n且找不到可回退的快照。")
        finally:
            self.running = False
            self.run_btn.config(state="normal")
            self.rebuild_btn.config(state="normal")
            self.after(0, self.hide_loading_popup)

    def refresh_notebook(self, result_dict):
        for sheet in ["CLIENT_BULLISH", "CLIENT_BULLISH_KEYK", "CLIENT_BEARISH", "CLIENT_BEARISH_KEYK"]:
            df = result_dict.get(sheet, pd.DataFrame())
            self.current_view_data[sheet] = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
            self.tree_sort_state[sheet] = {"column": None, "ascending": True}
            self.load_df_to_tree(self.sheet_trees[sheet], self.current_view_data[sheet], sheet)

        self.refresh_holdings_view(select_first=False)
        self.after(100, self.select_first_row_and_refresh)

    def refresh_holdings_view(self, select_first=True):
        holdings_df = build_holdings_view(self.user_holdings, self.latest_result)
        self.current_view_data["MY_HOLDINGS"] = holdings_df
        self.tree_sort_state["MY_HOLDINGS"] = {"column": None, "ascending": True}
        self.load_df_to_tree(self.sheet_trees["MY_HOLDINGS"], holdings_df, "MY_HOLDINGS")

        settle_date = self.latest_result.get("settle_date", "尚未執行")
        bull_count = len(self.latest_result.get("CLIENT_BULLISH", pd.DataFrame())) if self.latest_result else 0
        bull_keyk_count = len(self.latest_result.get("CLIENT_BULLISH_KEYK", pd.DataFrame())) if self.latest_result else 0
        bear_count = len(self.latest_result.get("CLIENT_BEARISH", pd.DataFrame())) if self.latest_result else 0
        bear_keyk_count = len(self.latest_result.get("CLIENT_BEARISH_KEYK", pd.DataFrame())) if self.latest_result else 0
        hold_count = len(self.user_holdings)

        if self.latest_result:
            self.settle_var.set(f"本次結算日：{settle_date}")
        self.summary_var.set(f"看多：{bull_count} 檔｜多方關鍵K：{bull_keyk_count} 檔｜看空：{bear_count} 檔｜空方關鍵K：{bear_keyk_count} 檔｜持股：{hold_count} 檔")

        if select_first:
            tree = self.sheet_trees["MY_HOLDINGS"]
            items = tree.get_children()
            if items:
                tree.selection_set(items[0])
                tree.focus(items[0])
                tree.see(items[0])
                self.update_selected_stock_display()

    def add_holding_code(self):
        code = normalize_code(self.holding_code_var.get())
        if not is_valid_stock_code(code):
            messagebox.showwarning("提醒", "請輸入正確的 4 碼股票代號。")
            return
        if code in self.user_holdings:
            messagebox.showinfo("提醒", f"{code} 已經在持股清單中。")
            return

        self.user_holdings.append(code)
        self.user_holdings = list(dict.fromkeys(self.user_holdings))
        save_holdings(self.user_holdings)
        self.holding_code_var.set("")
        self.log(f"已加入持股：{code}")
        self.refresh_holdings_view()

    def remove_selected_holding(self):
        tree = self.sheet_trees["MY_HOLDINGS"]
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("提醒", "請先選擇要刪除的持股。")
            return

        cols = list(tree["columns"])
        if "股票代號" not in cols:
            return
        idx = cols.index("股票代號")
        values = tree.item(selected[0], "values")
        code = str(values[idx])

        if code in self.user_holdings:
            self.user_holdings.remove(code)
            save_holdings(self.user_holdings)
            self.log(f"已刪除持股：{code}")
            self.refresh_holdings_view()

    def select_first_row_and_refresh(self):
        tree = self.get_current_tree()
        items = tree.get_children()
        if items:
            tree.selection_set(items[0])
            tree.focus(items[0])
            tree.see(items[0])
            self.update_selected_stock_display()

    def load_df_to_tree(self, tree, df, sheet_name=None):
        for item in tree.get_children():
            tree.delete(item)

        if df is None or df.empty:
            tree["columns"] = []
            return

        cols = list(df.columns)
        tree["columns"] = cols

        for c in cols:
            heading_text = c
            if sheet_name in self.tree_sort_state:
                sort_info = self.tree_sort_state[sheet_name]
                if sort_info["column"] == c:
                    heading_text = f"{c} {'▲' if sort_info['ascending'] else '▼'}"

            tree.heading(
                c,
                text=heading_text,
                command=lambda col=c, sheet=sheet_name: self.on_tree_heading_click(sheet, col)
            )

            if c == "股票名稱":
                tree.column(c, width=200, anchor="center")
            elif c == "產業別":
                tree.column(c, width=160, anchor="center")
            elif c == "週結算日期":
                tree.column(c, width=120, anchor="center")
            elif c == "星等":
                tree.column(c, width=100, anchor="center")
            elif c == "意見":
                tree.column(c, width=150, anchor="center")
            else:
                tree.column(c, width=110, anchor="center")

        show_df = df.copy().fillna("")
        max_rows = min(len(show_df), 3000)

        for i in range(max_rows):
            vals = [str(show_df.iloc[i][c]) for c in cols]
            tree.insert("", "end", values=vals)

    def parse_sort_value(self, value, column_name):
        if pd.isna(value):
            return (1, None)

        if column_name == "星等":
            s = str(value)
            return (0, s.count("★"))

        if column_name == "週結算日期":
            try:
                return (0, pd.to_datetime(value))
            except Exception:
                return (1, pd.Timestamp("1900-01-01"))

        if "Alarm" in str(column_name) or "提醒" in str(column_name):
            return (0, 1 if str(value) == "是" else 0)

        if "代號" in str(column_name):
            s = str(value).strip()
            if s.isdigit():
                return (0, int(s))
            return (0, s)

        num = parse_number(value)
        if num is not None:
            return (0, num)

        return (0, str(value))

    def on_tree_heading_click(self, sheet_name, column_name):
        if not sheet_name:
            return
        if sheet_name not in self.current_view_data:
            return

        df = self.current_view_data[sheet_name]
        if df is None or df.empty or column_name not in df.columns:
            return

        sort_state = self.tree_sort_state.get(sheet_name, {"column": None, "ascending": True})
        if sort_state["column"] == column_name:
            ascending = not sort_state["ascending"]
        else:
            ascending = False if column_name in ["星等", "星等數值", "StrongScore", "BearishScore"] else True

        tmp = df.copy()
        tmp["_sort_key_"] = tmp[column_name].apply(lambda x: self.parse_sort_value(x, column_name))
        tmp = tmp.sort_values("_sort_key_", ascending=ascending, kind="mergesort").drop(columns=["_sort_key_"]).reset_index(drop=True)

        if "項次" in tmp.columns:
            tmp["項次"] = range(1, len(tmp) + 1)

        self.current_view_data[sheet_name] = tmp
        self.tree_sort_state[sheet_name] = {"column": column_name, "ascending": ascending}

        tree = self.sheet_trees[sheet_name]
        self.load_df_to_tree(tree, tmp, sheet_name)

        items = tree.get_children()
        if items:
            tree.selection_set(items[0])
            tree.focus(items[0])
            tree.see(items[0])
            self.update_selected_stock_display()

    def get_current_sheet_name(self):
        cur_tab = self.notebook.select()
        cur_idx = self.notebook.index(cur_tab)
        return self.sheet_order[cur_idx]

    def get_current_tree(self):
        return self.sheet_trees[self.get_current_sheet_name()]

    def on_tree_select(self, event=None):
        self.after(50, self.update_selected_stock_display)

    def get_selected_code(self):
        tree = self.get_current_tree()
        selected = tree.selection()
        if not selected:
            return None

        values = tree.item(selected[0], "values")
        cols = list(tree["columns"])
        if "股票代號" not in cols:
            return None

        idx = cols.index("股票代號")
        return str(values[idx])

    def update_selected_stock_display(self):
        code = self.get_selected_code()
        if not code:
            return

        raw = self.latest_result.get("WEEKLY_MA_RAW", pd.DataFrame())
        if raw is None or raw.empty:
            self.chart_title_var.set("查無週K資料")
            self.set_readonly_text(self.detail_text, "目前沒有可顯示的週K資料。")
            self.set_readonly_text(self.company_text, "目前沒有可顯示的公司基本資訊。")
            return

        sub = raw[raw["股票代號"].astype(str) == str(code)].copy()
        if sub.empty:
            self.chart_title_var.set(f"{code}｜查無週K資料")
            self.set_readonly_text(self.detail_text, f"查無 {code} 的週K資料。")
            self.set_readonly_text(self.company_text, f"查無 {code} 的公司基本資訊。")
            return

        sub = sub.sort_values("週結算日期").tail(36).reset_index(drop=True)
        self.draw_weekly_chart(sub)
        self.update_detail_text(code)
        self.update_company_info_text(code)

    def draw_weekly_chart(self, sub):
        for child in self.chart_container.winfo_children():
            child.destroy()

        code = str(sub.iloc[-1]["股票代號"])
        name = str(sub.iloc[-1]["股票名稱"])
        self.chart_title_var.set(f"{code} {name}｜週K線圖")

        fig = Figure(figsize=(8.8, 4.8), dpi=100)
        ax = fig.add_subplot(111)

        x = list(range(len(sub)))
        width = 0.55

        highs = sub["週最高價"].tolist()
        lows = sub["週最低價"].tolist()
        opens = sub["週開盤價"].tolist()
        closes = sub["週收盤價"].tolist()

        for i in range(len(sub)):
            o = opens[i]
            h = highs[i]
            l = lows[i]
            c = closes[i]

            if any(pd.isna(v) for v in [o, h, l, c]):
                continue

            up = c >= o
            color = "#d62728" if up else "#2ca02c"

            ax.vlines(i, l, h, linewidth=1.0)
            lower = min(o, c)
            height = max(abs(c - o), 0.01)
            ax.add_patch(Rectangle((i - width / 2, lower), width, height, facecolor=color, edgecolor=color, linewidth=1.0))

        if "週3MA" in sub.columns:
            ax.plot(x, sub["週3MA"], linewidth=1.1, label="3MA")
        if "週5MA" in sub.columns:
            ax.plot(x, sub["週5MA"], linewidth=1.1, label="5MA")
        if "週20MA" in sub.columns:
            ax.plot(x, sub["週20MA"], linewidth=1.3, label="20MA")

        labels = [pd.to_datetime(d).strftime("%m-%d") for d in sub["週結算日期"]]
        step = max(1, len(x) // 10)
        ax.set_xticks(x[::step])
        ax.set_xticklabels(labels[::step], rotation=30)
        ax.set_title(f"{code} {name}")
        ax.set_ylabel("Price")
        ax.grid(True, alpha=0.3)
        ax.legend(loc="upper left")

        self.chart_canvas = FigureCanvasTkAgg(fig, master=self.chart_container)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill="both", expand=True)

    def update_detail_text(self, code):
        sheet = self.get_current_sheet_name()

        if sheet == "CLIENT_BULLISH":
            source_df = self.latest_result.get("TRAINING_POOL", pd.DataFrame())
            one = source_df[source_df["股票代號"].astype(str) == str(code)].copy()
            if one.empty:
                self.set_readonly_text(self.detail_text, f"查無 {code} 的看多詳細資料。")
                return
            row = one.iloc[0]

            lines = [
                f"股票代號：{row.get('股票代號', '')}",
                f"股票名稱：{row.get('股票名稱', '')}",
                f"產業別：{row.get('產業別', '')}",
                f"週結算日期：{row.get('週結算日期', '')}",
                "",
                f"買進星等：{row.get('星等', '')}",
                f"StrongScore：{row.get('StrongScore', '')}",
                f"乖離率(20MA)：{row.get('乖離率(%)', '')}%",
                "",
                f"短線停利Alarm：{row.get('短線停利Alarm', '否')}",
                f"長線停利Alarm：{row.get('長線停利Alarm', '否')}",
                "",
                f"技術標籤：{row.get('技術標籤', '')}",
                f"是否最新正式突破：{row.get('是否最新正式突破', '')}",
                f"趨勢線距離(%)：{row.get('趨勢線距離(%)', '')}",
                f"盤整區距離(%)：{row.get('盤整區距離(%)', '')}",
            ]

        elif sheet == "CLIENT_BULLISH_KEYK":
            source_df = self.latest_result.get("STRICT_BREAKOUT", pd.DataFrame())
            one = source_df[source_df["股票代號"].astype(str) == str(code)].copy()
            if one.empty:
                self.set_readonly_text(self.detail_text, f"查無 {code} 的多方關鍵K詳細資料。")
                return
            row = one.iloc[0]

            lines = [
                f"股票代號：{row.get('股票代號', '')}",
                f"股票名稱：{row.get('股票名稱', '')}",
                f"產業別：{row.get('產業別', '')}",
                f"週結算日期：{row.get('週結算日期', '')}",
                "",
                f"最新週收盤價：{row.get('最新週收盤價', '')}",
                f"週20MA：{row.get('週20MA', '')}",
                f"最新一週成交量(張)：{row.get('最新一週成交量(張)', '')}",
                "",
                f"第一高點日期：{row.get('第一高點日期', '')}",
                f"第一高點價格：{row.get('第一高點價格', '')}",
                f"第二高點日期：{row.get('第二高點日期', '')}",
                f"第二高點價格：{row.get('第二高點價格', '')}",
                f"盤整區高點：{row.get('盤整區高點', '')}",
                f"盤整區低點：{row.get('盤整區低點', '')}",
                f"最新壓力線價位：{row.get('最新壓力線價位', '')}",
                "",
                f"最後一根是否突破趨勢線：{row.get('最後一根是否突破趨勢線', '')}",
                f"最後一根是否突破盤整：{row.get('最後一根是否突破盤整', '')}",
                f"最後一根是否紅K：{row.get('最後一根是否紅K', '')}",
                f"是否實K突破趨勢線：{row.get('是否實K突破趨勢線', '')}",
                f"趨勢線距離(%)：{row.get('趨勢線距離(%)', '')}",
                f"盤整區距離(%)：{row.get('盤整區距離(%)', '')}",
                f"穿越K棒數：{row.get('穿越K棒數', '')}",
            ]

        elif sheet == "CLIENT_BEARISH":
            source_df = self.latest_result.get("BEARISH_TRAINING_POOL", pd.DataFrame())
            one = source_df[source_df["股票代號"].astype(str) == str(code)].copy()
            if one.empty:
                self.set_readonly_text(self.detail_text, f"查無 {code} 的看空詳細資料。")
                return
            row = one.iloc[0]

            lines = [
                f"股票代號：{row.get('股票代號', '')}",
                f"股票名稱：{row.get('股票名稱', '')}",
                f"產業別：{row.get('產業別', '')}",
                f"週結算日期：{row.get('週結算日期', '')}",
                "",
                f"放空星等：{row.get('星等', '')}",
                f"BearishScore：{row.get('BearishScore', '')}",
                f"乖離率(20MA)：{row.get('乖離率(%)', '')}%",
                "",
                f"短線回補Alarm：{row.get('短線回補Alarm', '否')}",
                f"長線回補Alarm：{row.get('長線回補Alarm', '否')}",
                "",
                f"空方技術標籤：{row.get('空方技術標籤', '')}",
                f"是否最新正式跌破：{row.get('是否最新正式跌破', '')}",
                f"趨勢線距離(%)：{row.get('趨勢線距離(%)', '')}",
            ]

        elif sheet == "CLIENT_BEARISH_KEYK":
            source_df = self.latest_result.get("BEARISH_KEY_BREAKDOWN", pd.DataFrame())
            one = source_df[source_df["股票代號"].astype(str) == str(code)].copy()
            if one.empty:
                self.set_readonly_text(self.detail_text, f"查無 {code} 的空方關鍵K詳細資料。")
                return
            row = one.iloc[0]

            lines = [
                f"股票代號：{row.get('股票代號', '')}",
                f"股票名稱：{row.get('股票名稱', '')}",
                f"產業別：{row.get('產業別', '')}",
                f"週結算日期：{row.get('週結算日期', '')}",
                "",
                f"最新週收盤價：{row.get('最新週收盤價', '')}",
                f"週20MA：{row.get('週20MA', '')}",
                f"最新一週成交量(張)：{row.get('最新一週成交量(張)', '')}",
                "",
                f"第一低點日期：{row.get('第一低點日期', '')}",
                f"第一低點價格：{row.get('第一低點價格', '')}",
                f"第二低點日期：{row.get('第二低點日期', '')}",
                f"第二低點價格：{row.get('第二低點價格', '')}",
                f"最新支撐線價位：{row.get('最新支撐線價位', '')}",
                "",
                f"最後一根是否跌破趨勢線：{row.get('最後一根是否跌破趨勢線', '')}",
                f"最後一根是否黑K：{row.get('最後一根是否黑K', '')}",
                f"是否實K跌破趨勢線：{row.get('是否實K跌破趨勢線', '')}",
                f"趨勢線距離(%)：{row.get('趨勢線距離(%)', '')}",
                f"穿越K棒數：{row.get('穿越K棒數', '')}",
            ]

        else:
            bull_df = self.latest_result.get("TRAINING_POOL", pd.DataFrame())
            bear_df = self.latest_result.get("BEARISH_TRAINING_POOL", pd.DataFrame())

            bull_one = bull_df[bull_df["股票代號"].astype(str) == str(code)] if isinstance(bull_df, pd.DataFrame) else pd.DataFrame()
            bear_one = bear_df[bear_df["股票代號"].astype(str) == str(code)] if isinstance(bear_df, pd.DataFrame) else pd.DataFrame()

            if not bull_one.empty:
                row = bull_one.iloc[0]
                lines = [
                    f"股票代號：{row.get('股票代號', '')}",
                    f"股票名稱：{row.get('股票名稱', '')}",
                    f"持股判定：系統看多池",
                    "",
                    f"星等：{row.get('星等', '')}",
                    f"StrongScore：{row.get('StrongScore', '')}",
                    f"乖離率(20MA)：{row.get('乖離率(%)', '')}%",
                    f"短線停利Alarm：{row.get('短線停利Alarm', '否')}",
                    f"長線停利Alarm：{row.get('長線停利Alarm', '否')}",
                    "",
                    f"技術標籤：{row.get('技術標籤', '')}",
                    f"系統意見：可提供意見",
                ]
            elif not bear_one.empty:
                row = bear_one.iloc[0]
                lines = [
                    f"股票代號：{row.get('股票代號', '')}",
                    f"股票名稱：{row.get('股票名稱', '')}",
                    f"持股判定：系統看空池",
                    "",
                    f"星等：{row.get('星等', '')}",
                    f"BearishScore：{row.get('BearishScore', '')}",
                    f"乖離率(20MA)：{row.get('乖離率(%)', '')}%",
                    f"短線回補Alarm：{row.get('短線回補Alarm', '否')}",
                    f"長線回補Alarm：{row.get('長線回補Alarm', '否')}",
                    "",
                    f"空方技術標籤：{row.get('空方技術標籤', '')}",
                    f"系統意見：可提供意見",
                ]
            else:
                lines = [
                    f"股票代號：{code}",
                    "",
                    "持股判定：非系統選股池",
                    "系統意見：無法提供意見",
                    "",
                    "說明：",
                    "1. 這檔股票不在目前的看多 / 看空選股池內。",
                    "2. 系統不提供買賣建議與判斷。",
                    "3. 仍可顯示週K線與基本資訊供參考。"
                ]

        self.set_readonly_text(self.detail_text, "\n".join(lines))

    def update_company_info_text(self, code):
        master = self.latest_result.get("MASTER_STOCK_LIST", pd.DataFrame())
        twse_all = self.latest_result.get("TWSE_ALL", pd.DataFrame())
        weekly_raw = self.latest_result.get("WEEKLY_MA_RAW", pd.DataFrame())

        master_one = master[master["股票代號"].astype(str) == str(code)].copy() if isinstance(master, pd.DataFrame) else pd.DataFrame()
        twse_one = twse_all[twse_all["股票代號"].astype(str) == str(code)].copy() if isinstance(twse_all, pd.DataFrame) else pd.DataFrame()
        raw_one = weekly_raw[weekly_raw["股票代號"].astype(str) == str(code)].copy() if isinstance(weekly_raw, pd.DataFrame) else pd.DataFrame()

        if master_one.empty and twse_one.empty and raw_one.empty:
            self.set_readonly_text(self.company_text, f"查無 {code} 的公司基本資訊。")
            return

        name = ""
        market = ""
        industry = ""
        settle_date = ""
        weekly_close = ""
        weekly_vol = ""
        ma20 = ""
        ma3 = ""
        ma5 = ""
        above20 = ""
        trend_status = ""

        if not master_one.empty:
            r = master_one.iloc[0]
            name = r.get("股票名稱", "")
            market = r.get("市場別", "")
            industry = r.get("產業別", "")

        if not twse_one.empty:
            r = twse_one.iloc[0]
            if not name:
                name = r.get("股票名稱", "")
            if not market:
                market = r.get("市場別", "")
            if not industry:
                industry = r.get("產業別", "")
            settle_date = r.get("週結算日期", "")
            weekly_close = r.get("每週五收盤後股價", "")
            weekly_vol = r.get("每週總成交量(張)", "")

        if not raw_one.empty:
            raw_one = raw_one.sort_values("週結算日期")
            r = raw_one.iloc[-1]
            ma20 = round(float(r["週20MA"]), 4) if pd.notna(r.get("週20MA")) else ""
            ma3 = round(float(r["週3MA"]), 4) if pd.notna(r.get("週3MA")) else ""
            ma5 = round(float(r["週5MA"]), 4) if pd.notna(r.get("週5MA")) else ""
            above20 = "是" if pd.notna(r.get("週20MA")) and pd.notna(r.get("週收盤價")) and r.get("週收盤價") >= r.get("週20MA") else "否"

            if pd.notna(r.get("週3MA")) and pd.notna(r.get("週5MA")) and pd.notna(r.get("週20MA")):
                if r["週3MA"] >= r["週5MA"] >= r["週20MA"]:
                    trend_status = "多方排列"
                elif r["週3MA"] <= r["週5MA"] <= r["週20MA"]:
                    trend_status = "空方排列"
                else:
                    trend_status = "整理中"

        lines = [
            f"股票代號：{code}",
            f"股票名稱：{name}",
            f"市場別：{market}",
            f"產業別：{industry}",
            "",
            f"最新週結算日：{settle_date}",
            f"最新週收盤價：{weekly_close}",
            f"最新週成交量(張)：{weekly_vol}",
            "",
            f"週3MA：{ma3}",
            f"週5MA：{ma5}",
            f"週20MA：{ma20}",
            f"是否站上20MA：{above20}",
            f"目前均線狀態：{trend_status}",
            "",
            "備註：",
            "1. 本頁文字為唯讀，不可直接修改。",
            "2. 若最新資料抓不到，系統會自動回退到最近可用結果。",
        ]

        self.set_readonly_text(self.company_text, "\n".join(lines))

    def show_alarm_summary(self):
        bull = self.latest_result.get("CLIENT_BULLISH", pd.DataFrame())
        bear = self.latest_result.get("CLIENT_BEARISH", pd.DataFrame())

        msgs = []

        if bull is not None and not bull.empty:
            short_list = bull[bull["短線停利Alarm"] == "是"]["股票代號"].tolist() if "短線停利Alarm" in bull.columns else []
            long_list = bull[bull["長線停利Alarm"] == "是"]["股票代號"].tolist() if "長線停利Alarm" in bull.columns else []

            if short_list:
                msgs.append("看多短線停利：" + "、".join(short_list[:12]))
            if long_list:
                msgs.append("看多長線停利：" + "、".join(long_list[:12]))

        if bear is not None and not bear.empty:
            short_list = bear[bear["短線回補Alarm"] == "是"]["股票代號"].tolist() if "短線回補Alarm" in bear.columns else []
            long_list = bear[bear["長線回補Alarm"] == "是"]["股票代號"].tolist() if "長線回補Alarm" in bear.columns else []

            if short_list:
                msgs.append("看空短線回補：" + "、".join(short_list[:12]))
            if long_list:
                msgs.append("看空長線回補：" + "、".join(long_list[:12]))

        if msgs:
            messagebox.showinfo("Alarm 提示", "\n".join(msgs))

    def open_excel(self):
        pass

    def open_output_folder(self):
        pass

    def start_auto_refresh(self):
        if not self.analysis_allowed:
            messagebox.showwarning("授權提醒", "目前尚未登入，或試用已到期／未訂閱，無法啟動自動刷新。")
            return
        try:
            mins = int(self.auto_var.get())
            if mins <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning("提醒", "請輸入正整數分鐘。")
            return

        self.stop_auto_refresh()
        self.log(f"已啟動自動刷新，每 {mins} 分鐘執行一次。")
        self.schedule_auto_refresh(mins)

    def schedule_auto_refresh(self, mins):
        ms = mins * 60 * 1000

        def job():
            self.run_analysis_thread()
            self.auto_refresh_job = self.after(ms, job)

        self.auto_refresh_job = self.after(ms, job)
        self.set_status(f"自動刷新中（每 {mins} 分鐘）")

    def stop_auto_refresh(self):
        if self.auto_refresh_job is not None:
            self.after_cancel(self.auto_refresh_job)
            self.auto_refresh_job = None
            self.log("已停止自動刷新。")
        self.set_status("待命中")


# =========================
# 執行
# =========================
if __name__ == "__main__":
    app = StockApp()
    app.mainloop()
